import json
import os
from timeit import default_timer as timer

import jsonpath
import openpyxl
from jproperties import Properties

from utils.Ephoc2DateTime import Ephoc2DateTime as tzconverter

configs = Properties()
ROOT_DIR1 = os.path.abspath(os.curdir)
ROOT_DIR = ROOT_DIR1.replace('writeExcel', 'configuration.properties')
with open(ROOT_DIR, 'rb') as config_file:
    configs.load(config_file)
Json_DIR = ROOT_DIR1.replace('writeExcel', 'readAPI/jsonfiles/')
start = timer()
input_jsonfile = Json_DIR + configs.get("clientName").data + "_AllSubscriptions.json"
output_file_name = ROOT_DIR1 + '/outputExcelFiles/' + configs.get("clientName").data + "_Subscription_Actual.xlsx"
clienttimezone = configs.get("clienttimezone").data

# -------------------Column Nos ----------------------#
customer_id_CNo = 1
subscription_id_CNo = 2
subscription_plan_id_0_CNo = 3
subscription_plan_quantity_CNo = 4
subscription_plan_unit_price_CNo = 5
subscription_status_CNo = 6
subscription_started_at_CNo = 7
subscription_current_term_start_CNo = 8
subscription_current_term_end_CNo = 9
subscription_cancelled_at_CNo = 10
subscription_auto_collection_CNo = 11
coupon_0_ids_CNo = 12
shipping_address_line1_CNo = 13
shipping_address_line2_CNo = 14
shipping_address_city_CNo = 15
shipping_address_state_CNo = 16
shipping_address_zip_CNo = 17
shipping_address_country_CNo = 18
subscription_setup_fee_CNo = 'NA'
subscription_start_date_CNo = 'NA'
subscription_trial_start_CNo = 'NA'
subscription_trial_end_CNo = 'NA'
subscription_pause_date_CNo = 'NA'
subscription_resume_date_CNo = 'NA'
billing_cycles_CNo = 'NA'
addons_0_id_CNo = 'NA'
addons_0_quantity_CNo = 'NA'
addons_0_unit_price_CNo = 'NA'
addons_1_id_CNo = 'NA'
addons_1_quantity_CNo = 'NA'
addons_1_unit_price_CNo = 'NA'
addons_2_id_CNo = 'NA'
addons_2_quantity_CNo = 'NA'
addons_2_unit_price_CNo = 'NA'

subscription_payment_source_id_CNo = 'NA'
subscription_invoice_notes_CNo = 'NA'
subscription_meta_data_CNo = 'NA'
shipping_address_first_name_CNo = 'NA'
shipping_address_last_name_CNo = 'NA'
shipping_address_email_CNo = 'NA'
shipping_address_company_CNo = 'NA'
shipping_address_phone_CNo = 'NA'
shipping_address_line3_CNo = 'NA'
shipping_address_state_code_CNo = 'NA'
subscription_po_number_CNo = 'NA'
customer_company_CNo = 'NA'
term_length_CNo = 'NA'
subscription_payment_frequency_CNo = 'NA'
discount_CNo = 'NA'
discount_on_CNo = 'NA'
SalesTax_CNo = 'NA'
shipping_address_validation_status_CNo = 'NA'
addons_3_id_CNo = 'NA'
addons_3_quantity_CNo = 'NA'
addons_3_unit_price_CNo = 'NA'
addons_4_id_CNo = 'NA'
addons_4_quantity_CNo = 'NA'
addons_4_unit_price_CNo = 'NA'
addons_5_id_CNo = 'NA'
addons_5_quantity_CNo = 'NA'
addons_5_unit_price_CNo = 'NA'
addons_6_id_CNo = 'NA'
addons_6_quantity_CNo = 'NA'
addons_6_unit_price_CNo = 'NA'
addons_7_id_CNo = 'NA'
addons_7_quantity_CNo = 'NA'
addons_7_unit_price_CNo = 'NA'
contract_term_created_at_CNo = 'NA'
contract_term_contract_start_CNo = 'NA'
contract_term_billing_cycle_CNo = 'NA'
contract_term_total_amount_raised_CNo = 'NA'
contract_term_action_at_term_end_CNo = 'NA'
contract_term_cancellation_cutoff_period_CNo = 'NA'
contract_term_billing_cycle_on_renewal_CNo = 'NA'
created_at_CNo = 19

jsondata = open(input_jsonfile)
Subscriptiondictionary = json.load(jsondata)
print("Final Json File", Subscriptiondictionary)
apitojsontime = timer()

jsonpathres = Subscriptiondictionary
l = jsonpath.jsonpath(jsonpathres, "list")
totalRecordCountInResp = len(l[0])
# create work book and sheet object
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "Subscriptions"

# ---------------------------write headers in excel -- first row---------------------------#

if customer_id_CNo != "NA":
    customer_id = my_sheet.cell(row=1, column=customer_id_CNo)
    customer_id.value = str("customer_id")

if customer_company_CNo != "NA":
    customer_company = my_sheet.cell(row=1, column=customer_company_CNo)
    customer_company.value = str("customer_company")

if subscription_id_CNo != "NA":
    subscription_id = my_sheet.cell(row=1, column=subscription_id_CNo)
    subscription_id.value = str("subscription_id")

if subscription_plan_id_0_CNo != "NA":
    subscription_plan_id_0 = my_sheet.cell(row=1, column=subscription_plan_id_0_CNo)
    subscription_plan_id_0.value = str("subscription_plan_id_0")

if subscription_plan_quantity_CNo != "NA":
    subscription_plan_quantity = my_sheet.cell(row=1, column=subscription_plan_quantity_CNo)
    subscription_plan_quantity.value = str("subscription_plan_quantity")

if subscription_plan_unit_price_CNo != "NA":
    subscription_plan_unit_price = my_sheet.cell(row=1, column=subscription_plan_unit_price_CNo)
    subscription_plan_unit_price.value = str("subscription_plan_unit_price")

if subscription_setup_fee_CNo != "NA":
    subscription_setup_fee = my_sheet.cell(row=1, column=subscription_setup_fee_CNo)
    subscription_setup_fee.value = str("subscription_setup_fee")

if subscription_status_CNo != "NA":
    subscription_status = my_sheet.cell(row=1, column=subscription_status_CNo)
    subscription_status.value = str("subscription_status")

if subscription_start_date_CNo != "NA":
    subscription_start_date = my_sheet.cell(row=1, column=subscription_start_date_CNo)
    subscription_start_date.value = str("subscription_start_date")

if subscription_trial_start_CNo != "NA":
    subscription_trial_start = my_sheet.cell(row=1, column=subscription_trial_start_CNo)
    subscription_trial_start.value = str("subscription_trial_start")

if subscription_trial_end_CNo != "NA":
    subscription_trial_end = my_sheet.cell(row=1, column=subscription_trial_end_CNo)
    subscription_trial_end.value = str("subscription_trial_end")

if subscription_started_at_CNo != "NA":
    subscription_started_at = my_sheet.cell(row=1, column=subscription_started_at_CNo)
    subscription_started_at.value = str("subscription_started_at")

if subscription_current_term_start_CNo != "NA":
    subscription_current_term_start = my_sheet.cell(row=1, column=subscription_current_term_start_CNo)
    subscription_current_term_start.value = str("subscription_current_term_start")

if subscription_current_term_end_CNo != "NA":
    subscription_current_term_end = my_sheet.cell(row=1, column=subscription_current_term_end_CNo)
    subscription_current_term_end.value = str("subscription_current_term_end")

if term_length_CNo != "NA":
    term_length = my_sheet.cell(row=1, column=term_length_CNo)
    term_length.value = str("term_length")

if subscription_payment_frequency_CNo != "NA":
    subscription_payment_frequency = my_sheet.cell(row=1, column=subscription_payment_frequency_CNo)
    subscription_payment_frequency.value = str("subscription_payment_frequency")

if billing_cycles_CNo != "NA":
    billing_cycles = my_sheet.cell(row=1, column=billing_cycles_CNo)
    billing_cycles.value = str("billing_cycles")

if subscription_auto_collection_CNo != "NA":
    subscription_auto_collection = my_sheet.cell(row=1, column=subscription_auto_collection_CNo)
    subscription_auto_collection.value = str("subscription_auto_collection")

if subscription_po_number_CNo != "NA":
    subscription_po_number = my_sheet.cell(row=1, column=subscription_po_number_CNo)
    subscription_po_number.value = str("subscription_po_number")

if discount_CNo != "NA":
    discount = my_sheet.cell(row=1, column=discount_CNo)
    discount.value = str("discount")

if discount_on_CNo != "NA":
    discount_on = my_sheet.cell(row=1, column=discount_on_CNo)
    discount_on.value = str("discount_on")

if SalesTax_CNo != "NA":
    SalesTax = my_sheet.cell(row=1, column=SalesTax_CNo)
    SalesTax.value = str("SalesTax")

if coupon_0_ids_CNo != "NA":
    coupon_0_ids = my_sheet.cell(row=1, column=coupon_0_ids_CNo)
    coupon_0_ids.value = str("coupon_0_ids")

if subscription_payment_source_id_CNo != "NA":
    subscription_payment_source_id = my_sheet.cell(row=1, column=subscription_payment_source_id_CNo)
    subscription_payment_source_id.value = str("subscription_payment_source_id")

if subscription_cancelled_at_CNo != "NA":
    subscription_cancelled_at = my_sheet.cell(row=1, column=subscription_cancelled_at_CNo)
    subscription_cancelled_at.value = str("subscription_cancelled_at")

if subscription_pause_date_CNo != "NA":
    subscription_pause_date = my_sheet.cell(row=1, column=subscription_pause_date_CNo)
    subscription_pause_date.value = str("subscription_pause_date")

if subscription_resume_date_CNo != "NA":
    subscription_resume_date = my_sheet.cell(row=1, column=subscription_resume_date_CNo)
    subscription_resume_date.value = str("subscription_resume_date")

if subscription_invoice_notes_CNo != "NA":
    subscription_invoice_notes = my_sheet.cell(row=1, column=subscription_invoice_notes_CNo)
    subscription_invoice_notes.value = str("subscription_invoice_notes")

if subscription_meta_data_CNo != "NA":
    subscription_meta_data = my_sheet.cell(row=1, column=subscription_meta_data_CNo)
    subscription_meta_data.value = str("subscription_meta_data")

if shipping_address_first_name_CNo != "NA":
    shipping_address_first_name = my_sheet.cell(row=1, column=shipping_address_first_name_CNo)
    shipping_address_first_name.value = str("shipping_address_first_name")

if shipping_address_last_name_CNo != "NA":
    shipping_address_last_name = my_sheet.cell(row=1, column=shipping_address_last_name_CNo)
    shipping_address_last_name.value = str("shipping_address_last_name")

if shipping_address_email_CNo != "NA":
    shipping_address_email = my_sheet.cell(row=1, column=shipping_address_email_CNo)
    shipping_address_email.value = str("shipping_address_email")

if shipping_address_company_CNo != "NA":
    shipping_address_company = my_sheet.cell(row=1, column=shipping_address_company_CNo)
    shipping_address_company.value = str("shipping_address_company")

if shipping_address_phone_CNo != "NA":
    shipping_address_phone = my_sheet.cell(row=1, column=shipping_address_phone_CNo)
    shipping_address_phone.value = str("shipping_address_phone")

if shipping_address_line1_CNo != "NA":
    shipping_address_line1 = my_sheet.cell(row=1, column=shipping_address_line1_CNo)
    shipping_address_line1.value = str("shipping_address_line1")

if shipping_address_line2_CNo != "NA":
    shipping_address_line2 = my_sheet.cell(row=1, column=shipping_address_line2_CNo)
    shipping_address_line2.value = str("shipping_address_line2")

if shipping_address_line3_CNo != "NA":
    shipping_address_line3 = my_sheet.cell(row=1, column=shipping_address_line3_CNo)
    shipping_address_line3.value = str("shipping_address_line3")

if shipping_address_city_CNo != "NA":
    shipping_address_city = my_sheet.cell(row=1, column=shipping_address_city_CNo)
    shipping_address_city.value = str("shipping_address_city")

if shipping_address_state_code_CNo != "NA":
    shipping_address_state_code = my_sheet.cell(row=1, column=shipping_address_state_code_CNo)
    shipping_address_state_code.value = str("shipping_address_state_code")

if shipping_address_state_CNo != "NA":
    shipping_address_state = my_sheet.cell(row=1, column=shipping_address_state_CNo)
    shipping_address_state.value = str("shipping_address_state")

if shipping_address_zip_CNo != "NA":
    shipping_address_zip = my_sheet.cell(row=1, column=shipping_address_zip_CNo)
    shipping_address_zip.value = str("shipping_address_zip")

if shipping_address_country_CNo != "NA":
    shipping_address_country = my_sheet.cell(row=1, column=shipping_address_country_CNo)
    shipping_address_country.value = str("shipping_address_country")

if shipping_address_validation_status_CNo != "NA":
    shipping_address_validation_status = my_sheet.cell(row=1, column=shipping_address_validation_status_CNo)
    shipping_address_validation_status.value = str("shipping_address_validation_status")

if addons_0_id_CNo != "NA":
    addons_0_id = my_sheet.cell(row=1, column=addons_0_id_CNo)
    addons_0_id.value = str("addons_0_id")

if addons_0_quantity_CNo != "NA":
    addons_0_quantity = my_sheet.cell(row=1, column=addons_0_quantity_CNo)
    addons_0_quantity.value = str("addons_0_quantity")

if addons_0_unit_price_CNo != "NA":
    addons_0_unit_price = my_sheet.cell(row=1, column=addons_0_unit_price_CNo)
    addons_0_unit_price.value = str("addons_0_unit_price")

if addons_1_id_CNo != "NA":
    addons_1_id = my_sheet.cell(row=1, column=addons_1_id_CNo)
    addons_1_id.value = str("addons_1_id")

if addons_1_quantity_CNo != "NA":
    addons_1_quantity = my_sheet.cell(row=1, column=addons_1_quantity_CNo)
    addons_1_quantity.value = str("addons_1_quantity")

if addons_1_unit_price_CNo != "NA":
    addons_1_unit_price = my_sheet.cell(row=1, column=addons_1_unit_price_CNo)
    addons_1_unit_price.value = str("addons_1_unit_price")

if addons_2_id_CNo != "NA":
    addons_2_id = my_sheet.cell(row=1, column=addons_2_id_CNo)
    addons_2_id.value = str("addons_2_id")

if addons_2_quantity_CNo != "NA":
    addons_2_quantity = my_sheet.cell(row=1, column=addons_2_quantity_CNo)
    addons_2_quantity.value = str("addons_2_quantity")

if addons_2_unit_price_CNo != "NA":
    addons_2_unit_price = my_sheet.cell(row=1, column=addons_2_unit_price_CNo)
    addons_2_unit_price.value = str("addons_2_unit_price")

if addons_3_id_CNo != "NA":
    addons_3_id = my_sheet.cell(row=1, column=addons_3_id_CNo)
    addons_3_id.value = str("addons_3_id")

if addons_3_quantity_CNo != "NA":
    addons_3_quantity = my_sheet.cell(row=1, column=addons_3_quantity_CNo)
    addons_3_quantity.value = str("addons_3_quantity")

if addons_3_unit_price_CNo != "NA":
    addons_3_unit_price = my_sheet.cell(row=1, column=addons_3_unit_price_CNo)
    addons_3_unit_price.value = str("addons_3_unit_price")

if addons_4_id_CNo != "NA":
    addons_4_id = my_sheet.cell(row=1, column=addons_4_id_CNo)
    addons_4_id.value = str("addons_4_id")

if addons_4_quantity_CNo != "NA":
    addons_4_quantity = my_sheet.cell(row=1, column=addons_4_quantity_CNo)
    addons_4_quantity.value = str("addons_4_quantity")

if addons_4_unit_price_CNo != "NA":
    addons_4_unit_price = my_sheet.cell(row=1, column=addons_4_unit_price_CNo)
    addons_4_unit_price.value = str("addons_4_unit_price")

if addons_5_id_CNo != "NA":
    addons_5_id = my_sheet.cell(row=1, column=addons_5_id_CNo)
    addons_5_id.value = str("addons_5_id")

if addons_5_quantity_CNo != "NA":
    addons_5_quantity = my_sheet.cell(row=1, column=addons_5_quantity_CNo)
    addons_5_quantity.value = str("addons_5_quantity")

if addons_5_unit_price_CNo != "NA":
    addons_5_unit_price = my_sheet.cell(row=1, column=addons_5_unit_price_CNo)
    addons_5_unit_price.value = str("addons_5_unit_price")

if addons_6_id_CNo != "NA":
    addons_6_id = my_sheet.cell(row=1, column=addons_6_id_CNo)
    addons_6_id.value = str("addons_6_id")

if addons_6_quantity_CNo != "NA":
    addons_6_quantity = my_sheet.cell(row=1, column=addons_6_quantity_CNo)
    addons_6_quantity.value = str("addons_6_quantity")

if addons_6_unit_price_CNo != "NA":
    addons_6_unit_price = my_sheet.cell(row=1, column=addons_6_unit_price_CNo)
    addons_6_unit_price.value = str("addons_6_unit_price")

if addons_7_id_CNo != "NA":
    addons_7_id = my_sheet.cell(row=1, column=addons_7_id_CNo)
    addons_7_id.value = str("addons_7_id")

if addons_7_quantity_CNo != "NA":
    addons_7_quantity = my_sheet.cell(row=1, column=addons_7_quantity_CNo)
    addons_7_quantity.value = str("addons_7_quantity")

if addons_7_unit_price_CNo != "NA":
    addons_7_unit_price = my_sheet.cell(row=1, column=addons_7_unit_price_CNo)
    addons_7_unit_price.value = str("addons_7_unit_price")

if contract_term_created_at_CNo != "NA":
    contract_term_created_at = my_sheet.cell(row=1, column=contract_term_created_at_CNo)
    contract_term_created_at.value = str("contract_term_created_at")

if contract_term_contract_start_CNo != "NA":
    contract_term_contract_start = my_sheet.cell(row=1, column=contract_term_contract_start_CNo)
    contract_term_contract_start.value = str("contract_term_contract_start")

if contract_term_billing_cycle_CNo != "NA":
    contract_term_billing_cycle = my_sheet.cell(row=1, column=contract_term_billing_cycle_CNo)
    contract_term_billing_cycle.value = str("contract_term_billing_cycle")

if contract_term_total_amount_raised_CNo != "NA":
    contract_term_total_amount_raised = my_sheet.cell(row=1, column=contract_term_total_amount_raised_CNo)
    contract_term_total_amount_raised.value = str("contract_term_total_amount_raised")

if contract_term_action_at_term_end_CNo != "NA":
    contract_term_action_at_term_end = my_sheet.cell(row=1, column=contract_term_action_at_term_end_CNo)
    contract_term_action_at_term_end.value = str("contract_term_action_at_term_end")

if contract_term_cancellation_cutoff_period_CNo != "NA":
    contract_term_cancellation_cutoff_period = my_sheet.cell(row=1, column=contract_term_cancellation_cutoff_period_CNo)
    contract_term_cancellation_cutoff_period.value = str("contract_term_cancellation_cutoff_period")

if contract_term_billing_cycle_on_renewal_CNo != "NA":
    contract_term_billing_cycle_on_renewal = my_sheet.cell(row=1, column=contract_term_billing_cycle_on_renewal_CNo)
    contract_term_billing_cycle_on_renewal.value = str("contract_term_billing_cycle_on_renewal")

if created_at_CNo != "NA":
    sub_created_at = my_sheet.cell(row=1, column=created_at_CNo)
    sub_created_at.value = str("sub_created_at")


# parse each element/field from response and return

def get_customer_id(i):
    try:
        customer_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.id")
    except:
        print("in status exception")
    return customer_id


def get_customer_company(i):
    try:
        customer_company = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.company")
    except:
        print("in status exception")
    return customer_company


def get_subscription_id(i):
    try:
        subscription_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.id")
    except:
        print("in status exception")
    return subscription_id


def get_subscription_plan_id_0(i):
    try:
        subscription_plan_id_0 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.plan_id")
    except:
        print("in status exception")
    return subscription_plan_id_0


# def get_subscription_plan_id_0(i):
#     try:
#         subscription_plan_id_0 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.subscription_items[0].item_price_id")
#     except:
#         print("in status exception")
#     return subscription_plan_id_0


def get_subscription_plan_quantity(i):
    try:
        subscription_plan_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.plan_quantity")
    except:
        print("in status exception")
    return subscription_plan_quantity


def get_subscription_plan_unit_price(i):
    try:
        subscription_plan_unit_price = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].subscription.plan_unit_price")
    except:
        print("in status exception")
    return subscription_plan_unit_price


def get_subscription_setup_fee(i):
    try:
        subscription_setup_fee = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.setup_fee")
    except:
        print("in status exception")
    return subscription_setup_fee


def get_subscription_status(i):
    try:
        subscription_status = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.status")
    except:
        print("in status exception")
    return subscription_status


def get_subscription_start_date(i):
    try:
        subscription_start_date = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.start_date")
    except:
        print("in status exception")
    return subscription_start_date


def get_subscription_trial_start(i):
    try:
        subscription_trial_start = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.trial_start")
    except:
        print("in status exception")
    return subscription_trial_start


def get_subscription_trial_end(i):
    try:
        subscription_trial_end = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.trial_end")
    except:
        print("in status exception")
    return subscription_trial_end


def get_subscription_started_at(i):
    try:
        subscription_started_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.started_at")
    except:
        print("in status exception")
    return subscription_started_at


def get_subscription_current_term_start(i):
    try:
        subscription_current_term_start = jsonpath.jsonpath(jsonpathres,
                                                            "list[" + str(i) + "].subscription.current_term_start")
    except:
        print("in status exception")
    return subscription_current_term_start


def get_subscription_current_term_end(i):
    try:
        subscription_current_term_end = jsonpath.jsonpath(jsonpathres,
                                                          "list[" + str(i) + "].subscription.current_term_end")
    except:
        print("in status exception")
    return subscription_current_term_end


def get_term_length(i):
    try:
        term_length = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.term_length")
    except:
        print("in status exception")
    return term_length


def get_subscription_payment_frequency(i):
    try:
        subscription_payment_frequency = jsonpath.jsonpath(jsonpathres,
                                                           "list[" + str(i) + "].subscription.payment_frequency")
    except:
        print("in status exception")
    return subscription_payment_frequency


def get_billing_cycles(i):
    try:
        billing_cycles = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].billing_cycles.")
    except:
        print("in status exception")
    return billing_cycles


def get_subscription_auto_collection(i):
    try:
        subscription_auto_collection = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].subscription.auto_collection")
    except:
        print("in status exception")
    return subscription_auto_collection


def get_subscription_po_number(i):
    try:
        subscription_po_number = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.po_number")
    except:
        print("in status exception")
    return subscription_po_number


def get_discount(i):
    try:
        discount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.discount")
    except:
        print("in status exception")
    return discount


def get_discount_on(i):
    try:
        discount_on = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.discount_on")
    except:
        print("in status exception")
    return discount_on


def get_SalesTax(i):
    try:
        SalesTax = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.salestax")
    except:
        print("in status exception")
    return SalesTax


def get_coupon_0_ids(i):
    try:
        coupon_0_ids = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.coupons[0].coupon_id")
    except:
        print("in status exception")
    return coupon_0_ids


def get_subscription_payment_source_id(i):
    try:
        subscription_payment_source_id = jsonpath.jsonpath(jsonpathres,
                                                           "list[" + str(i) + "].subscription.payment_source_id")
    except:
        print("in status exception")
    return subscription_payment_source_id


def get_subscription_cancelled_at(i):
    try:
        subscription_cancelled_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.cancelled_at")
    except:
        print("in status exception")
    return subscription_cancelled_at


def get_subscription_pause_date(i):
    try:
        subscription_pause_date = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.pause_date")
    except:
        print("in status exception")
    return subscription_pause_date


def get_subscription_resume_date(i):
    try:
        subscription_resume_date = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.resume_date")
    except:
        print("in status exception")
    return subscription_resume_date


def get_subscription_invoice_notes(i):
    try:
        subscription_invoice_notes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.invoice_notes")
    except:
        print("in status exception")
    return subscription_invoice_notes


def get_subscription_meta_data(i):
    try:
        subscription_meta_data = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.meta_data")
    except:
        print("in status exception")
    return subscription_meta_data


def get_shipping_address_first_name(i):
    try:
        shipping_address_first_name = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].subscription.shipping_address.first_name")
    except:
        print("in status exception")
    return shipping_address_first_name


def get_shipping_address_last_name(i):
    try:
        shipping_address_last_name = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].subscription.shipping_address.last_name")
    except:
        print("in status exception")
    return shipping_address_last_name


def get_shipping_address_email(i):
    try:
        shipping_address_email = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].subscription.shipping_address.email")
    except:
        print("in status exception")
    return shipping_address_email


def get_shipping_address_company(i):
    try:
        shipping_address_company = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].subscription.shipping_address.company")
    except:
        print("in status exception")
    return shipping_address_company


def get_shipping_address_phone(i):
    try:
        shipping_address_phone = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].subscription.shipping_address.phone")
    except:
        print("in status exception")
    return shipping_address_phone


def get_shipping_address_line1(i):
    try:
        shipping_address_line1 = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].subscription.shipping_address.line1")
    except:
        print("in status exception")
    return shipping_address_line1


def get_shipping_address_line2(i):
    try:
        shipping_address_line2 = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].subscription.shipping_address.line2")
    except:
        print("in status exception")
    return shipping_address_line2


def get_shipping_address_line3(i):
    try:
        shipping_address_line3 = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].subscription.shipping_address.line3")
    except:
        print("in status exception")
    return shipping_address_line3


def get_shipping_address_city(i):
    try:
        shipping_address_city = jsonpath.jsonpath(jsonpathres,
                                                  "list[" + str(i) + "].subscription.shipping_address.city")
    except:
        print("in status exception")
    return shipping_address_city


def get_shipping_address_state_code(i):
    try:
        shipping_address_state_code = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].subscription.shipping_address.state_code")
    except:
        print("in status exception")
    return shipping_address_state_code


def get_shipping_address_state(i):
    try:
        shipping_address_state = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].subscription.shipping_address.state")
    except:
        print("in status exception")
    return shipping_address_state


def get_shipping_address_zip(i):
    try:
        shipping_address_zip = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.shipping_address.zip")
    except:
        print("in status exception")
    return shipping_address_zip


def get_shipping_address_country(i):
    try:
        shipping_address_country = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].subscription.shipping_address.country")
    except:
        print("in status exception")
    return shipping_address_country


def get_shipping_address_validation_status(i):
    try:
        shipping_address_validation_status = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].subscription.shipping_address.validation_status")
    except:
        print("in status exception")
    return shipping_address_validation_status


def get_addons_0_id(i):
    try:
        addons_0_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[0].id")
    except:
        print("in status exception")
    return addons_0_id


def get_addons_0_quantity(i):
    try:
        addons_0_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[0].quantity")
    except:
        print("in status exception")
    return addons_0_quantity


def get_addons_0_unit_price(i):
    try:
        addons_0_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[0].unit_price")
    except:
        print("in status exception")
    return addons_0_unit_price


def get_addons_1_id(i):
    try:
        addons_1_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[1].id")
    except:
        print("in status exception")
    return addons_1_id


def get_addons_1_quantity(i):
    try:
        addons_1_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[1].quantity")
    except:
        print("in status exception")
    return addons_1_quantity


def get_addons_1_unit_price(i):
    try:
        addons_1_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[1].unit_price")
    except:
        print("in status exception")
    return addons_1_unit_price


def get_addons_2_id(i):
    try:
        addons_2_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[2].id")
    except:
        print("in status exception")
    return addons_2_id


def get_addons_2_quantity(i):
    try:
        addons_2_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[2].quantity")
    except:
        print("in status exception")
    return addons_2_quantity


def get_addons_2_unit_price(i):
    try:
        addons_2_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[2].unit_price")
    except:
        print("in status exception")
    return addons_2_unit_price


def get_addons_3_id(i):
    try:
        addons_3_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[3].id")
    except:
        print("in status exception")
    return addons_3_id


def get_addons_3_quantity(i):
    try:
        addons_3_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[3].quantity")
    except:
        print("in status exception")
    return addons_3_quantity


def get_addons_3_unit_price(i):
    try:
        addons_3_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[3].unit_price")
    except:
        print("in status exception")
    return addons_3_unit_price


def get_addons_4_id(i):
    try:
        addons_4_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[4].id")
    except:
        print("in status exception")
    return addons_4_id


def get_addons_4_quantity(i):
    try:
        addons_4_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[4].quantity")
    except:
        print("in status exception")
    return addons_4_quantity


def get_addons_4_unit_price(i):
    try:
        addons_4_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[4].unit_price")
    except:
        print("in status exception")
    return addons_4_unit_price


def get_addons_5_id(i):
    try:
        addons_5_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[5].id")
    except:
        print("in status exception")
    return addons_5_id


def get_addons_5_quantity(i):
    try:
        addons_5_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[5].quantity")
    except:
        print("in status exception")
    return addons_5_quantity


def get_addons_5_unit_price(i):
    try:
        addons_5_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[5].unit_price")
    except:
        print("in status exception")
    return addons_5_unit_price


def get_addons_6_id(i):
    try:
        addons_6_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[6].id")
    except:
        print("in status exception")
    return addons_6_id


def get_addons_6_quantity(i):
    try:
        addons_6_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[6].quantity")
    except:
        print("in status exception")
    return addons_6_quantity


def get_addons_6_unit_price(i):
    try:
        addons_6_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[6].unit_price")
    except:
        print("in status exception")
    return addons_6_unit_price


def get_addons_7_id(i):
    try:
        addons_7_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[7].id")
    except:
        print("in status exception")
    return addons_7_id


def get_addons_7_quantity(i):
    try:
        addons_7_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[7].quantity")
    except:
        print("in status exception")
    return addons_7_quantity


def get_addons_7_unit_price(i):
    try:
        addons_7_unit_price = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.addons[7].unit_price")
    except:
        print("in status exception")
    return addons_7_unit_price


def get_contract_term_created_at(i):
    try:
        contract_term_created_at = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].subscription.contract_term.created_at")
    except:
        print("in status exception")
    return contract_term_created_at


def get_contract_term_contract_start(i):
    try:
        contract_term_contract_start = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].subscription.contract_term.contract_start")
    except:
        print("in status exception")
    return contract_term_contract_start


def get_contract_term_billing_cycle(i):
    try:
        contract_term_billing_cycle = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].subscription.contract_term.billing_cycle")
    except:
        print("in status exception")
    return contract_term_billing_cycle


def get_contract_term_total_amount_raised(i):
    try:
        contract_term_total_amount_raised = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].subscription.contract_term.total_amount_raised")
    except:
        print("in status exception")
    return contract_term_total_amount_raised


def get_contract_term_action_at_term_end(i):
    try:
        contract_term_action_at_term_end = jsonpath.jsonpath(jsonpathres,
                                                             "list[" + str(i) + "].contract_term.action_at_term_end")
    except:
        print("in status exception")
    return contract_term_action_at_term_end


def get_contract_term_cancellation_cutoff_period(i):
    try:
        contract_term_cancellation_cutoff_period = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].contract_term.cancellation_cutoff_period")
    except:
        print("in status exception")
    return contract_term_cancellation_cutoff_period


def get_contract_term_billing_cycle_on_renewal(i):
    try:
        contract_term_billing_cycle_on_renewal = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].contract_term.billing_cycle_on_renewal")
    except:
        print("in status exception")
    return contract_term_billing_cycle_on_renewal


def get_sub_created_at(i):
    try:
        sub_created_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].subscription.created_at")
    except:
        print("in status exception")
    return sub_created_at



# ---------------------------loop through all records  and write to excel---------------------------#
for i in range(0, totalRecordCountInResp):
    try:

        if customer_id_CNo != "NA":
            customer_id = get_customer_id(i)
            customer_id_cell = my_sheet.cell(row=i + 2, column=customer_id_CNo)
            if customer_id == False:
                customer_id_cell.value = customer_id
            else:
                customer_id_cell.value = str(customer_id[0])

        if customer_company_CNo != "NA":
            customer_company = get_customer_company(i)
            customer_company_cell = my_sheet.cell(row=i + 2, column=customer_company_CNo)
            if customer_company == False:
                customer_company_cell.value = customer_company
            else:
                customer_company_cell.value = str(customer_company[0])

        if subscription_id_CNo != "NA":
            subscription_id = get_subscription_id(i)
            subscription_id_cell = my_sheet.cell(row=i + 2, column=subscription_id_CNo)
            if subscription_id == False:
                subscription_id_cell.value = subscription_id
            else:
                subscription_id_cell.value = str(subscription_id[0])

        if subscription_plan_id_0_CNo != "NA":
            subscription_plan_id_0 = get_subscription_plan_id_0(i)
            subscription_plan_id_0_cell = my_sheet.cell(row=i + 2, column=subscription_plan_id_0_CNo)
            if subscription_plan_id_0 == False:
                subscription_plan_id_0_cell.value = subscription_plan_id_0
            else:
                subscription_plan_id_0_cell.value = str(subscription_plan_id_0[0])

        if subscription_plan_quantity_CNo != "NA":
            subscription_plan_quantity = get_subscription_plan_quantity(i)
            subscription_plan_quantity_cell = my_sheet.cell(row=i + 2, column=subscription_plan_quantity_CNo)
            if subscription_plan_quantity == False:
                subscription_plan_quantity_cell.value = subscription_plan_quantity
            else:
                subscription_plan_quantity_cell.value = str(subscription_plan_quantity[0])

        if subscription_plan_unit_price_CNo != "NA":
            subscription_plan_unit_price = get_subscription_plan_unit_price(i)
            subscription_plan_unit_price_cell = my_sheet.cell(row=i + 2, column=subscription_plan_unit_price_CNo)
            if subscription_plan_unit_price == False:
                subscription_plan_unit_price_cell.value = subscription_plan_unit_price
            else:
                # temp = tzconverter.centToDollar(subscription_plan_unit_price[0])
                subscription_plan_unit_price_cell.value = str(subscription_plan_unit_price[0])

        if subscription_setup_fee_CNo != "NA":
            subscription_setup_fee = get_subscription_setup_fee(i)
            subscription_setup_fee_cell = my_sheet.cell(row=i + 2, column=subscription_setup_fee_CNo)
            if subscription_setup_fee == False:
                subscription_setup_fee_cell.value = subscription_setup_fee
            else:
                temp = tzconverter.centToDollar(subscription_setup_fee[0])
                subscription_setup_fee_cell.value = str(temp)

        if subscription_status_CNo != "NA":
            subscription_status = get_subscription_status(i)
            subscription_status_cell = my_sheet.cell(row=i + 2, column=subscription_status_CNo)
            if subscription_status == False:
                subscription_status_cell.value = subscription_status
            else:
                subscription_status_cell.value = str(subscription_status[0])

        if subscription_start_date_CNo != "NA":
            subscription_start_date = get_subscription_start_date(i)
            subscription_start_date_cell = my_sheet.cell(row=i + 2, column=subscription_start_date_CNo)
            if subscription_start_date == False:
                subscription_start_date_cell.value = subscription_start_date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_start_date[0], clienttimezone)

                subscription_start_date_cell.value = str(modifiedtimestamp)

        if subscription_trial_start_CNo != "NA":
            subscription_trial_start = get_subscription_trial_start(i)
            subscription_trial_start_cell = my_sheet.cell(row=i + 2, column=subscription_trial_start_CNo)
            if subscription_trial_start == False:
                subscription_trial_start_cell.value = subscription_trial_start
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_trial_start[0], clienttimezone)
                subscription_trial_start_cell.value = str(modifiedtimestamp)

        if subscription_trial_end_CNo != "NA":
            subscription_trial_end = get_subscription_trial_end(i)
            subscription_trial_end_cell = my_sheet.cell(row=i + 2, column=subscription_trial_end_CNo)
            if subscription_trial_end == False:
                subscription_trial_end_cell.value = subscription_trial_end
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_trial_end[0], clienttimezone)
                subscription_trial_end_cell.value = str(modifiedtimestamp)

        if subscription_started_at_CNo != "NA":
            subscription_started_at = get_subscription_started_at(i)
            subscription_started_at_cell = my_sheet.cell(row=i + 2, column=subscription_started_at_CNo)
            if subscription_started_at == False:
                subscription_started_at_cell.value = subscription_started_at
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_started_at[0], clienttimezone)

                subscription_started_at_cell.value = str(modifiedtimestamp)

        if subscription_current_term_start_CNo != "NA":
            subscription_current_term_start = get_subscription_current_term_start(i)
            subscription_current_term_start_cell = my_sheet.cell(row=i + 2, column=subscription_current_term_start_CNo)
            if subscription_current_term_start == False:
                subscription_current_term_start_cell.value = subscription_current_term_start
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_current_term_start[0],
                                                                          clienttimezone)
                subscription_current_term_start_cell.value = str(modifiedtimestamp)

        if subscription_current_term_end_CNo != "NA":
            subscription_current_term_end = get_subscription_current_term_end(i)
            subscription_current_term_end_cell = my_sheet.cell(row=i + 2, column=subscription_current_term_end_CNo)
            if subscription_current_term_end == False:
                subscription_current_term_end_cell.value = subscription_current_term_end
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_current_term_end[0],
                                                                          clienttimezone)
                subscription_current_term_end_cell.value = str(modifiedtimestamp)

        if term_length_CNo != "NA":
            term_length = get_term_length(i)
            term_length_cell = my_sheet.cell(row=i + 2, column=term_length_CNo)
            if term_length == False:
                term_length_cell.value = term_length
            else:
                term_length_cell.value = str(term_length[0])

        if subscription_payment_frequency_CNo != "NA":
            subscription_payment_frequency = get_subscription_payment_frequency(i)
            subscription_payment_frequency_cell = my_sheet.cell(row=i + 2, column=subscription_payment_frequency_CNo)
            if subscription_payment_frequency == False:
                subscription_payment_frequency_cell.value = subscription_payment_frequency
            else:
                subscription_payment_frequency_cell.value = str(subscription_payment_frequency[0])

        if billing_cycles_CNo != "NA":
            billing_cycles = get_billing_cycles(i)
            billing_cycles_cell = my_sheet.cell(row=i + 2, column=billing_cycles_CNo)
            if billing_cycles == False:
                billing_cycles_cell.value = billing_cycles
            else:
                billing_cycles_cell.value = str(billing_cycles[0])

        if subscription_auto_collection_CNo != "NA":
            subscription_auto_collection = get_subscription_auto_collection(i)
            subscription_auto_collection_cell = my_sheet.cell(row=i + 2, column=subscription_auto_collection_CNo)
            if subscription_auto_collection == False:
                subscription_auto_collection_cell.value = subscription_auto_collection
            else:
                subscription_auto_collection_cell.value = str(subscription_auto_collection[0])

        if subscription_po_number_CNo != "NA":
            subscription_po_number = get_subscription_po_number(i)
            subscription_po_number_cell = my_sheet.cell(row=i + 2, column=subscription_po_number_CNo)
            if subscription_po_number == False:
                subscription_po_number_cell.value = subscription_po_number
            else:
                subscription_po_number_cell.value = str(subscription_po_number[0])

        if discount_CNo != "NA":
            discount = get_discount(i)
            discount_cell = my_sheet.cell(row=i + 2, column=discount_CNo)
            if discount == False:
                discount_cell.value = discount
            else:
                discount_cell.value = str(discount[0])

        if discount_on_CNo != "NA":
            discount_on = get_discount_on(i)
            discount_on_cell = my_sheet.cell(row=i + 2, column=discount_on_CNo)
            if discount_on == False:
                discount_on_cell.value = discount_on
            else:
                discount_on_cell.value = str(discount_on[0])

        if SalesTax_CNo != "NA":
            SalesTax = get_SalesTax(i)
            SalesTax_cell = my_sheet.cell(row=i + 2, column=SalesTax_CNo)
            if SalesTax == False:
                SalesTax_cell.value = SalesTax
            else:
                SalesTax_cell.value = str(SalesTax[0])

        if coupon_0_ids_CNo != "NA":
            coupon_0_ids = get_coupon_0_ids(i)
            coupon_0_ids_cell = my_sheet.cell(row=i + 2, column=coupon_0_ids_CNo)
            if coupon_0_ids == False:
                coupon_0_ids_cell.value = coupon_0_ids
            else:
                coupon_0_ids_cell.value = str(coupon_0_ids[0])

        if subscription_payment_source_id_CNo != "NA":
            subscription_payment_source_id = get_subscription_payment_source_id(i)
            subscription_payment_source_id_cell = my_sheet.cell(row=i + 2, column=subscription_payment_source_id_CNo)
            if subscription_payment_source_id == False:
                subscription_payment_source_id_cell.value = subscription_payment_source_id
            else:
                subscription_payment_source_id_cell.value = str(subscription_payment_source_id[0])

        if subscription_cancelled_at_CNo != "NA":
            subscription_cancelled_at = get_subscription_cancelled_at(i)
            subscription_cancelled_at_cell = my_sheet.cell(row=i + 2, column=subscription_cancelled_at_CNo)
            if subscription_cancelled_at == False:
                subscription_cancelled_at_cell.value = subscription_cancelled_at
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_cancelled_at[0], clienttimezone)
                subscription_cancelled_at_cell.value = str(modifiedtimestamp)

        if subscription_pause_date_CNo != "NA":
            subscription_pause_date = get_subscription_pause_date(i)
            subscription_pause_date_cell = my_sheet.cell(row=i + 2, column=subscription_pause_date_CNo)
            if subscription_pause_date == False:
                subscription_pause_date_cell.value = subscription_pause_date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_pause_date[0], clienttimezone)
                subscription_pause_date_cell.value = str(modifiedtimestamp)

        if subscription_resume_date_CNo != "NA":
            subscription_resume_date = get_subscription_resume_date(i)
            subscription_resume_date_cell = my_sheet.cell(row=i + 2, column=subscription_resume_date_CNo)
            if subscription_resume_date == False:
                subscription_resume_date_cell.value = subscription_resume_date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(subscription_resume_date[0], clienttimezone)
                subscription_resume_date_cell.value = str(modifiedtimestamp)

        if subscription_invoice_notes_CNo != "NA":
            subscription_invoice_notes = get_subscription_invoice_notes(i)
            subscription_invoice_notes_cell = my_sheet.cell(row=i + 2, column=subscription_invoice_notes_CNo)
            if subscription_invoice_notes == False:
                subscription_invoice_notes_cell.value = subscription_invoice_notes
            else:
                subscription_invoice_notes_cell.value = str(subscription_invoice_notes[0])

        if subscription_meta_data_CNo != "NA":
            subscription_meta_data = get_subscription_meta_data(i)
            subscription_meta_data_cell = my_sheet.cell(row=i + 2, column=subscription_meta_data_CNo)
            if subscription_meta_data == False:
                subscription_meta_data_cell.value = subscription_meta_data
            else:
                subscription_meta_data_cell.value = str(subscription_meta_data[0])

        if shipping_address_first_name_CNo != "NA":
            shipping_address_first_name = get_shipping_address_first_name(i)
            shipping_address_first_name_cell = my_sheet.cell(row=i + 2, column=shipping_address_first_name_CNo)
            if shipping_address_first_name == False:
                shipping_address_first_name_cell.value = shipping_address_first_name
            else:
                shipping_address_first_name_cell.value = str(shipping_address_first_name[0])

        if shipping_address_last_name_CNo != "NA":
            shipping_address_last_name = get_shipping_address_last_name(i)
            shipping_address_last_name_cell = my_sheet.cell(row=i + 2, column=shipping_address_last_name_CNo)
            if shipping_address_last_name == False:
                shipping_address_last_name_cell.value = shipping_address_last_name
            else:
                shipping_address_last_name_cell.value = str(shipping_address_last_name[0])

        if shipping_address_email_CNo != "NA":
            shipping_address_email = get_shipping_address_email(i)
            shipping_address_email_cell = my_sheet.cell(row=i + 2, column=shipping_address_email_CNo)
            if shipping_address_email == False:
                shipping_address_email_cell.value = shipping_address_email
            else:
                emailvalue = shipping_address_email[0]
                emailvalue = emailvalue.replace("_AT_", "@")
                emailvalue = emailvalue.replace("@example.com", "")
                shipping_address_email_cell.value = str(emailvalue)

        if shipping_address_company_CNo != "NA":
            shipping_address_company = get_shipping_address_company(i)
            shipping_address_company_cell = my_sheet.cell(row=i + 2, column=shipping_address_company_CNo)
            if shipping_address_company == False:
                shipping_address_company_cell.value = shipping_address_company
            else:
                shipping_address_company_cell.value = str(shipping_address_company[0])

        if shipping_address_phone_CNo != "NA":
            shipping_address_phone = get_shipping_address_phone(i)
            shipping_address_phone_cell = my_sheet.cell(row=i + 2, column=shipping_address_phone_CNo)
            if shipping_address_phone == False:
                shipping_address_phone_cell.value = shipping_address_phone
            else:
                shipping_address_phone_cell.value = str(shipping_address_phone[0])

        if shipping_address_line1_CNo != "NA":
            shipping_address_line1 = get_shipping_address_line1(i)
            shipping_address_line1_cell = my_sheet.cell(row=i + 2, column=shipping_address_line1_CNo)
            if shipping_address_line1 == False:
                shipping_address_line1_cell.value = shipping_address_line1
            else:
                shipping_address_line1_cell.value = str(shipping_address_line1[0])

        if shipping_address_line2_CNo != "NA":
            shipping_address_line2 = get_shipping_address_line2(i)
            shipping_address_line2_cell = my_sheet.cell(row=i + 2, column=shipping_address_line2_CNo)
            if shipping_address_line2 == False:
                shipping_address_line2_cell.value = shipping_address_line2
            else:
                shipping_address_line2_cell.value = str(shipping_address_line2[0])

        if shipping_address_line3_CNo != "NA":
            shipping_address_line3 = get_shipping_address_line3(i)
            shipping_address_line3_cell = my_sheet.cell(row=i + 2, column=shipping_address_line3_CNo)
            if shipping_address_line3 == False:
                shipping_address_line3_cell.value = shipping_address_line3
            else:
                shipping_address_line3_cell.value = str(shipping_address_line3[0])

        if shipping_address_city_CNo != "NA":
            shipping_address_city = get_shipping_address_city(i)
            shipping_address_city_cell = my_sheet.cell(row=i + 2, column=shipping_address_city_CNo)
            if shipping_address_city == False:
                shipping_address_city_cell.value = shipping_address_city
            else:
                shipping_address_city_cell.value = str(shipping_address_city[0])

        if shipping_address_state_code_CNo != "NA":
            shipping_address_state_code = get_shipping_address_state_code(i)
            shipping_address_state_code_cell = my_sheet.cell(row=i + 2, column=shipping_address_state_code_CNo)
            if shipping_address_state_code == False:
                shipping_address_state_code_cell.value = shipping_address_state_code
            else:
                shipping_address_state_code_cell.value = str(shipping_address_state_code[0])

        if shipping_address_state_CNo != "NA":
            shipping_address_state = get_shipping_address_state(i)
            shipping_address_state_cell = my_sheet.cell(row=i + 2, column=shipping_address_state_CNo)
            if shipping_address_state == False:
                shipping_address_state_cell.value = shipping_address_state
            else:
                shipping_address_state_cell.value = str(shipping_address_state[0])

        if shipping_address_zip_CNo != "NA":
            shipping_address_zip = get_shipping_address_zip(i)
            shipping_address_zip_cell = my_sheet.cell(row=i + 2, column=shipping_address_zip_CNo)
            if shipping_address_zip == False:
                shipping_address_zip_cell.value = shipping_address_zip
            else:
                shipping_address_zip_cell.value = str(shipping_address_zip[0])

        if shipping_address_country_CNo != "NA":
            shipping_address_country = get_shipping_address_country(i)
            shipping_address_country_cell = my_sheet.cell(row=i + 2, column=shipping_address_country_CNo)
            if shipping_address_country == False:
                shipping_address_country_cell.value = shipping_address_country
            else:
                shipping_address_country_cell.value = str(shipping_address_country[0])

        if shipping_address_validation_status_CNo != "NA":
            shipping_address_validation_status = get_shipping_address_validation_status(i)
            shipping_address_validation_status_cell = my_sheet.cell(row=i + 2,
                                                                    column=shipping_address_validation_status_CNo)
            if shipping_address_validation_status == False:
                shipping_address_validation_status_cell.value = shipping_address_validation_status
            else:
                shipping_address_validation_status_cell.value = str(shipping_address_validation_status[0])

        if addons_0_id_CNo != "NA":
            addons_0_id = get_addons_0_id(i)
            addons_0_id_cell = my_sheet.cell(row=i + 2, column=addons_0_id_CNo)
            if addons_0_id == False:
                addons_0_id_cell.value = addons_0_id
            else:
                addons_0_id_cell.value = str(addons_0_id[0])

        if addons_0_quantity_CNo != "NA":
            addons_0_quantity = get_addons_0_quantity(i)
            addons_0_quantity_cell = my_sheet.cell(row=i + 2, column=addons_0_quantity_CNo)
            if addons_0_quantity == False:
                addons_0_quantity_cell.value = addons_0_quantity
            else:
                addons_0_quantity_cell.value = str(addons_0_quantity[0])

        if addons_0_unit_price_CNo != "NA":
            addons_0_unit_price = get_addons_0_unit_price(i)
            addons_0_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_0_unit_price_CNo)
            if addons_0_unit_price == False:
                addons_0_unit_price_cell.value = addons_0_unit_price
            else:
                temp = tzconverter.centToDollar(addons_0_unit_price[0])
                addons_0_unit_price_cell.value = str(temp)

        if addons_1_id_CNo != "NA":
            addons_1_id = get_addons_1_id(i)
            addons_1_id_cell = my_sheet.cell(row=i + 2, column=addons_1_id_CNo)
            if addons_1_id == False:
                addons_1_id_cell.value = addons_1_id
            else:
                addons_1_id_cell.value = str(addons_1_id[0])

        if addons_1_quantity_CNo != "NA":
            addons_1_quantity = get_addons_1_quantity(i)
            addons_1_quantity_cell = my_sheet.cell(row=i + 2, column=addons_1_quantity_CNo)
            if addons_1_quantity == False:
                addons_1_quantity_cell.value = addons_1_quantity
            else:
                addons_1_quantity_cell.value = str(addons_1_quantity[0])

        if addons_1_unit_price_CNo != "NA":
            addons_1_unit_price = get_addons_1_unit_price(i)
            addons_1_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_1_unit_price_CNo)
            if addons_1_unit_price == False:
                addons_1_unit_price_cell.value = addons_1_unit_price
            else:
                temp = tzconverter.centToDollar(addons_1_unit_price[0])
                addons_1_unit_price_cell.value = str(temp)

        if addons_2_id_CNo != "NA":
            addons_2_id = get_addons_2_id(i)
            addons_2_id_cell = my_sheet.cell(row=i + 2, column=addons_2_id_CNo)
            if addons_2_id == False:
                addons_2_id_cell.value = addons_2_id
            else:
                addons_2_id_cell.value = str(addons_2_id[0])

        if addons_2_quantity_CNo != "NA":
            addons_2_quantity = get_addons_2_quantity(i)
            addons_2_quantity_cell = my_sheet.cell(row=i + 2, column=addons_2_quantity_CNo)
            if addons_2_quantity == False:
                addons_2_quantity_cell.value = addons_2_quantity
            else:
                addons_2_quantity_cell.value = str(addons_2_quantity[0])

        if addons_2_unit_price_CNo != "NA":
            addons_2_unit_price = get_addons_2_unit_price(i)
            addons_2_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_2_unit_price_CNo)
            if addons_2_unit_price == False:
                addons_2_unit_price_cell.value = addons_2_unit_price
            else:
                temp = tzconverter.centToDollar(addons_2_unit_price[0])
                addons_2_unit_price_cell.value = str(temp)

        if addons_3_id_CNo != "NA":
            addons_3_id = get_addons_3_id(i)
            addons_3_id_cell = my_sheet.cell(row=i + 2, column=addons_3_id_CNo)
            if addons_3_id == False:
                addons_3_id_cell.value = addons_3_id
            else:
                addons_3_id_cell.value = str(addons_3_id[0])

        if addons_3_quantity_CNo != "NA":
            addons_3_quantity = get_addons_3_quantity(i)
            addons_3_quantity_cell = my_sheet.cell(row=i + 2, column=addons_3_quantity_CNo)
            if addons_3_quantity == False:
                addons_3_quantity_cell.value = addons_3_quantity
            else:
                addons_3_quantity_cell.value = str(addons_3_quantity[0])

        if addons_3_unit_price_CNo != "NA":
            addons_3_unit_price = get_addons_3_unit_price(i)
            addons_3_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_3_unit_price_CNo)
            if addons_3_unit_price == False:
                addons_3_unit_price_cell.value = addons_3_unit_price
            else:
                temp = tzconverter.centToDollar(addons_3_unit_price[0])
                addons_3_unit_price_cell.value = str(temp)

        if addons_4_id_CNo != "NA":
            addons_4_id = get_addons_4_id(i)
            addons_4_id_cell = my_sheet.cell(row=i + 2, column=addons_4_id_CNo)
            if addons_4_id == False:
                addons_4_id_cell.value = addons_4_id
            else:
                addons_4_id_cell.value = str(addons_4_id[0])

        if addons_4_quantity_CNo != "NA":
            addons_4_quantity = get_addons_4_quantity(i)
            addons_4_quantity_cell = my_sheet.cell(row=i + 2, column=addons_4_quantity_CNo)
            if addons_4_quantity == False:
                addons_4_quantity_cell.value = addons_4_quantity
            else:
                addons_4_quantity_cell.value = str(addons_4_quantity[0])

        if addons_4_unit_price_CNo != "NA":
            addons_4_unit_price = get_addons_4_unit_price(i)
            addons_4_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_4_unit_price_CNo)
            if addons_4_unit_price == False:
                addons_4_unit_price_cell.value = addons_4_unit_price
            else:
                temp = tzconverter.centToDollar(addons_4_unit_price[0])
                addons_4_unit_price_cell.value = str(temp)

        if addons_5_id_CNo != "NA":
            addons_5_id = get_addons_5_id(i)
            addons_5_id_cell = my_sheet.cell(row=i + 2, column=addons_5_id_CNo)
            if addons_5_id == False:
                addons_5_id_cell.value = addons_5_id
            else:
                addons_5_id_cell.value = str(addons_5_id[0])

        if addons_5_quantity_CNo != "NA":
            addons_5_quantity = get_addons_5_quantity(i)
            addons_5_quantity_cell = my_sheet.cell(row=i + 2, column=addons_5_quantity_CNo)
            if addons_5_quantity == False:
                addons_5_quantity_cell.value = addons_5_quantity
            else:
                addons_5_quantity_cell.value = str(addons_5_quantity[0])

        if addons_5_unit_price_CNo != "NA":
            addons_5_unit_price = get_addons_5_unit_price(i)
            addons_5_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_5_unit_price_CNo)
            if addons_5_unit_price == False:
                addons_5_unit_price_cell.value = addons_5_unit_price
            else:
                addons_5_unit_price_cell.value = str(addons_5_unit_price[0])

        if addons_6_id_CNo != "NA":
            addons_6_id = get_addons_6_id(i)
            addons_6_id_cell = my_sheet.cell(row=i + 2, column=addons_6_id_CNo)
            if addons_6_id == False:
                addons_6_id_cell.value = addons_6_id
            else:
                addons_6_id_cell.value = str(addons_6_id[0])

        if addons_6_quantity_CNo != "NA":
            addons_6_quantity = get_addons_6_quantity(i)
            addons_6_quantity_cell = my_sheet.cell(row=i + 2, column=addons_6_quantity_CNo)
            if addons_6_quantity == False:
                addons_6_quantity_cell.value = addons_6_quantity
            else:
                addons_6_quantity_cell.value = str(addons_6_quantity[0])

        if addons_6_unit_price_CNo != "NA":
            addons_6_unit_price = get_addons_6_unit_price(i)
            addons_6_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_6_unit_price_CNo)
            if addons_6_unit_price == False:
                addons_6_unit_price_cell.value = addons_6_unit_price
            else:
                addons_6_unit_price_cell.value = str(addons_6_unit_price[0])

        if addons_7_id_CNo != "NA":
            addons_7_id = get_addons_7_id(i)
            addons_7_id_cell = my_sheet.cell(row=i + 2, column=addons_7_id_CNo)
            if addons_7_id == False:
                addons_7_id_cell.value = addons_7_id
            else:
                addons_7_id_cell.value = str(addons_7_id[0])

        if addons_7_quantity_CNo != "NA":
            addons_7_quantity = get_addons_7_quantity(i)
            addons_7_quantity_cell = my_sheet.cell(row=i + 2, column=addons_7_quantity_CNo)
            if addons_7_quantity == False:
                addons_7_quantity_cell.value = addons_7_quantity
            else:
                addons_7_quantity_cell.value = str(addons_7_quantity[0])

        if addons_7_unit_price_CNo != "NA":
            addons_7_unit_price = get_addons_7_unit_price(i)
            addons_7_unit_price_cell = my_sheet.cell(row=i + 2, column=addons_7_unit_price_CNo)
            if addons_7_unit_price == False:
                addons_7_unit_price_cell.value = addons_7_unit_price
            else:
                addons_7_unit_price_cell.value = str(addons_7_unit_price[0])

        if contract_term_created_at_CNo != "NA":
            contract_term_created_at = get_contract_term_created_at(i)
            contract_term_created_at_cell = my_sheet.cell(row=i + 2, column=contract_term_created_at_CNo)
            if contract_term_created_at == False:
                contract_term_created_at_cell.value = contract_term_created_at
            else:
                contract_term_created_at_cell.value = str(contract_term_created_at[0])

        if contract_term_contract_start_CNo != "NA":
            contract_term_contract_start = get_contract_term_contract_start(i)
            contract_term_contract_start_cell = my_sheet.cell(row=i + 2, column=contract_term_contract_start_CNo)
            if contract_term_contract_start == False:
                contract_term_contract_start_cell.value = contract_term_contract_start
            else:
                contract_term_contract_start_cell.value = str(contract_term_contract_start[0])

        if contract_term_billing_cycle_CNo != "NA":
            contract_term_billing_cycle = get_contract_term_billing_cycle(i)
            contract_term_billing_cycle_cell = my_sheet.cell(row=i + 2, column=contract_term_billing_cycle_CNo)
            if contract_term_billing_cycle == False:
                contract_term_billing_cycle_cell.value = contract_term_billing_cycle
            else:
                contract_term_billing_cycle_cell.value = str(contract_term_billing_cycle[0])

        if contract_term_total_amount_raised_CNo != "NA":
            contract_term_total_amount_raised = get_contract_term_total_amount_raised(i)
            contract_term_total_amount_raised_cell = my_sheet.cell(row=i + 2,
                                                                   column=contract_term_total_amount_raised_CNo)
            if contract_term_total_amount_raised == False:
                contract_term_total_amount_raised_cell.value = contract_term_total_amount_raised
            else:
                contract_term_total_amount_raised_cell.value = str(contract_term_total_amount_raised[0])

        if contract_term_action_at_term_end_CNo != "NA":
            contract_term_action_at_term_end = get_contract_term_action_at_term_end(i)
            contract_term_action_at_term_end_cell = my_sheet.cell(row=i + 2,
                                                                  column=contract_term_action_at_term_end_CNo)
            if contract_term_action_at_term_end == False:
                contract_term_action_at_term_end_cell.value = contract_term_action_at_term_end
            else:
                contract_term_action_at_term_end_cell.value = str(contract_term_action_at_term_end[0])

        if contract_term_cancellation_cutoff_period_CNo != "NA":
            contract_term_cancellation_cutoff_period = get_contract_term_cancellation_cutoff_period(i)
            contract_term_cancellation_cutoff_period_cell = my_sheet.cell(row=i + 2,
                                                                          column=contract_term_cancellation_cutoff_period_CNo)
            if contract_term_cancellation_cutoff_period == False:
                contract_term_cancellation_cutoff_period_cell.value = contract_term_cancellation_cutoff_period
            else:
                contract_term_cancellation_cutoff_period_cell.value = str(contract_term_cancellation_cutoff_period[0])

        if contract_term_billing_cycle_on_renewal_CNo != "NA":
            contract_term_billing_cycle_on_renewal = get_contract_term_billing_cycle_on_renewal(i)
            contract_term_billing_cycle_on_renewal_cell = my_sheet.cell(row=i + 2,
                                                                        column=contract_term_billing_cycle_on_renewal_CNo)
            if contract_term_billing_cycle_on_renewal == False:
                contract_term_billing_cycle_on_renewal_cell.value = contract_term_billing_cycle_on_renewal
            else:
                contract_term_billing_cycle_on_renewal_cell.value = str(contract_term_billing_cycle_on_renewal[0])

        if created_at_CNo != "NA":
            sub_created_at = get_sub_created_at(i)
            sub_created_at_cell = my_sheet.cell(row=i + 2, column=created_at_CNo)
            if sub_created_at == False:
                sub_created_at_cell.value = sub_created_at
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(sub_created_at[0], clienttimezone)
                sub_created_at_cell.value = str(modifiedtimestamp)



    except Exception as e:
        print("in exception", e)

# finally save excel
my_wb.save(output_file_name)
print("Execution completed")
end = timer()
print('Total time taken: ', end - start, ' seconds')
