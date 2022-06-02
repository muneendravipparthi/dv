import json
import os
from timeit import default_timer as timer

import jsonpath
import openpyxl
from jproperties import Properties

from utils.Ephoc2DateTime import Ephoc2DateTime as tzconverter

configs = Properties()
ROOT_DIR1 = os.path.abspath(os.curdir)
ROOT_DIR = ROOT_DIR1.replace('writeExcel', 'configuration.properties')
with open(ROOT_DIR, 'rb') as config_file:
    configs.load(config_file)
Json_DIR = ROOT_DIR1.replace('writeExcel', 'readAPI/jsonfiles/')
start = timer()
input_jsonfile = Json_DIR + configs.get("clientName").data + "_AllCreditNotes.json"
output_file_name = ROOT_DIR1 + '/outputExcelFiles/' + configs.get("clientName").data + "_CreditNotes_Actual.xlsx"
clienttimezone = configs.get("clienttimezone").data

# -------------------Column Nos ----------------------#
cn_map = 1
comment = 2
credit_note_date_CNo = 3
credit_note_reason_code_CNo = 4
credit_note_reference_invoice_id_CNo = 5
credit_note_reference_stripe_invoice_id_CNo = 6
credit_note_stripe_invoice_not_required_CNo = 7
credit_note_stripe_invoice_number_not_required_CNo = 8
credit_note_total_CNo = 9
credit_note_type_CNo = 10
customer_notes_CNo = 11
credit_note_line_items_0_amount_CNo = 12
credit_note_line_items_0_date_from_CNo = 13
credit_note_line_items_0_date_to_CNo = 14
credit_note_line_items_0_description_CNo = 15
credit_note_line_items_0_quantity_CNo = 16
credit_note_line_items_0_reference_line_item_id_CNo = 17
credit_note_line_items_0_unit_amount_CNo = 18
credit_note_id_CNo = "NA"
credit_note_customer_id_CNo = "NA"
credit_note_subscription_id_CNo = "NA"
credit_note_status_CNo = "NA"
credit_note_price_type_CNo = "NA"
credit_note_exchange_rate_CNo = "NA"
credit_note_amount_allocated_CNo = "NA"
credit_note_amount_refunded_CNo = "NA"
credit_note_amount_available_CNo = "NA"
credit_note_updated_at_CNo = "NA"
credit_note_resource_version_CNo = "NA"
credit_note_deleted_CNo = "NA"
credit_note_object_CNo = "NA"
credit_note_create_reason_code_CNo = "NA"
credit_note_currency_code_CNo = "NA"
credit_note_round_off_amount_CNo = "NA"
credit_note_fractional_correction_CNo = "NA"
credit_note_base_currency_code_CNo = "NA"
credit_note_sub_total_CNo = "NA"
credit_note_line_items_CNo = "NA"
credit_note_line_items_0__CNo = "NA"
credit_note_line_items_0_id_CNo = "NA"
credit_note_line_items_0_pricing_model_CNo = "NA"
credit_note_line_items_0_is_taxed_CNo = "NA"
credit_note_line_items_0_tax_amount_CNo = "NA"
credit_note_line_items_0_object_CNo = "NA"
credit_note_line_items_0_subscription_id_CNo = "NA"
credit_note_line_items_0_customer_id_CNo = "NA"
credit_note_line_items_0_entity_type_CNo = "NA"
credit_note_line_items_0_entity_id_CNo = "NA"
credit_note_line_items_0_discount_amount_CNo = "NA"
credit_note_line_items_0_item_level_discount_amount_CNo = "NA"
credit_note_taxes_CNo = "NA"
credit_note_line_item_taxes_CNo = "NA"
credit_note_line_item_discounts_CNo = "NA"
credit_note_linked_refunds_CNo = "NA"
credit_note_allocations_CNo = "NA"
credit_note_refunded_at_CNo = "NA"
credit_note_line_items_1__CNo = "NA"
credit_note_line_items_1_id_CNo = "NA"
credit_note_line_items_1_date_from_CNo = "NA"
credit_note_line_items_1_date_to_CNo = "NA"
credit_note_line_items_1_unit_amount_CNo = "NA"
credit_note_line_items_1_quantity_CNo = "NA"
credit_note_line_items_1_amount_CNo = "NA"
credit_note_line_items_1_pricing_model_CNo = "NA"
credit_note_line_items_1_is_taxed_CNo = "NA"
credit_note_line_items_1_tax_amount_CNo = "NA"
credit_note_line_items_1_object_CNo = "NA"
credit_note_line_items_1_subscription_id_CNo = "NA"
credit_note_line_items_1_customer_id_CNo = "NA"
credit_note_line_items_1_description_CNo = "NA"
credit_note_line_items_1_entity_type_CNo = "NA"
credit_note_line_items_1_entity_id_CNo = "NA"
credit_note_line_items_1_discount_amount_CNo = "NA"
credit_note_line_items_1_item_level_discount_amount_CNo = "NA"
credit_note_allocations_0__CNo = "NA"
credit_note_allocations_0_allocated_amount_CNo = "NA"
credit_note_allocations_0_allocated_at_CNo = "NA"
credit_note_allocations_0_invoice_id_CNo = "NA"
credit_note_allocations_0_invoice_date_CNo = "NA"
credit_note_allocations_0_invoice_status_CNo = "NA"

jsondata = open(input_jsonfile)
CNdictionary = json.load(jsondata)
print("Final Json File", CNdictionary)

jsonpathres = CNdictionary
l = jsonpath.jsonpath(jsonpathres, "list")
totalRecordCountInResp = len(l[0])
# create work book and sheet object
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "CreditNotes"
# ---------------------------write headers in excel -- first row---------------------------#
if credit_note_id_CNo != "NA":
    credit_note_id = my_sheet.cell(row=1, column=credit_note_id_CNo)
    credit_note_id.value = str("credit_note_id")

if credit_note_customer_id_CNo != "NA":
    credit_note_customer_id = my_sheet.cell(row=1, column=credit_note_customer_id_CNo)
    credit_note_customer_id.value = str("credit_note_customer_id")

if credit_note_subscription_id_CNo != "NA":
    credit_note_subscription_id = my_sheet.cell(row=1, column=credit_note_subscription_id_CNo)
    credit_note_subscription_id.value = str("credit_note_subscription_id")

if credit_note_reference_invoice_id_CNo != "NA":
    credit_note_reference_invoice_id = my_sheet.cell(row=1, column=credit_note_reference_invoice_id_CNo)
    credit_note_reference_invoice_id.value = str("credit_note_reference_invoice_id")

if credit_note_type_CNo != "NA":
    credit_note_type = my_sheet.cell(row=1, column=credit_note_type_CNo)
    credit_note_type.value = str("credit_note_type")

if credit_note_reason_code_CNo != "NA":
    credit_note_reason_code = my_sheet.cell(row=1, column=credit_note_reason_code_CNo)
    credit_note_reason_code.value = str("credit_note_reason_code")

if credit_note_status_CNo != "NA":
    credit_note_status = my_sheet.cell(row=1, column=credit_note_status_CNo)
    credit_note_status.value = str("credit_note_status")

if credit_note_date_CNo != "NA":
    credit_note_date = my_sheet.cell(row=1, column=credit_note_date_CNo)
    credit_note_date.value = str("credit_note_date")

if credit_note_price_type_CNo != "NA":
    credit_note_price_type = my_sheet.cell(row=1, column=credit_note_price_type_CNo)
    credit_note_price_type.value = str("credit_note_price_type")

if credit_note_exchange_rate_CNo != "NA":
    credit_note_exchange_rate = my_sheet.cell(row=1, column=credit_note_exchange_rate_CNo)
    credit_note_exchange_rate.value = str("credit_note_exchange_rate")

if credit_note_total_CNo != "NA":
    credit_note_total = my_sheet.cell(row=1, column=credit_note_total_CNo)
    credit_note_total.value = str("credit_note_total")

if credit_note_amount_allocated_CNo != "NA":
    credit_note_amount_allocated = my_sheet.cell(row=1, column=credit_note_amount_allocated_CNo)
    credit_note_amount_allocated.value = str("credit_note_amount_allocated")

if credit_note_amount_refunded_CNo != "NA":
    credit_note_amount_refunded = my_sheet.cell(row=1, column=credit_note_amount_refunded_CNo)
    credit_note_amount_refunded.value = str("credit_note_amount_refunded")

if credit_note_amount_available_CNo != "NA":
    credit_note_amount_available = my_sheet.cell(row=1, column=credit_note_amount_available_CNo)
    credit_note_amount_available.value = str("credit_note_amount_available")

if credit_note_updated_at_CNo != "NA":
    credit_note_updated_at = my_sheet.cell(row=1, column=credit_note_updated_at_CNo)
    credit_note_updated_at.value = str("credit_note_updated_at")

if credit_note_resource_version_CNo != "NA":
    credit_note_resource_version = my_sheet.cell(row=1, column=credit_note_resource_version_CNo)
    credit_note_resource_version.value = str("credit_note_resource_version")

if credit_note_deleted_CNo != "NA":
    credit_note_deleted = my_sheet.cell(row=1, column=credit_note_deleted_CNo)
    credit_note_deleted.value = str("credit_note_deleted")

if credit_note_object_CNo != "NA":
    credit_note_object = my_sheet.cell(row=1, column=credit_note_object_CNo)
    credit_note_object.value = str("credit_note_object")

if credit_note_create_reason_code_CNo != "NA":
    credit_note_create_reason_code = my_sheet.cell(row=1, column=credit_note_create_reason_code_CNo)
    credit_note_create_reason_code.value = str("credit_note_create_reason_code")

if credit_note_currency_code_CNo != "NA":
    credit_note_currency_code = my_sheet.cell(row=1, column=credit_note_currency_code_CNo)
    credit_note_currency_code.value = str("credit_note_currency_code")

if credit_note_round_off_amount_CNo != "NA":
    credit_note_round_off_amount = my_sheet.cell(row=1, column=credit_note_round_off_amount_CNo)
    credit_note_round_off_amount.value = str("credit_note_round_off_amount")

if credit_note_fractional_correction_CNo != "NA":
    credit_note_fractional_correction = my_sheet.cell(row=1, column=credit_note_fractional_correction_CNo)
    credit_note_fractional_correction.value = str("credit_note_fractional_correction")

if credit_note_base_currency_code_CNo != "NA":
    credit_note_base_currency_code = my_sheet.cell(row=1, column=credit_note_base_currency_code_CNo)
    credit_note_base_currency_code.value = str("credit_note_base_currency_code")

if credit_note_sub_total_CNo != "NA":
    credit_note_sub_total = my_sheet.cell(row=1, column=credit_note_sub_total_CNo)
    credit_note_sub_total.value = str("credit_note_sub_total")

if credit_note_line_items_CNo != "NA":
    credit_note_line_items = my_sheet.cell(row=1, column=credit_note_line_items_CNo)
    credit_note_line_items.value = str("credit_note_line_items")

if credit_note_line_items_0__CNo != "NA":
    credit_note_line_items_0_ = my_sheet.cell(row=1, column=credit_note_line_items_0__CNo)
    credit_note_line_items_0_.value = str("credit_note_line_items_0_")

if credit_note_line_items_0_id_CNo != "NA":
    credit_note_line_items_0_id = my_sheet.cell(row=1, column=credit_note_line_items_0_id_CNo)
    credit_note_line_items_0_id.value = str("credit_note_line_items_0_id")

if credit_note_line_items_0_date_from_CNo != "NA":
    credit_note_line_items_0_date_from = my_sheet.cell(row=1, column=credit_note_line_items_0_date_from_CNo)
    credit_note_line_items_0_date_from.value = str("credit_note_line_items_0_date_from")

if credit_note_line_items_0_date_to_CNo != "NA":
    credit_note_line_items_0_date_to = my_sheet.cell(row=1, column=credit_note_line_items_0_date_to_CNo)
    credit_note_line_items_0_date_to.value = str("credit_note_line_items_0_date_to")

if credit_note_line_items_0_unit_amount_CNo != "NA":
    credit_note_line_items_0_unit_amount = my_sheet.cell(row=1, column=credit_note_line_items_0_unit_amount_CNo)
    credit_note_line_items_0_unit_amount.value = str("credit_note_line_items_0_unit_amount")

if credit_note_line_items_0_quantity_CNo != "NA":
    credit_note_line_items_0_quantity = my_sheet.cell(row=1, column=credit_note_line_items_0_quantity_CNo)
    credit_note_line_items_0_quantity.value = str("credit_note_line_items_0_quantity")

if credit_note_line_items_0_amount_CNo != "NA":
    credit_note_line_items_0_amount = my_sheet.cell(row=1, column=credit_note_line_items_0_amount_CNo)
    credit_note_line_items_0_amount.value = str("credit_note_line_items_0_amount")

if credit_note_line_items_0_pricing_model_CNo != "NA":
    credit_note_line_items_0_pricing_model = my_sheet.cell(row=1, column=credit_note_line_items_0_pricing_model_CNo)
    credit_note_line_items_0_pricing_model.value = str("credit_note_line_items_0_pricing_model")

if credit_note_line_items_0_is_taxed_CNo != "NA":
    credit_note_line_items_0_is_taxed = my_sheet.cell(row=1, column=credit_note_line_items_0_is_taxed_CNo)
    credit_note_line_items_0_is_taxed.value = str("credit_note_line_items_0_is_taxed")

if credit_note_line_items_0_tax_amount_CNo != "NA":
    credit_note_line_items_0_tax_amount = my_sheet.cell(row=1, column=credit_note_line_items_0_tax_amount_CNo)
    credit_note_line_items_0_tax_amount.value = str("credit_note_line_items_0_tax_amount")

if credit_note_line_items_0_object_CNo != "NA":
    credit_note_line_items_0_object = my_sheet.cell(row=1, column=credit_note_line_items_0_object_CNo)
    credit_note_line_items_0_object.value = str("credit_note_line_items_0_object")

if credit_note_line_items_0_subscription_id_CNo != "NA":
    credit_note_line_items_0_subscription_id = my_sheet.cell(row=1,
                                                             column=credit_note_line_items_0_subscription_id_CNo)
    credit_note_line_items_0_subscription_id.value = str("credit_note_line_items_0_subscription_id")

if credit_note_line_items_0_customer_id_CNo != "NA":
    credit_note_line_items_0_customer_id = my_sheet.cell(row=1, column=credit_note_line_items_0_customer_id_CNo)
    credit_note_line_items_0_customer_id.value = str("credit_note_line_items_0_customer_id")

if credit_note_line_items_0_description_CNo != "NA":
    credit_note_line_items_0_description = my_sheet.cell(row=1, column=credit_note_line_items_0_description_CNo)
    credit_note_line_items_0_description.value = str("credit_note_line_items_0_description")

if credit_note_line_items_0_entity_type_CNo != "NA":
    credit_note_line_items_0_entity_type = my_sheet.cell(row=1, column=credit_note_line_items_0_entity_type_CNo)
    credit_note_line_items_0_entity_type.value = str("credit_note_line_items_0_entity_type")

if credit_note_line_items_0_entity_id_CNo != "NA":
    credit_note_line_items_0_entity_id = my_sheet.cell(row=1, column=credit_note_line_items_0_entity_id_CNo)
    credit_note_line_items_0_entity_id.value = str("credit_note_line_items_0_entity_id")

if credit_note_line_items_0_discount_amount_CNo != "NA":
    credit_note_line_items_0_discount_amount = my_sheet.cell(row=1,
                                                             column=credit_note_line_items_0_discount_amount_CNo)
    credit_note_line_items_0_discount_amount.value = str("credit_note_line_items_0_discount_amount")

if credit_note_line_items_0_item_level_discount_amount_CNo != "NA":
    credit_note_line_items_0_item_level_discount_amount = my_sheet.cell(row=1,
                                                                        column=credit_note_line_items_0_item_level_discount_amount_CNo)
    credit_note_line_items_0_item_level_discount_amount.value = str(
        "credit_note_line_items_0_item_level_discount_amount")

if credit_note_taxes_CNo != "NA":
    credit_note_taxes = my_sheet.cell(row=1, column=credit_note_taxes_CNo)
    credit_note_taxes.value = str("credit_note_taxes")

if credit_note_line_item_taxes_CNo != "NA":
    credit_note_line_item_taxes = my_sheet.cell(row=1, column=credit_note_line_item_taxes_CNo)
    credit_note_line_item_taxes.value = str("credit_note_line_item_taxes")

if credit_note_line_item_discounts_CNo != "NA":
    credit_note_line_item_discounts = my_sheet.cell(row=1, column=credit_note_line_item_discounts_CNo)
    credit_note_line_item_discounts.value = str("credit_note_line_item_discounts")

if credit_note_linked_refunds_CNo != "NA":
    credit_note_linked_refunds = my_sheet.cell(row=1, column=credit_note_linked_refunds_CNo)
    credit_note_linked_refunds.value = str("credit_note_linked_refunds")

if credit_note_allocations_CNo != "NA":
    credit_note_allocations = my_sheet.cell(row=1, column=credit_note_allocations_CNo)
    credit_note_allocations.value = str("credit_note_allocations")

if credit_note_refunded_at_CNo != "NA":
    credit_note_refunded_at = my_sheet.cell(row=1, column=credit_note_refunded_at_CNo)
    credit_note_refunded_at.value = str("credit_note_refunded_at")

if credit_note_line_items_1__CNo != "NA":
    credit_note_line_items_1_ = my_sheet.cell(row=1, column=credit_note_line_items_1__CNo)
    credit_note_line_items_1_.value = str("credit_note_line_items_1_")

if credit_note_line_items_1_id_CNo != "NA":
    credit_note_line_items_1_id = my_sheet.cell(row=1, column=credit_note_line_items_1_id_CNo)
    credit_note_line_items_1_id.value = str("credit_note_line_items_1_id")

if credit_note_line_items_1_date_from_CNo != "NA":
    credit_note_line_items_1_date_from = my_sheet.cell(row=1, column=credit_note_line_items_1_date_from_CNo)
    credit_note_line_items_1_date_from.value = str("credit_note_line_items_1_date_from")

if credit_note_line_items_1_date_to_CNo != "NA":
    credit_note_line_items_1_date_to = my_sheet.cell(row=1, column=credit_note_line_items_1_date_to_CNo)
    credit_note_line_items_1_date_to.value = str("credit_note_line_items_1_date_to")

if credit_note_line_items_1_unit_amount_CNo != "NA":
    credit_note_line_items_1_unit_amount = my_sheet.cell(row=1, column=credit_note_line_items_1_unit_amount_CNo)
    credit_note_line_items_1_unit_amount.value = str("credit_note_line_items_1_unit_amount")

if credit_note_line_items_1_quantity_CNo != "NA":
    credit_note_line_items_1_quantity = my_sheet.cell(row=1, column=credit_note_line_items_1_quantity_CNo)
    credit_note_line_items_1_quantity.value = str("credit_note_line_items_1_quantity")

if credit_note_line_items_1_amount_CNo != "NA":
    credit_note_line_items_1_amount = my_sheet.cell(row=1, column=credit_note_line_items_1_amount_CNo)
    credit_note_line_items_1_amount.value = str("credit_note_line_items_1_amount")

if credit_note_line_items_1_pricing_model_CNo != "NA":
    credit_note_line_items_1_pricing_model = my_sheet.cell(row=1, column=credit_note_line_items_1_pricing_model_CNo)
    credit_note_line_items_1_pricing_model.value = str("credit_note_line_items_1_pricing_model")

if credit_note_line_items_1_is_taxed_CNo != "NA":
    credit_note_line_items_1_is_taxed = my_sheet.cell(row=1, column=credit_note_line_items_1_is_taxed_CNo)
    credit_note_line_items_1_is_taxed.value = str("credit_note_line_items_1_is_taxed")

if credit_note_line_items_1_tax_amount_CNo != "NA":
    credit_note_line_items_1_tax_amount = my_sheet.cell(row=1, column=credit_note_line_items_1_tax_amount_CNo)
    credit_note_line_items_1_tax_amount.value = str("credit_note_line_items_1_tax_amount")

if credit_note_line_items_1_object_CNo != "NA":
    credit_note_line_items_1_object = my_sheet.cell(row=1, column=credit_note_line_items_1_object_CNo)
    credit_note_line_items_1_object.value = str("credit_note_line_items_1_object")

if credit_note_line_items_1_subscription_id_CNo != "NA":
    credit_note_line_items_1_subscription_id = my_sheet.cell(row=1,
                                                             column=credit_note_line_items_1_subscription_id_CNo)
    credit_note_line_items_1_subscription_id.value = str("credit_note_line_items_1_subscription_id")

if credit_note_line_items_1_customer_id_CNo != "NA":
    credit_note_line_items_1_customer_id = my_sheet.cell(row=1, column=credit_note_line_items_1_customer_id_CNo)
    credit_note_line_items_1_customer_id.value = str("credit_note_line_items_1_customer_id")

if credit_note_line_items_1_description_CNo != "NA":
    credit_note_line_items_1_description = my_sheet.cell(row=1, column=credit_note_line_items_1_description_CNo)
    credit_note_line_items_1_description.value = str("credit_note_line_items_1_description")

if credit_note_line_items_1_entity_type_CNo != "NA":
    credit_note_line_items_1_entity_type = my_sheet.cell(row=1, column=credit_note_line_items_1_entity_type_CNo)
    credit_note_line_items_1_entity_type.value = str("credit_note_line_items_1_entity_type")

if credit_note_line_items_1_entity_id_CNo != "NA":
    credit_note_line_items_1_entity_id = my_sheet.cell(row=1, column=credit_note_line_items_1_entity_id_CNo)
    credit_note_line_items_1_entity_id.value = str("credit_note_line_items_1_entity_id")

if credit_note_line_items_1_discount_amount_CNo != "NA":
    credit_note_line_items_1_discount_amount = my_sheet.cell(row=1,
                                                             column=credit_note_line_items_1_discount_amount_CNo)
    credit_note_line_items_1_discount_amount.value = str("credit_note_line_items_1_discount_amount")

if credit_note_line_items_1_item_level_discount_amount_CNo != "NA":
    credit_note_line_items_1_item_level_discount_amount = my_sheet.cell(row=1,
                                                                        column=credit_note_line_items_1_item_level_discount_amount_CNo)
    credit_note_line_items_1_item_level_discount_amount.value = str(
        "credit_note_line_items_1_item_level_discount_amount")

if credit_note_allocations_0__CNo != "NA":
    credit_note_allocations_0_ = my_sheet.cell(row=1, column=credit_note_allocations_0__CNo)
    credit_note_allocations_0_.value = str("credit_note_allocations_0_")

if credit_note_allocations_0_allocated_amount_CNo != "NA":
    credit_note_allocations_0_allocated_amount = my_sheet.cell(row=1,
                                                               column=credit_note_allocations_0_allocated_amount_CNo)
    credit_note_allocations_0_allocated_amount.value = str("credit_note_allocations_0_allocated_amount")

if credit_note_allocations_0_allocated_at_CNo != "NA":
    credit_note_allocations_0_allocated_at = my_sheet.cell(row=1, column=credit_note_allocations_0_allocated_at_CNo)
    credit_note_allocations_0_allocated_at.value = str("credit_note_allocations_0_allocated_at")

if credit_note_allocations_0_invoice_id_CNo != "NA":
    credit_note_allocations_0_invoice_id = my_sheet.cell(row=1, column=credit_note_allocations_0_invoice_id_CNo)
    credit_note_allocations_0_invoice_id.value = str("credit_note_allocations_0_invoice_id")

if credit_note_allocations_0_invoice_date_CNo != "NA":
    credit_note_allocations_0_invoice_date = my_sheet.cell(row=1, column=credit_note_allocations_0_invoice_date_CNo)
    credit_note_allocations_0_invoice_date.value = str("credit_note_allocations_0_invoice_date")

if credit_note_allocations_0_invoice_status_CNo != "NA":
    credit_note_allocations_0_invoice_status = my_sheet.cell(row=1,
                                                             column=credit_note_allocations_0_invoice_status_CNo)
    credit_note_allocations_0_invoice_status.value = str("credit_note_allocations_0_invoice_status")



# parse each element/field from response and return

def get_credit_note_id(i):
    try:
        credit_note_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.id")
    except:
        print("in status exception")
    return credit_note_id


def get_credit_note_customer_id(i):
    try:
        credit_note_customer_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.customer_id")
    except:
        print("in status exception")
    return credit_note_customer_id


def get_credit_note_subscription_id(i):
    try:
        credit_note_subscription_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.subscription_id")
    except:
        print("in status exception")
    return credit_note_subscription_id


def get_credit_note_reference_invoice_id(i):
    try:
        credit_note_reference_invoice_id = jsonpath.jsonpath(jsonpathres,
                                                             "list[" + str(i) + "].credit_note.reference_invoice_id")
    except:
        print("in status exception")
    return credit_note_reference_invoice_id


def get_credit_note_type(i):
    try:
        credit_note_type = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.type")
    except:
        print("in status exception")
    return credit_note_type


def get_credit_note_reason_code(i):
    try:
        credit_note_reason_code = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.reason_code")
    except:
        print("in status exception")
    return credit_note_reason_code


def get_credit_note_status(i):
    try:
        credit_note_status = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.status")
    except:
        print("in status exception")
    return credit_note_status


def get_credit_note_date(i):
    try:
        credit_note_date = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.date")
    except:
        print("in status exception")
    return credit_note_date


def get_credit_note_price_type(i):
    try:
        credit_note_price_type = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.price_type")
    except:
        print("in status exception")
    return credit_note_price_type


def get_credit_note_exchange_rate(i):
    try:
        credit_note_exchange_rate = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.exchange_rate")
    except:
        print("in status exception")
    return credit_note_exchange_rate


def get_credit_note_total(i):
    try:
        credit_note_total = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.total")
    except:
        print("in status exception")
    return credit_note_total


def get_credit_note_amount_allocated(i):
    try:
        credit_note_amount_allocated = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].credit_note.amount_allocated")
    except:
        print("in status exception")
    return credit_note_amount_allocated


def get_credit_note_amount_refunded(i):
    try:
        credit_note_amount_refunded = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.amount_refunded")
    except:
        print("in status exception")
    return credit_note_amount_refunded


def get_credit_note_amount_available(i):
    try:
        credit_note_amount_available = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].credit_note.amount_available")
    except:
        print("in status exception")
    return credit_note_amount_available


def get_credit_note_updated_at(i):
    try:
        credit_note_updated_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.updated_at")
    except:
        print("in status exception")
    return credit_note_updated_at


def get_credit_note_resource_version(i):
    try:
        credit_note_resource_version = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].credit_note.resource_version")
    except:
        print("in status exception")
    return credit_note_resource_version


def get_credit_note_deleted(i):
    try:
        credit_note_deleted = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.deleted")
    except:
        print("in status exception")
    return credit_note_deleted


def get_credit_note_object(i):
    try:
        credit_note_object = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.object")
    except:
        print("in status exception")
    return credit_note_object


def get_credit_note_create_reason_code(i):
    try:
        credit_note_create_reason_code = jsonpath.jsonpath(jsonpathres,
                                                           "list[" + str(i) + "].credit_note.create_reason_code")
    except:
        print("in status exception")
    return credit_note_create_reason_code


def get_credit_note_currency_code(i):
    try:
        credit_note_currency_code = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.currency_code")
    except:
        print("in status exception")
    return credit_note_currency_code


def get_credit_note_round_off_amount(i):
    try:
        credit_note_round_off_amount = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].credit_note.round_off_amount")
    except:
        print("in status exception")
    return credit_note_round_off_amount


def get_credit_note_fractional_correction(i):
    try:
        credit_note_fractional_correction = jsonpath.jsonpath(jsonpathres,
                                                              "list[" + str(i) + "].credit_note.fractional_correction")
    except:
        print("in status exception")
    return credit_note_fractional_correction


def get_credit_note_base_currency_code(i):
    try:
        credit_note_base_currency_code = jsonpath.jsonpath(jsonpathres,
                                                           "list[" + str(i) + "].credit_note.base_currency_code")
    except:
        print("in status exception")
    return credit_note_base_currency_code


def get_credit_note_sub_total(i):
    try:
        credit_note_sub_total = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.sub_total")
    except:
        print("in status exception")
    return credit_note_sub_total


def get_credit_note_line_items(i):
    try:
        credit_note_line_items = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.line_items")
    except:
        print("in status exception")
    return credit_note_line_items


def get_credit_note_line_items_0_(i):
    try:
        credit_note_line_items_0_ = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.line_items[0]")
    except:
        print("in status exception")
    return credit_note_line_items_0_


def get_credit_note_line_items_0_id(i):
    try:
        credit_note_line_items_0_id = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].credit_note.line_items[0].id")
    except:
        print("in status exception")
    return credit_note_line_items_0_id


def get_credit_note_line_items_0_date_from(i):
    try:
        credit_note_line_items_0_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].date_from")
    except:
        print("in status exception")
    return credit_note_line_items_0_date_from


def get_credit_note_line_items_0_date_to(i):
    try:
        credit_note_line_items_0_date_to = jsonpath.jsonpath(jsonpathres,
                                                             "list[" + str(i) + "].credit_note.line_items[0].date_to")
    except:
        print("in status exception")
    return credit_note_line_items_0_date_to


def get_credit_note_line_items_0_unit_amount(i):
    try:
        credit_note_line_items_0_unit_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].unit_amount")
    except:
        print("in status exception")
    return credit_note_line_items_0_unit_amount


def get_credit_note_line_items_0_quantity(i):
    try:
        credit_note_line_items_0_quantity = jsonpath.jsonpath(jsonpathres,
                                                              "list[" + str(i) + "].credit_note.line_items[0].quantity")
    except:
        print("in status exception")
    return credit_note_line_items_0_quantity


def get_credit_note_line_items_0_amount(i):
    try:
        credit_note_line_items_0_amount = jsonpath.jsonpath(jsonpathres,
                                                            "list[" + str(i) + "].credit_note.line_items[0].amount")
    except:
        print("in status exception")
    return credit_note_line_items_0_amount


def get_credit_note_line_items_0_pricing_model(i):
    try:
        credit_note_line_items_0_pricing_model = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].pricing_model")
    except:
        print("in status exception")
    return credit_note_line_items_0_pricing_model


def get_credit_note_line_items_0_is_taxed(i):
    try:
        credit_note_line_items_0_is_taxed = jsonpath.jsonpath(jsonpathres,
                                                              "list[" + str(i) + "].credit_note.line_items[0].is_taxed")
    except:
        print("in status exception")
    return credit_note_line_items_0_is_taxed


def get_credit_note_line_items_0_tax_amount(i):
    try:
        credit_note_line_items_0_tax_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].tax_amount")
    except:
        print("in status exception")
    return credit_note_line_items_0_tax_amount


def get_credit_note_line_items_0_object(i):
    try:
        credit_note_line_items_0_object = jsonpath.jsonpath(jsonpathres,
                                                            "list[" + str(i) + "].credit_note.line_items[0].object")
    except:
        print("in status exception")
    return credit_note_line_items_0_object


def get_credit_note_line_items_0_subscription_id(i):
    try:
        credit_note_line_items_0_subscription_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].subscription_id")
    except:
        print("in status exception")
    return credit_note_line_items_0_subscription_id


def get_credit_note_line_items_0_customer_id(i):
    try:
        credit_note_line_items_0_customer_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].customer_id")
    except:
        print("in status exception")
    return credit_note_line_items_0_customer_id


def get_credit_note_line_items_0_description(i):
    try:
        credit_note_line_items_0_description = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].description")
    except:
        print("in status exception")
    return credit_note_line_items_0_description


def get_credit_note_line_items_0_entity_type(i):
    try:
        credit_note_line_items_0_entity_type = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].entity_type")
    except:
        print("in status exception")
    return credit_note_line_items_0_entity_type


def get_credit_note_line_items_0_entity_id(i):
    try:
        credit_note_line_items_0_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].entity_id")
    except:
        print("in status exception")
    return credit_note_line_items_0_entity_id


def get_credit_note_line_items_0_discount_amount(i):
    try:
        credit_note_line_items_0_discount_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].discount_amount")
    except:
        print("in status exception")
    return credit_note_line_items_0_discount_amount


def get_credit_note_line_items_0_item_level_discount_amount(i):
    try:
        credit_note_line_items_0_item_level_discount_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[0].item_level_discount_amount")
    except:
        print("in status exception")
    return credit_note_line_items_0_item_level_discount_amount


def get_credit_note_taxes(i):
    try:
        credit_note_taxes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.taxes")
    except:
        print("in status exception")
    return credit_note_taxes


def get_credit_note_line_item_taxes(i):
    try:
        credit_note_line_item_taxes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.line_item_taxes")
    except:
        print("in status exception")
    return credit_note_line_item_taxes


def get_credit_note_line_item_discounts(i):
    try:
        credit_note_line_item_discounts = jsonpath.jsonpath(jsonpathres,
                                                            "list[" + str(i) + "].credit_note.line_item_discounts")
    except:
        print("in status exception")
    return credit_note_line_item_discounts


def get_credit_note_linked_refunds(i):
    try:
        credit_note_linked_refunds = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.linked_refunds")
    except:
        print("in status exception")
    return credit_note_linked_refunds


def get_credit_note_allocations(i):
    try:
        credit_note_allocations = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.allocations")
    except:
        print("in status exception")
    return credit_note_allocations


def get_credit_note_refunded_at(i):
    try:
        credit_note_refunded_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.refunded_at")
    except:
        print("in status exception")
    return credit_note_refunded_at


def get_credit_note_line_items_1_(i):
    try:
        credit_note_line_items_1_ = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.line_items[1]")
    except:
        print("in status exception")
    return credit_note_line_items_1_


def get_credit_note_line_items_1_id(i):
    try:
        credit_note_line_items_1_id = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].credit_note.line_items[1].id")
    except:
        print("in status exception")
    return credit_note_line_items_1_id


def get_credit_note_line_items_1_date_from(i):
    try:
        credit_note_line_items_1_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].date_from")
    except:
        print("in status exception")
    return credit_note_line_items_1_date_from


def get_credit_note_line_items_1_date_to(i):
    try:
        credit_note_line_items_1_date_to = jsonpath.jsonpath(jsonpathres,
                                                             "list[" + str(i) + "].credit_note.line_items[1].date_to")
    except:
        print("in status exception")
    return credit_note_line_items_1_date_to


def get_credit_note_line_items_1_unit_amount(i):
    try:
        credit_note_line_items_1_unit_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].unit_amount")
    except:
        print("in status exception")
    return credit_note_line_items_1_unit_amount


def get_credit_note_line_items_1_quantity(i):
    try:
        credit_note_line_items_1_quantity = jsonpath.jsonpath(jsonpathres,
                                                              "list[" + str(i) + "].credit_note.line_items[1].quantity")
    except:
        print("in status exception")
    return credit_note_line_items_1_quantity


def get_credit_note_line_items_1_amount(i):
    try:
        credit_note_line_items_1_amount = jsonpath.jsonpath(jsonpathres,
                                                            "list[" + str(i) + "].credit_note.line_items[1].amount")
    except:
        print("in status exception")
    return credit_note_line_items_1_amount


def get_credit_note_line_items_1_pricing_model(i):
    try:
        credit_note_line_items_1_pricing_model = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].pricing_model")
    except:
        print("in status exception")
    return credit_note_line_items_1_pricing_model


def get_credit_note_line_items_1_is_taxed(i):
    try:
        credit_note_line_items_1_is_taxed = jsonpath.jsonpath(jsonpathres,
                                                              "list[" + str(i) + "].credit_note.line_items[1].is_taxed")
    except:
        print("in status exception")
    return credit_note_line_items_1_is_taxed


def get_credit_note_line_items_1_tax_amount(i):
    try:
        credit_note_line_items_1_tax_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].tax_amount")
    except:
        print("in status exception")
    return credit_note_line_items_1_tax_amount


def get_credit_note_line_items_1_object(i):
    try:
        credit_note_line_items_1_object = jsonpath.jsonpath(jsonpathres,
                                                            "list[" + str(i) + "].credit_note.line_items[1].object")
    except:
        print("in status exception")
    return credit_note_line_items_1_object


def get_credit_note_line_items_1_subscription_id(i):
    try:
        credit_note_line_items_1_subscription_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].subscription_id")
    except:
        print("in status exception")
    return credit_note_line_items_1_subscription_id


def get_credit_note_line_items_1_customer_id(i):
    try:
        credit_note_line_items_1_customer_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].customer_id")
    except:
        print("in status exception")
    return credit_note_line_items_1_customer_id


def get_credit_note_line_items_1_description(i):
    try:
        credit_note_line_items_1_description = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].description")
    except:
        print("in status exception")
    return credit_note_line_items_1_description


def get_credit_note_line_items_1_entity_type(i):
    try:
        credit_note_line_items_1_entity_type = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].entity_type")
    except:
        print("in status exception")
    return credit_note_line_items_1_entity_type


def get_credit_note_line_items_1_entity_id(i):
    try:
        credit_note_line_items_1_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].entity_id")
    except:
        print("in status exception")
    return credit_note_line_items_1_entity_id


def get_credit_note_line_items_1_discount_amount(i):
    try:
        credit_note_line_items_1_discount_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].discount_amount")
    except:
        print("in status exception")
    return credit_note_line_items_1_discount_amount


def get_credit_note_line_items_1_item_level_discount_amount(i):
    try:
        credit_note_line_items_1_item_level_discount_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.line_items[1].item_level_discount_amount")
    except:
        print("in status exception")
    return credit_note_line_items_1_item_level_discount_amount


def get_credit_note_allocations_0_(i):
    try:
        credit_note_allocations_0_ = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].credit_note.allocations[0]")
    except:
        print("in status exception")
    return credit_note_allocations_0_


def get_credit_note_allocations_0_allocated_amount(i):
    try:
        credit_note_allocations_0_allocated_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.allocations[0].allocated_amount")
    except:
        print("in status exception")
    return credit_note_allocations_0_allocated_amount


def get_credit_note_allocations_0_allocated_at(i):
    try:
        credit_note_allocations_0_allocated_at = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.allocations[0].allocated_at")
    except:
        print("in status exception")
    return credit_note_allocations_0_allocated_at


def get_credit_note_allocations_0_invoice_id(i):
    try:
        credit_note_allocations_0_invoice_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.allocations[0].invoice_id")
    except:
        print("in status exception")
    return credit_note_allocations_0_invoice_id


def get_credit_note_allocations_0_invoice_date(i):
    try:
        credit_note_allocations_0_invoice_date = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.allocations[0].invoice_date")
    except:
        print("in status exception")
    return credit_note_allocations_0_invoice_date


def get_credit_note_allocations_0_invoice_status(i):
    try:
        credit_note_allocations_0_invoice_status = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].credit_note.allocations[0].invoice_status")
    except:
        print("in status exception")
    return credit_note_allocations_0_invoice_status


# ---------------------------loop through all records  and write to excel---------------------------#
for i in range(0, totalRecordCountInResp):
    try:

        if credit_note_id_CNo != 'NA':
            credit_note_id = get_credit_note_id(i)
            credit_note_id_cell = my_sheet.cell(row=i + 2, column=credit_note_id_CNo)
            if credit_note_id == False:
                credit_note_id_cell.value = credit_note_id
            else:
                credit_note_id_cell.value = str(credit_note_id[0])

        if credit_note_customer_id_CNo != 'NA':
            credit_note_customer_id = get_credit_note_customer_id(i)
            credit_note_customer_id_cell = my_sheet.cell(row=i + 2, column=credit_note_customer_id_CNo)
            if credit_note_customer_id == False:
                credit_note_customer_id_cell.value = credit_note_customer_id
            else:
                credit_note_customer_id_cell.value = str(credit_note_customer_id[0])

        if credit_note_subscription_id_CNo != 'NA':
            credit_note_subscription_id = get_credit_note_subscription_id(i)
            credit_note_subscription_id_cell = my_sheet.cell(row=i + 2, column=credit_note_subscription_id_CNo)
            if credit_note_subscription_id == False:
                credit_note_subscription_id_cell.value = credit_note_subscription_id
            else:
                credit_note_subscription_id_cell.value = str(credit_note_subscription_id[0])

        if credit_note_reference_invoice_id_CNo != 'NA':
            credit_note_reference_invoice_id = get_credit_note_reference_invoice_id(i)
            credit_note_reference_invoice_id_cell = my_sheet.cell(row=i + 2,
                                                                  column=credit_note_reference_invoice_id_CNo)
            if credit_note_reference_invoice_id == False:
                credit_note_reference_invoice_id_cell.value = credit_note_reference_invoice_id
            else:
                credit_note_reference_invoice_id_cell.value = str(credit_note_reference_invoice_id[0])

        if credit_note_type_CNo != 'NA':
            credit_note_type = get_credit_note_type(i)
            credit_note_type_cell = my_sheet.cell(row=i + 2, column=credit_note_type_CNo)
            if credit_note_type == False:
                credit_note_type_cell.value = credit_note_type
            else:
                credit_note_type_cell.value = str(credit_note_type[0])

        if credit_note_reason_code_CNo != 'NA':
            credit_note_reason_code = get_credit_note_reason_code(i)
            credit_note_reason_code_cell = my_sheet.cell(row=i + 2, column=credit_note_reason_code_CNo)
            if credit_note_reason_code == False:
                credit_note_reason_code_cell.value = credit_note_reason_code
            else:
                credit_note_reason_code_cell.value = str(credit_note_reason_code[0])

        if credit_note_status_CNo != 'NA':
            credit_note_status = get_credit_note_status(i)
            credit_note_status_cell = my_sheet.cell(row=i + 2, column=credit_note_status_CNo)
            if credit_note_status == False:
                credit_note_status_cell.value = credit_note_status
            else:
                credit_note_status_cell.value = str(credit_note_status[0])

        if credit_note_date_CNo != 'NA':
            credit_note_date = get_credit_note_date(i)
            credit_note_date_cell = my_sheet.cell(row=i + 2, column=credit_note_date_CNo)
            if credit_note_date == False:
                credit_note_date_cell.value = credit_note_date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(credit_note_date[0], clienttimezone)
                credit_note_date_cell.value = str(modifiedtimestamp)


        if credit_note_price_type_CNo != 'NA':
            credit_note_price_type = get_credit_note_price_type(i)
            credit_note_price_type_cell = my_sheet.cell(row=i + 2, column=credit_note_price_type_CNo)
            if credit_note_price_type == False:
                credit_note_price_type_cell.value = credit_note_price_type
            else:
                credit_note_price_type_cell.value = str(credit_note_price_type[0])

        if credit_note_exchange_rate_CNo != 'NA':
            credit_note_exchange_rate = get_credit_note_exchange_rate(i)
            credit_note_exchange_rate_cell = my_sheet.cell(row=i + 2, column=credit_note_exchange_rate_CNo)
            if credit_note_exchange_rate == False:
                credit_note_exchange_rate_cell.value = credit_note_exchange_rate
            else:
                credit_note_exchange_rate_cell.value = str(credit_note_exchange_rate[0])

        if credit_note_total_CNo != 'NA':
            credit_note_total = get_credit_note_total(i)
            credit_note_total_cell = my_sheet.cell(row=i + 2, column=credit_note_total_CNo)
            if credit_note_total == False:
                credit_note_total_cell.value = credit_note_total
            else:
                credit_note_total_cell.value = str(credit_note_total[0])

        if credit_note_amount_allocated_CNo != 'NA':
            credit_note_amount_allocated = get_credit_note_amount_allocated(i)
            credit_note_amount_allocated_cell = my_sheet.cell(row=i + 2, column=credit_note_amount_allocated_CNo)
            if credit_note_amount_allocated == False:
                credit_note_amount_allocated_cell.value = credit_note_amount_allocated
            else:
                credit_note_amount_allocated_cell.value = str(credit_note_amount_allocated[0])

        if credit_note_amount_refunded_CNo != 'NA':
            credit_note_amount_refunded = get_credit_note_amount_refunded(i)
            credit_note_amount_refunded_cell = my_sheet.cell(row=i + 2, column=credit_note_amount_refunded_CNo)
            if credit_note_amount_refunded == False:
                credit_note_amount_refunded_cell.value = credit_note_amount_refunded
            else:
                credit_note_amount_refunded_cell.value = str(credit_note_amount_refunded[0])

        if credit_note_amount_available_CNo != 'NA':
            credit_note_amount_available = get_credit_note_amount_available(i)
            credit_note_amount_available_cell = my_sheet.cell(row=i + 2, column=credit_note_amount_available_CNo)
            if credit_note_amount_available == False:
                credit_note_amount_available_cell.value = credit_note_amount_available
            else:
                credit_note_amount_available_cell.value = str(credit_note_amount_available[0])

        if credit_note_updated_at_CNo != 'NA':
            credit_note_updated_at = get_credit_note_updated_at(i)
            credit_note_updated_at_cell = my_sheet.cell(row=i + 2, column=credit_note_updated_at_CNo)
            if credit_note_updated_at == False:
                credit_note_updated_at_cell.value = credit_note_updated_at
            else:
                credit_note_updated_at_cell.value = str(credit_note_updated_at[0])

        if credit_note_resource_version_CNo != 'NA':
            credit_note_resource_version = get_credit_note_resource_version(i)
            credit_note_resource_version_cell = my_sheet.cell(row=i + 2, column=credit_note_resource_version_CNo)
            if credit_note_resource_version == False:
                credit_note_resource_version_cell.value = credit_note_resource_version
            else:
                credit_note_resource_version_cell.value = str(credit_note_resource_version[0])

        if credit_note_deleted_CNo != 'NA':
            credit_note_deleted = get_credit_note_deleted(i)
            credit_note_deleted_cell = my_sheet.cell(row=i + 2, column=credit_note_deleted_CNo)
            if credit_note_deleted == False:
                credit_note_deleted_cell.value = credit_note_deleted
            else:
                credit_note_deleted_cell.value = str(credit_note_deleted[0])

        if credit_note_object_CNo != 'NA':
            credit_note_object = get_credit_note_object(i)
            credit_note_object_cell = my_sheet.cell(row=i + 2, column=credit_note_object_CNo)
            if credit_note_object == False:
                credit_note_object_cell.value = credit_note_object
            else:
                credit_note_object_cell.value = str(credit_note_object[0])

        if credit_note_create_reason_code_CNo != 'NA':
            credit_note_create_reason_code = get_credit_note_create_reason_code(i)
            credit_note_create_reason_code_cell = my_sheet.cell(row=i + 2, column=credit_note_create_reason_code_CNo)
            if credit_note_create_reason_code == False:
                credit_note_create_reason_code_cell.value = credit_note_create_reason_code
            else:
                credit_note_create_reason_code_cell.value = str(credit_note_create_reason_code[0])

        if credit_note_currency_code_CNo != 'NA':
            credit_note_currency_code = get_credit_note_currency_code(i)
            credit_note_currency_code_cell = my_sheet.cell(row=i + 2, column=credit_note_currency_code_CNo)
            if credit_note_currency_code == False:
                credit_note_currency_code_cell.value = credit_note_currency_code
            else:
                credit_note_currency_code_cell.value = str(credit_note_currency_code[0])

        if credit_note_round_off_amount_CNo != 'NA':
            credit_note_round_off_amount = get_credit_note_round_off_amount(i)
            credit_note_round_off_amount_cell = my_sheet.cell(row=i + 2, column=credit_note_round_off_amount_CNo)
            if credit_note_round_off_amount == False:
                credit_note_round_off_amount_cell.value = credit_note_round_off_amount
            else:
                credit_note_round_off_amount_cell.value = str(credit_note_round_off_amount[0])

        if credit_note_fractional_correction_CNo != 'NA':
            credit_note_fractional_correction = get_credit_note_fractional_correction(i)
            credit_note_fractional_correction_cell = my_sheet.cell(row=i + 2,
                                                                   column=credit_note_fractional_correction_CNo)
            if credit_note_fractional_correction == False:
                credit_note_fractional_correction_cell.value = credit_note_fractional_correction
            else:
                credit_note_fractional_correction_cell.value = str(credit_note_fractional_correction[0])

        if credit_note_base_currency_code_CNo != 'NA':
            credit_note_base_currency_code = get_credit_note_base_currency_code(i)
            credit_note_base_currency_code_cell = my_sheet.cell(row=i + 2, column=credit_note_base_currency_code_CNo)
            if credit_note_base_currency_code == False:
                credit_note_base_currency_code_cell.value = credit_note_base_currency_code
            else:
                credit_note_base_currency_code_cell.value = str(credit_note_base_currency_code[0])

        if credit_note_sub_total_CNo != 'NA':
            credit_note_sub_total = get_credit_note_sub_total(i)
            credit_note_sub_total_cell = my_sheet.cell(row=i + 2, column=credit_note_sub_total_CNo)
            if credit_note_sub_total == False:
                credit_note_sub_total_cell.value = credit_note_sub_total
            else:
                credit_note_sub_total_cell.value = str(credit_note_sub_total[0])

        if credit_note_line_items_CNo != 'NA':
            credit_note_line_items = get_credit_note_line_items(i)
            credit_note_line_items_cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_CNo)
            if credit_note_line_items == False:
                credit_note_line_items_cell.value = credit_note_line_items
            else:
                credit_note_line_items_cell.value = str(credit_note_line_items[0])

        if credit_note_line_items_0__CNo != 'NA':
            credit_note_line_items_0_ = get_credit_note_line_items_0_(i)
            credit_note_line_items_0__cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_0__CNo)
            if credit_note_line_items_0_ == False:
                credit_note_line_items_0__cell.value = credit_note_line_items_0_
            else:
                credit_note_line_items_0__cell.value = str(credit_note_line_items_0_[0])

        if credit_note_line_items_0_id_CNo != 'NA':
            credit_note_line_items_0_id = get_credit_note_line_items_0_id(i)
            credit_note_line_items_0_id_cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_0_id_CNo)
            if credit_note_line_items_0_id == False:
                credit_note_line_items_0_id_cell.value = credit_note_line_items_0_id
            else:
                credit_note_line_items_0_id_cell.value = str(credit_note_line_items_0_id[0])

        if credit_note_line_items_0_date_from_CNo != 'NA':
            credit_note_line_items_0_date_from = get_credit_note_line_items_0_date_from(i)
            credit_note_line_items_0_date_from_cell = my_sheet.cell(row=i + 2,
                                                                    column=credit_note_line_items_0_date_from_CNo)
            if credit_note_line_items_0_date_from == False:
                credit_note_line_items_0_date_from_cell.value = credit_note_line_items_0_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(credit_note_line_items_0_date_from[0],
                                                                          clienttimezone)
                credit_note_line_items_0_date_from_cell.value = str(modifiedtimestamp)


        if credit_note_line_items_0_date_to_CNo != 'NA':
            credit_note_line_items_0_date_to = get_credit_note_line_items_0_date_to(i)
            credit_note_line_items_0_date_to_cell = my_sheet.cell(row=i + 2,
                                                                  column=credit_note_line_items_0_date_to_CNo)
            if credit_note_line_items_0_date_to == False:
                credit_note_line_items_0_date_to_cell.value = credit_note_line_items_0_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(credit_note_line_items_0_date_to[0],
                                                                          clienttimezone)
                credit_note_line_items_0_date_to_cell.value = str(modifiedtimestamp)


        if credit_note_line_items_0_unit_amount_CNo != 'NA':
            credit_note_line_items_0_unit_amount = get_credit_note_line_items_0_unit_amount(i)
            credit_note_line_items_0_unit_amount_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_0_unit_amount_CNo)
            if credit_note_line_items_0_unit_amount == False:
                credit_note_line_items_0_unit_amount_cell.value = credit_note_line_items_0_unit_amount
            else:
                credit_note_line_items_0_unit_amount_cell.value = str(credit_note_line_items_0_unit_amount[0])

        if credit_note_line_items_0_quantity_CNo != 'NA':
            credit_note_line_items_0_quantity = get_credit_note_line_items_0_quantity(i)
            credit_note_line_items_0_quantity_cell = my_sheet.cell(row=i + 2,
                                                                   column=credit_note_line_items_0_quantity_CNo)
            if credit_note_line_items_0_quantity == False:
                credit_note_line_items_0_quantity_cell.value = credit_note_line_items_0_quantity
            else:
                credit_note_line_items_0_quantity_cell.value = str(credit_note_line_items_0_quantity[0])

        if credit_note_line_items_0_amount_CNo != 'NA':
            credit_note_line_items_0_amount = get_credit_note_line_items_0_amount(i)
            credit_note_line_items_0_amount_cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_0_amount_CNo)
            if credit_note_line_items_0_amount == False:
                credit_note_line_items_0_amount_cell.value = credit_note_line_items_0_amount
            else:
                credit_note_line_items_0_amount_cell.value = str(credit_note_line_items_0_amount[0])

        if credit_note_line_items_0_pricing_model_CNo != 'NA':
            credit_note_line_items_0_pricing_model = get_credit_note_line_items_0_pricing_model(i)
            credit_note_line_items_0_pricing_model_cell = my_sheet.cell(row=i + 2,
                                                                        column=credit_note_line_items_0_pricing_model_CNo)
            if credit_note_line_items_0_pricing_model == False:
                credit_note_line_items_0_pricing_model_cell.value = credit_note_line_items_0_pricing_model
            else:
                credit_note_line_items_0_pricing_model_cell.value = str(credit_note_line_items_0_pricing_model[0])

        if credit_note_line_items_0_is_taxed_CNo != 'NA':
            credit_note_line_items_0_is_taxed = get_credit_note_line_items_0_is_taxed(i)
            credit_note_line_items_0_is_taxed_cell = my_sheet.cell(row=i + 2,
                                                                   column=credit_note_line_items_0_is_taxed_CNo)
            if credit_note_line_items_0_is_taxed == False:
                credit_note_line_items_0_is_taxed_cell.value = credit_note_line_items_0_is_taxed
            else:
                credit_note_line_items_0_is_taxed_cell.value = str(credit_note_line_items_0_is_taxed[0])

        if credit_note_line_items_0_tax_amount_CNo != 'NA':
            credit_note_line_items_0_tax_amount = get_credit_note_line_items_0_tax_amount(i)
            credit_note_line_items_0_tax_amount_cell = my_sheet.cell(row=i + 2,
                                                                     column=credit_note_line_items_0_tax_amount_CNo)
            if credit_note_line_items_0_tax_amount == False:
                credit_note_line_items_0_tax_amount_cell.value = credit_note_line_items_0_tax_amount
            else:
                credit_note_line_items_0_tax_amount_cell.value = str(credit_note_line_items_0_tax_amount[0])

        if credit_note_line_items_0_object_CNo != 'NA':
            credit_note_line_items_0_object = get_credit_note_line_items_0_object(i)
            credit_note_line_items_0_object_cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_0_object_CNo)
            if credit_note_line_items_0_object == False:
                credit_note_line_items_0_object_cell.value = credit_note_line_items_0_object
            else:
                credit_note_line_items_0_object_cell.value = str(credit_note_line_items_0_object[0])

        if credit_note_line_items_0_subscription_id_CNo != 'NA':
            credit_note_line_items_0_subscription_id = get_credit_note_line_items_0_subscription_id(i)
            credit_note_line_items_0_subscription_id_cell = my_sheet.cell(row=i + 2,
                                                                          column=credit_note_line_items_0_subscription_id_CNo)
            if credit_note_line_items_0_subscription_id == False:
                credit_note_line_items_0_subscription_id_cell.value = credit_note_line_items_0_subscription_id
            else:
                credit_note_line_items_0_subscription_id_cell.value = str(credit_note_line_items_0_subscription_id[0])

        if credit_note_line_items_0_customer_id_CNo != 'NA':
            credit_note_line_items_0_customer_id = get_credit_note_line_items_0_customer_id(i)
            credit_note_line_items_0_customer_id_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_0_customer_id_CNo)
            if credit_note_line_items_0_customer_id == False:
                credit_note_line_items_0_customer_id_cell.value = credit_note_line_items_0_customer_id
            else:
                credit_note_line_items_0_customer_id_cell.value = str(credit_note_line_items_0_customer_id[0])

        if credit_note_line_items_0_description_CNo != 'NA':
            credit_note_line_items_0_description = get_credit_note_line_items_0_description(i)
            credit_note_line_items_0_description_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_0_description_CNo)
            if credit_note_line_items_0_description == False:
                credit_note_line_items_0_description_cell.value = credit_note_line_items_0_description
            else:
                credit_note_line_items_0_description_cell.value = str(credit_note_line_items_0_description[0])

        if credit_note_line_items_0_entity_type_CNo != 'NA':
            credit_note_line_items_0_entity_type = get_credit_note_line_items_0_entity_type(i)
            credit_note_line_items_0_entity_type_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_0_entity_type_CNo)
            if credit_note_line_items_0_entity_type == False:
                credit_note_line_items_0_entity_type_cell.value = credit_note_line_items_0_entity_type
            else:
                credit_note_line_items_0_entity_type_cell.value = str(credit_note_line_items_0_entity_type[0])

        if credit_note_line_items_0_entity_id_CNo != 'NA':
            credit_note_line_items_0_entity_id = get_credit_note_line_items_0_entity_id(i)
            credit_note_line_items_0_entity_id_cell = my_sheet.cell(row=i + 2,
                                                                    column=credit_note_line_items_0_entity_id_CNo)
            if credit_note_line_items_0_entity_id == False:
                credit_note_line_items_0_entity_id_cell.value = credit_note_line_items_0_entity_id
            else:
                credit_note_line_items_0_entity_id_cell.value = str(credit_note_line_items_0_entity_id[0])

        if credit_note_line_items_0_discount_amount_CNo != 'NA':
            credit_note_line_items_0_discount_amount = get_credit_note_line_items_0_discount_amount(i)
            credit_note_line_items_0_discount_amount_cell = my_sheet.cell(row=i + 2,
                                                                          column=credit_note_line_items_0_discount_amount_CNo)
            if credit_note_line_items_0_discount_amount == False:
                credit_note_line_items_0_discount_amount_cell.value = credit_note_line_items_0_discount_amount
            else:
                credit_note_line_items_0_discount_amount_cell.value = str(credit_note_line_items_0_discount_amount[0])

        if credit_note_line_items_0_item_level_discount_amount_CNo != 'NA':
            credit_note_line_items_0_item_level_discount_amount = get_credit_note_line_items_0_item_level_discount_amount(
                i)
            credit_note_line_items_0_item_level_discount_amount_cell = my_sheet.cell(row=i + 2,
                                                                                     column=credit_note_line_items_0_item_level_discount_amount_CNo)
            if credit_note_line_items_0_item_level_discount_amount == False:
                credit_note_line_items_0_item_level_discount_amount_cell.value = credit_note_line_items_0_item_level_discount_amount
            else:
                credit_note_line_items_0_item_level_discount_amount_cell.value = str(
                    credit_note_line_items_0_item_level_discount_amount[0])

        if credit_note_taxes_CNo != 'NA':
            credit_note_taxes = get_credit_note_taxes(i)
            credit_note_taxes_cell = my_sheet.cell(row=i + 2, column=credit_note_taxes_CNo)
            if credit_note_taxes == False:
                credit_note_taxes_cell.value = credit_note_taxes
            else:
                credit_note_taxes_cell.value = str(credit_note_taxes[0])

        if credit_note_line_item_taxes_CNo != 'NA':
            credit_note_line_item_taxes = get_credit_note_line_item_taxes(i)
            credit_note_line_item_taxes_cell = my_sheet.cell(row=i + 2, column=credit_note_line_item_taxes_CNo)
            if credit_note_line_item_taxes == False:
                credit_note_line_item_taxes_cell.value = credit_note_line_item_taxes
            else:
                credit_note_line_item_taxes_cell.value = str(credit_note_line_item_taxes[0])

        if credit_note_line_item_discounts_CNo != 'NA':
            credit_note_line_item_discounts = get_credit_note_line_item_discounts(i)
            credit_note_line_item_discounts_cell = my_sheet.cell(row=i + 2, column=credit_note_line_item_discounts_CNo)
            if credit_note_line_item_discounts == False:
                credit_note_line_item_discounts_cell.value = credit_note_line_item_discounts
            else:
                credit_note_line_item_discounts_cell.value = str(credit_note_line_item_discounts[0])

        if credit_note_linked_refunds_CNo != 'NA':
            credit_note_linked_refunds = get_credit_note_linked_refunds(i)
            credit_note_linked_refunds_cell = my_sheet.cell(row=i + 2, column=credit_note_linked_refunds_CNo)
            if credit_note_linked_refunds == False:
                credit_note_linked_refunds_cell.value = credit_note_linked_refunds
            else:
                credit_note_linked_refunds_cell.value = str(credit_note_linked_refunds[0])

        if credit_note_allocations_CNo != 'NA':
            credit_note_allocations = get_credit_note_allocations(i)
            credit_note_allocations_cell = my_sheet.cell(row=i + 2, column=credit_note_allocations_CNo)
            if credit_note_allocations == False:
                credit_note_allocations_cell.value = credit_note_allocations
            else:
                credit_note_allocations_cell.value = str(credit_note_allocations[0])

        if credit_note_refunded_at_CNo != 'NA':
            credit_note_refunded_at = get_credit_note_refunded_at(i)
            credit_note_refunded_at_cell = my_sheet.cell(row=i + 2, column=credit_note_refunded_at_CNo)
            if credit_note_refunded_at == False:
                credit_note_refunded_at_cell.value = credit_note_refunded_at
            else:
                credit_note_refunded_at_cell.value = str(credit_note_refunded_at[0])

        if credit_note_line_items_1__CNo != 'NA':
            credit_note_line_items_1_ = get_credit_note_line_items_1_(i)
            credit_note_line_items_1__cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_1__CNo)
            if credit_note_line_items_1_ == False:
                credit_note_line_items_1__cell.value = credit_note_line_items_1_
            else:
                credit_note_line_items_1__cell.value = str(credit_note_line_items_1_[0])

        if credit_note_line_items_1_id_CNo != 'NA':
            credit_note_line_items_1_id = get_credit_note_line_items_1_id(i)
            credit_note_line_items_1_id_cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_1_id_CNo)
            if credit_note_line_items_1_id == False:
                credit_note_line_items_1_id_cell.value = credit_note_line_items_1_id
            else:
                credit_note_line_items_1_id_cell.value = str(credit_note_line_items_1_id[0])

        if credit_note_line_items_1_date_from_CNo != 'NA':
            credit_note_line_items_1_date_from = get_credit_note_line_items_1_date_from(i)
            credit_note_line_items_1_date_from_cell = my_sheet.cell(row=i + 2,
                                                                    column=credit_note_line_items_1_date_from_CNo)
            if credit_note_line_items_1_date_from == False:
                credit_note_line_items_1_date_from_cell.value = credit_note_line_items_1_date_from
            else:
                credit_note_line_items_1_date_from_cell.value = str(credit_note_line_items_1_date_from[0])

        if credit_note_line_items_1_date_to_CNo != 'NA':
            credit_note_line_items_1_date_to = get_credit_note_line_items_1_date_to(i)
            credit_note_line_items_1_date_to_cell = my_sheet.cell(row=i + 2,
                                                                  column=credit_note_line_items_1_date_to_CNo)
            if credit_note_line_items_1_date_to == False:
                credit_note_line_items_1_date_to_cell.value = credit_note_line_items_1_date_to
            else:
                credit_note_line_items_1_date_to_cell.value = str(credit_note_line_items_1_date_to[0])

        if credit_note_line_items_1_unit_amount_CNo != 'NA':
            credit_note_line_items_1_unit_amount = get_credit_note_line_items_1_unit_amount(i)
            credit_note_line_items_1_unit_amount_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_1_unit_amount_CNo)
            if credit_note_line_items_1_unit_amount == False:
                credit_note_line_items_1_unit_amount_cell.value = credit_note_line_items_1_unit_amount
            else:
                credit_note_line_items_1_unit_amount_cell.value = str(credit_note_line_items_1_unit_amount[0])

        if credit_note_line_items_1_quantity_CNo != 'NA':
            credit_note_line_items_1_quantity = get_credit_note_line_items_1_quantity(i)
            credit_note_line_items_1_quantity_cell = my_sheet.cell(row=i + 2,
                                                                   column=credit_note_line_items_1_quantity_CNo)
            if credit_note_line_items_1_quantity == False:
                credit_note_line_items_1_quantity_cell.value = credit_note_line_items_1_quantity
            else:
                credit_note_line_items_1_quantity_cell.value = str(credit_note_line_items_1_quantity[0])

        if credit_note_line_items_1_amount_CNo != 'NA':
            credit_note_line_items_1_amount = get_credit_note_line_items_1_amount(i)
            credit_note_line_items_1_amount_cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_1_amount_CNo)
            if credit_note_line_items_1_amount == False:
                credit_note_line_items_1_amount_cell.value = credit_note_line_items_1_amount
            else:
                credit_note_line_items_1_amount_cell.value = str(credit_note_line_items_1_amount[0])

        if credit_note_line_items_1_pricing_model_CNo != 'NA':
            credit_note_line_items_1_pricing_model = get_credit_note_line_items_1_pricing_model(i)
            credit_note_line_items_1_pricing_model_cell = my_sheet.cell(row=i + 2,
                                                                        column=credit_note_line_items_1_pricing_model_CNo)
            if credit_note_line_items_1_pricing_model == False:
                credit_note_line_items_1_pricing_model_cell.value = credit_note_line_items_1_pricing_model
            else:
                credit_note_line_items_1_pricing_model_cell.value = str(credit_note_line_items_1_pricing_model[0])

        if credit_note_line_items_1_is_taxed_CNo != 'NA':
            credit_note_line_items_1_is_taxed = get_credit_note_line_items_1_is_taxed(i)
            credit_note_line_items_1_is_taxed_cell = my_sheet.cell(row=i + 2,
                                                                   column=credit_note_line_items_1_is_taxed_CNo)
            if credit_note_line_items_1_is_taxed == False:
                credit_note_line_items_1_is_taxed_cell.value = credit_note_line_items_1_is_taxed
            else:
                credit_note_line_items_1_is_taxed_cell.value = str(credit_note_line_items_1_is_taxed[0])

        if credit_note_line_items_1_tax_amount_CNo != 'NA':
            credit_note_line_items_1_tax_amount = get_credit_note_line_items_1_tax_amount(i)
            credit_note_line_items_1_tax_amount_cell = my_sheet.cell(row=i + 2,
                                                                     column=credit_note_line_items_1_tax_amount_CNo)
            if credit_note_line_items_1_tax_amount == False:
                credit_note_line_items_1_tax_amount_cell.value = credit_note_line_items_1_tax_amount
            else:
                credit_note_line_items_1_tax_amount_cell.value = str(credit_note_line_items_1_tax_amount[0])

        if credit_note_line_items_1_object_CNo != 'NA':
            credit_note_line_items_1_object = get_credit_note_line_items_1_object(i)
            credit_note_line_items_1_object_cell = my_sheet.cell(row=i + 2, column=credit_note_line_items_1_object_CNo)
            if credit_note_line_items_1_object == False:
                credit_note_line_items_1_object_cell.value = credit_note_line_items_1_object
            else:
                credit_note_line_items_1_object_cell.value = str(credit_note_line_items_1_object[0])

        if credit_note_line_items_1_subscription_id_CNo != 'NA':
            credit_note_line_items_1_subscription_id = get_credit_note_line_items_1_subscription_id(i)
            credit_note_line_items_1_subscription_id_cell = my_sheet.cell(row=i + 2,
                                                                          column=credit_note_line_items_1_subscription_id_CNo)
            if credit_note_line_items_1_subscription_id == False:
                credit_note_line_items_1_subscription_id_cell.value = credit_note_line_items_1_subscription_id
            else:
                credit_note_line_items_1_subscription_id_cell.value = str(credit_note_line_items_1_subscription_id[0])

        if credit_note_line_items_1_customer_id_CNo != 'NA':
            credit_note_line_items_1_customer_id = get_credit_note_line_items_1_customer_id(i)
            credit_note_line_items_1_customer_id_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_1_customer_id_CNo)
            if credit_note_line_items_1_customer_id == False:
                credit_note_line_items_1_customer_id_cell.value = credit_note_line_items_1_customer_id
            else:
                credit_note_line_items_1_customer_id_cell.value = str(credit_note_line_items_1_customer_id[0])

        if credit_note_line_items_1_description_CNo != 'NA':
            credit_note_line_items_1_description = get_credit_note_line_items_1_description(i)
            credit_note_line_items_1_description_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_1_description_CNo)
            if credit_note_line_items_1_description == False:
                credit_note_line_items_1_description_cell.value = credit_note_line_items_1_description
            else:
                credit_note_line_items_1_description_cell.value = str(credit_note_line_items_1_description[0])

        if credit_note_line_items_1_entity_type_CNo != 'NA':
            credit_note_line_items_1_entity_type = get_credit_note_line_items_1_entity_type(i)
            credit_note_line_items_1_entity_type_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_line_items_1_entity_type_CNo)
            if credit_note_line_items_1_entity_type == False:
                credit_note_line_items_1_entity_type_cell.value = credit_note_line_items_1_entity_type
            else:
                credit_note_line_items_1_entity_type_cell.value = str(credit_note_line_items_1_entity_type[0])

        if credit_note_line_items_1_entity_id_CNo != 'NA':
            credit_note_line_items_1_entity_id = get_credit_note_line_items_1_entity_id(i)
            credit_note_line_items_1_entity_id_cell = my_sheet.cell(row=i + 2,
                                                                    column=credit_note_line_items_1_entity_id_CNo)
            if credit_note_line_items_1_entity_id == False:
                credit_note_line_items_1_entity_id_cell.value = credit_note_line_items_1_entity_id
            else:
                credit_note_line_items_1_entity_id_cell.value = str(credit_note_line_items_1_entity_id[0])

        if credit_note_line_items_1_discount_amount_CNo != 'NA':
            credit_note_line_items_1_discount_amount = get_credit_note_line_items_1_discount_amount(i)
            credit_note_line_items_1_discount_amount_cell = my_sheet.cell(row=i + 2,
                                                                          column=credit_note_line_items_1_discount_amount_CNo)
            if credit_note_line_items_1_discount_amount == False:
                credit_note_line_items_1_discount_amount_cell.value = credit_note_line_items_1_discount_amount
            else:
                credit_note_line_items_1_discount_amount_cell.value = str(credit_note_line_items_1_discount_amount[0])

        if credit_note_line_items_1_item_level_discount_amount_CNo != 'NA':
            credit_note_line_items_1_item_level_discount_amount = get_credit_note_line_items_1_item_level_discount_amount(
                i)
            credit_note_line_items_1_item_level_discount_amount_cell = my_sheet.cell(row=i + 2,
                                                                                     column=credit_note_line_items_1_item_level_discount_amount_CNo)
            if credit_note_line_items_1_item_level_discount_amount == False:
                credit_note_line_items_1_item_level_discount_amount_cell.value = credit_note_line_items_1_item_level_discount_amount
            else:
                credit_note_line_items_1_item_level_discount_amount_cell.value = str(
                    credit_note_line_items_1_item_level_discount_amount[0])

        if credit_note_allocations_0__CNo != 'NA':
            credit_note_allocations_0_ = get_credit_note_allocations_0_(i)
            credit_note_allocations_0__cell = my_sheet.cell(row=i + 2, column=credit_note_allocations_0__CNo)
            if credit_note_allocations_0_ == False:
                credit_note_allocations_0__cell.value = credit_note_allocations_0_
            else:
                credit_note_allocations_0__cell.value = str(credit_note_allocations_0_[0])

        if credit_note_allocations_0_allocated_amount_CNo != 'NA':
            credit_note_allocations_0_allocated_amount = get_credit_note_allocations_0_allocated_amount(i)
            credit_note_allocations_0_allocated_amount_cell = my_sheet.cell(row=i + 2,
                                                                            column=credit_note_allocations_0_allocated_amount_CNo)
            if credit_note_allocations_0_allocated_amount == False:
                credit_note_allocations_0_allocated_amount_cell.value = credit_note_allocations_0_allocated_amount
            else:
                credit_note_allocations_0_allocated_amount_cell.value = str(
                    credit_note_allocations_0_allocated_amount[0])

        if credit_note_allocations_0_allocated_at_CNo != 'NA':
            credit_note_allocations_0_allocated_at = get_credit_note_allocations_0_allocated_at(i)
            credit_note_allocations_0_allocated_at_cell = my_sheet.cell(row=i + 2,
                                                                        column=credit_note_allocations_0_allocated_at_CNo)
            if credit_note_allocations_0_allocated_at == False:
                credit_note_allocations_0_allocated_at_cell.value = credit_note_allocations_0_allocated_at
            else:
                credit_note_allocations_0_allocated_at_cell.value = str(credit_note_allocations_0_allocated_at[0])

        if credit_note_allocations_0_invoice_id_CNo != 'NA':
            credit_note_allocations_0_invoice_id = get_credit_note_allocations_0_invoice_id(i)
            credit_note_allocations_0_invoice_id_cell = my_sheet.cell(row=i + 2,
                                                                      column=credit_note_allocations_0_invoice_id_CNo)
            if credit_note_allocations_0_invoice_id == False:
                credit_note_allocations_0_invoice_id_cell.value = credit_note_allocations_0_invoice_id
            else:
                credit_note_allocations_0_invoice_id_cell.value = str(credit_note_allocations_0_invoice_id[0])

        if credit_note_allocations_0_invoice_date_CNo != 'NA':
            credit_note_allocations_0_invoice_date = get_credit_note_allocations_0_invoice_date(i)
            credit_note_allocations_0_invoice_date_cell = my_sheet.cell(row=i + 2,
                                                                        column=credit_note_allocations_0_invoice_date_CNo)
            if credit_note_allocations_0_invoice_date == False:
                credit_note_allocations_0_invoice_date_cell.value = credit_note_allocations_0_invoice_date
            else:
                credit_note_allocations_0_invoice_date_cell.value = str(credit_note_allocations_0_invoice_date[0])

        if credit_note_allocations_0_invoice_status_CNo != 'NA':
            credit_note_allocations_0_invoice_status = get_credit_note_allocations_0_invoice_status(i)
            credit_note_allocations_0_invoice_status_cell = my_sheet.cell(row=i + 2,
                                                                          column=credit_note_allocations_0_invoice_status_CNo)
            if credit_note_allocations_0_invoice_status == False:
                credit_note_allocations_0_invoice_status_cell.value = credit_note_allocations_0_invoice_status
            else:
                credit_note_allocations_0_invoice_status_cell.value = str(credit_note_allocations_0_invoice_status[0])



    except Exception as e:
        print("in exception", e)

# finally save excel
my_wb.save(output_file_name)
print("Execution completed")
end = timer()
print('Total time taken: ', end - start, ' seconds')
