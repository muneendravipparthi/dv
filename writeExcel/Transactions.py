import json
import os
from timeit import default_timer as timer

import jsonpath
import openpyxl
from jproperties import Properties

from utils.Ephoc2DateTime import Ephoc2DateTime as tzconverter

configs = Properties()
ROOT_DIR1 = os.path.abspath(os.curdir)
ROOT_DIR = ROOT_DIR1.replace('writeExcel', 'configuration.properties')
with open(ROOT_DIR, 'rb') as config_file:
    configs.load(config_file)
Json_DIR = ROOT_DIR1.replace('writeExcel', 'readAPI/jsonfiles/')
start = timer()
input_jsonfile = Json_DIR + configs.get("clientName").data + "_AllTransactions.json"
output_file_name = ROOT_DIR1 + '/outputExcelFiles/' + configs.get("clientName").data + "_Transactions_Actual.xlsx"
clienttimezone = configs.get("clienttimezone").data

# -------------------Column Nos ----------------------#
transactionId_CNo = 1
gateway_account_id_CNo = 2
payment_method_CNo = 3
gateway_CNo = 4
type_CNo = 5
date_CNo = 6
amount_CNo = 7
status_CNo = 8
currency_code_CNo = 9
linked_invoices_CNo = 10
masked_card_number_CNo = 11

jsondata = open(input_jsonfile)
Transactiondictionary = json.load(jsondata)
print("Final Json File", Transactiondictionary)
apitojsontime = timer()

jsonpathres = Transactiondictionary
l = jsonpath.jsonpath(jsonpathres, "list")
totalRecordCountInResp = len(l[0])
# create work book and sheet object
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "Transaction"


if transactionId_CNo != 'NA':
    transactionId = my_sheet.cell(row=1, column=transactionId_CNo)
    transactionId.value = str("transactionId")

if gateway_account_id_CNo != 'NA':
    gateway_account_id = my_sheet.cell(row=1, column=gateway_account_id_CNo)
    gateway_account_id.value = str("gateway_account_id")

if payment_method_CNo != 'NA':
    payment_method = my_sheet.cell(row=1, column=payment_method_CNo)
    payment_method.value = str("payment_method")

if gateway_CNo != 'NA':
    gateway = my_sheet.cell(row=1, column=gateway_CNo)
    gateway.value = str("gateway")

if type_CNo != 'NA':
    type = my_sheet.cell(row=1, column=type_CNo)
    type.value = str("type")

if date_CNo != 'NA':
    date = my_sheet.cell(row=1, column=date_CNo)
    date.value = str("date")

if amount_CNo != 'NA':
    amount = my_sheet.cell(row=1, column=amount_CNo)
    amount.value = str("amount")

if status_CNo != 'NA':
    status = my_sheet.cell(row=1, column=status_CNo)
    status.value = str("status")

if currency_code_CNo != 'NA':
    currency_code = my_sheet.cell(row=1, column=currency_code_CNo)
    currency_code.value = str("currency_code")

if linked_invoices_CNo != 'NA':
    linked_invoices = my_sheet.cell(row=1, column=linked_invoices_CNo)
    linked_invoices.value = str("linked_invoices")


if masked_card_number_CNo != 'NA':
    masked_card_number = my_sheet.cell(row=1, column=masked_card_number_CNo)
    masked_card_number.value = str("masked_card_number")


def get_masked_card_number(i):
    try:
        masked_card_number = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.masked_card_number')
    except:
        print('in status exception')
    return masked_card_number


def get_transactionId(i):
    try:
        transactionId = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.id')
    except:
        print('in status exception')
    return transactionId


def get_gateway_account_id(i):
    try:
        gateway_account_id = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.gateway_account_id')
    except:
        print('in status exception')
    return gateway_account_id


def get_payment_method(i):
    try:
        payment_method = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.payment_method')
    except:
        print('in status exception')
    return payment_method


def get_gateway(i):
    try:
        gateway = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.gateway')
    except:
        print('in status exception')
    return gateway


def get_type(i):
    try:
        type = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + ']transaction.type')
    except:
        print('in status exception')
    return type


def get_date(i):
    try:
        date = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.date')
    except:
        print('in status exception')
    return date


def get_amount(i):
    try:
        amount = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.amount')
    except:
        print('in status exception')
    return amount


def get_status(i):
    try:
        status = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.status')
    except:
        print('in status exception')
    return status


def get_currency_code(i):
    try:
        currency_code = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.currency_code')
    except:
        print('in status exception')
    return currency_code


def get_linked_invoices(i):
    try:
        linked_invoices = jsonpath.jsonpath(jsonpathres, 'list[' + str(i) + '].transaction.linked_invoices')
    except:
        print('in status exception')
    return linked_invoices
    # ---------------------------loop through all records  and write to excel---------------------------#


for i in range(0, totalRecordCountInResp):
    try:
        if transactionId_CNo != 'NA':
            transactionId = get_transactionId(i)
            transactionId_cell = my_sheet.cell(row=i + 2, column=transactionId_CNo)
            if transactionId == False:
                transactionId_cell.value = transactionId
            else:
                transactionId_cell.value = str(transactionId[0])
        if gateway_account_id_CNo != 'NA':
            gateway_account_id = get_gateway_account_id(i)
            gateway_account_id_cell = my_sheet.cell(row=i + 2, column=gateway_account_id_CNo)
            if gateway_account_id == False:
                gateway_account_id_cell.value = gateway_account_id
            else:
                gateway_account_id_cell.value = str(gateway_account_id[0])
        if payment_method_CNo != 'NA':
            payment_method = get_payment_method(i)
            payment_method_cell = my_sheet.cell(row=i + 2, column=payment_method_CNo)
            if payment_method == False:
                payment_method_cell.value = payment_method
            else:
                payment_method_cell.value = str(payment_method[0])
        if gateway_CNo != 'NA':
            gateway = get_gateway(i)
            gateway_cell = my_sheet.cell(row=i + 2, column=gateway_CNo)
            if gateway == False:
                gateway_cell.value = gateway
            else:
                gateway_cell.value = str(gateway[0])
        if type_CNo != 'NA':
            type = get_type(i)
            type_cell = my_sheet.cell(row=i + 2, column=type_CNo)
            if type == False:
                type_cell.value = type
            else:
                type_cell.value = str(type[0])
        if date_CNo != 'NA':
            date = get_date(i)
            date_cell = my_sheet.cell(row=i + 2, column=date_CNo)
            if date == False:
                date_cell.value = date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(date[0], clienttimezone)
                date_cell.value = str(modifiedtimestamp)

        if amount_CNo != 'NA':
            amount = get_amount(i)
            amount_cell = my_sheet.cell(row=i + 2, column=amount_CNo)
            if amount == False:
                amount_cell.value = amount
            else:
                amount_cell.value = str(amount[0])
        if status_CNo != 'NA':
            status = get_status(i)
            status_cell = my_sheet.cell(row=i + 2, column=status_CNo)
            if status == False:
                status_cell.value = status
            else:
                status_cell.value = str(status[0])
        if currency_code_CNo != 'NA':
            currency_code = get_currency_code(i)
            currency_code_cell = my_sheet.cell(row=i + 2, column=currency_code_CNo)
            if currency_code == False:
                currency_code_cell.value = currency_code
            else:
                currency_code_cell.value = str(currency_code[0])
        if linked_invoices_CNo != 'NA':
            linked_invoices = get_linked_invoices(i)
            linked_invoices_cell = my_sheet.cell(row=i + 2, column=linked_invoices_CNo)
            if linked_invoices == False:
                linked_invoices_cell.value = linked_invoices
            else:
                linked_invoices_cell.value = str(linked_invoices[0])

        if masked_card_number_CNo != 'NA':
            masked_card_number = get_masked_card_number(i)
            masked_card_number_cell = my_sheet.cell(row=i + 2, column=masked_card_number_CNo)
            if masked_card_number == False:
                masked_card_number_cell.value = masked_card_number
            else:
                masked_card_number_cell.value = str(masked_card_number[0])


    except Exception as e:
        print("in exception", e)

# finally save excel
my_wb.save(output_file_name)
print("Execution completed")
end = timer()
print('Total time taken: ', end - start, ' seconds')
