# Python program to illustrate the concept
# of threading
import threading
import os
import json
import os
from timeit import default_timer as timer

import jsonpath
import openpyxl
from jproperties import Properties

from utils.Ephoc2DateTime import Ephoc2DateTime as tzconverter

configs = Properties()
ROOT_DIR1 = os.path.abspath(os.curdir)
ROOT_DIR = ROOT_DIR1.replace('writeExcel', 'configuration.properties')
with open(ROOT_DIR, 'rb') as config_file:
    configs.load(config_file)
Json_DIR = ROOT_DIR1.replace('writeExcel', 'readAPI/jsonfiles/')
start = timer()
# input_jsonfile = Json_DIR + configs.get("clientName").data + "_AllInvoices.json"
# output_file_name = ROOT_DIR1 + '/outputExcelFiles/' + configs.get("clientName").data + "_Invoice_Actual2.xlsx"
# clienttimezone = configs.get("clienttimezone").data
#
# jsondata = open(input_jsonfile)
# Invoicesdictionary = json.load(jsondata)
# print("Final Json File", Invoicesdictionary)
# apitojsontime = timer()
#
# jsonpathres = Invoicesdictionary
# l = jsonpath.jsonpath(jsonpathres, "list")
# totalRecordCountInResp = len(l[0])
# create work book and sheet object
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "Invoices"

datefields = ["invoice[date]", "invoice[due_date]", "line_items[date_from][0]", "line_items[date_to][0]",
              "payments[date][0]", "line_items[date_from][1]", "line_items[date_to][1]", "line_items[date_from][2]",
              "line_items[date_to][2]"]
amountfields = ["invoice[total]", "line_items[unit_amount][0]", "line_items[amount][0]",
                "line_items[item_level_discount1_amount][0]", "line_items[tax1_amount][0]", "discounts[amount][0]",
                "payments[amount][0]", "line_items[unit_amount][1]", "line_items[amount][1]",
                "line_items[unit_amount][2]", "line_items[amount][2]"]
columns = ["invoice[id]", "invoice[currency_code]", "invoice[customer_id]", "invoice[subscription_id]",
           "invoice[status]", "invoice[date]", "invoice[po_number]", "invoice[price_type]", "tax_override_reason",
           "invoice[vat_number]", "round_off", "invoice[total]", "invoice[due_date]", "invoice[net_term_days]",
           "use_for_proration", "line_items[id][0]", "line_items[entity_type][0]", "line_items[entity_id][0]",
           "line_items[description][0]", "line_items[date_from][0]", "line_items[date_to][0]",
           "line_items[quantity][0]", "line_items[unit_amount][0]", "line_items[amount][0]",
           "line_items[item_level_discount1_entity_id][0]", "line_items[item_level_discount1_amount][0]",
           "line_items[tax1_name][0]", "line_items[tax1_amount][0]", "discounts[entity_type][0]",
           "discounts[entity_id][0]", "discounts[description][0]", "discounts[amount][0]", "payments[amount][0]",
           "payments[date][0]", "line_items[entity_type][1]", "line_items[entity_id][1]", "line_items[description][1]",
           "line_items[date_from][1]", "line_items[date_to][1]", "line_items[quantity][1]",
           "line_items[unit_amount][1]", "line_items[amount][1]", "line_items[entity_type][2]",
           "line_items[entity_id][2]", "line_items[description][2]", "line_items[date_from][2]",
           "line_items[date_to][2]", "line_items[quantity][2]", "line_items[unit_amount][2]", "line_items[amount][2]",
           "billing_address[first_name]", "billing_address[last_name]", "billing_address[email]",
           "billing_address[company]", "billing_address[phone]", "billing_address[line1]", "billing_address[line2]",
           "billing_address[line3]", "billing_address[state_code]", "billing_address[city]", "billing_address[state]",
           "billing_address[zip]", "billing_address[country]", "shipping_address[first_name]",
           "shipping_address[last_name]", "shipping_address[email]", "shipping_address[company]",
           "shipping_address[phone]", "shipping_address[line1]", "shipping_address[line2]", "shipping_address[line3]",
           "shipping_address[city]", "shipping_address[state_code]", "shipping_address[state]", "shipping_address[zip]",
           "shipping_address[country]"]

jsoncall = [".invoice.id", ".invoice.currency_code", ".invoice.customer_id", ".invoice.subscription_id",
            ".invoice.status", ".invoice.date", ".invoice.po_number", ".invoice.price_type",
            ".invoice.tax_override_reason", ".invoice.vat_number", ".invoice.round_off_amount", ".invoice.total",
            ".invoice.due_date", ".invoice.net_term_days", ".invoice.use_for_proration", ".invoice.line_items[0].id",
            ".invoice.line_items[0].entity_type", ".invoice.line_items[0].entity_id",
            ".invoice.line_items[0].description", ".invoice.line_items[0].date_from", ".invoice.line_items[0].date_to",
            ".invoice.line_items[0].quantity", ".invoice.line_items[0].unit_amount", ".invoice.line_items[0].amount",
            ".invoice.line_items[0].item_level_discount_entity_id", ".invoice.line_items[0].item_level_discount_amount",
            ".invoice.line_items[0].tax_name", ".invoice.line_items[0].tax_amount", ".invoice.discounts[0].entity_type",
            ".invoice.discounts[0].entity_id", ".invoice.discounts[0].description", ".invoice.discounts[0].amount",
            ".invoice.linked_payments[0].txn_amount", ".invoice.linked_payments[0].txn_date",
            ".invoice.line_items[1].entity_type", ".invoice.line_items[1].entity_id",
            ".invoice.line_items[1].description", ".invoice.line_items[1].date_from", ".invoice.line_items[1].date_to",
            ".invoice.line_items[1].quantity", ".invoice.line_items[1].unit_amount", ".invoice.line_items[1].amount",
            ".invoice.line_items[2].entity_type", ".invoice.line_items[2].entity_id",
            ".invoice.line_items[2].description", ".invoice.line_items[2].date_from", ".invoice.line_items[2].date_to",
            ".invoice.line_items[2].quantity", ".invoice.line_items[2].unit_amount", ".invoice.line_items[2].amount",
            ".invoice.billing_address.first_name", ".invoice.billing_address.last_name",
            ".invoice.billing_address.email", ".invoice.billing_address.company", ".invoice.billing_address.phone",
            ".invoice.billing_address.line1", ".invoice.billing_address.line2", ".invoice.billing_address.line3",
            ".invoice.billing_address.state_code", ".invoice.billing_address.city", ".invoice.billing_address.state",
            ".invoice.billing_address.zip", ".invoice.billing_address.country", ".invoice.shipping_address.first_name",
            ".invoice.shipping_address.last_name", ".invoice.shipping_address.email",
            ".invoice.shipping_address.company", ".invoice.shipping_address.phone", ".invoice.shipping_address.line1",
            ".invoice.shipping_address.line2", ".invoice.shipping_address.line3",
            ".invoice.shipping_address.state_code", ".invoice.shipping_address.city", ".invoice.shipping_address.state",
            ".invoice.shipping_address.zip", ".invoice.shipping_address.country"]


def getdata(i, col):
    try:
        data = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "]" + jsoncall[col])
    except:
        print("in status exception : ", jsoncall[col])
    return data


# ---------------------------loop through all records  and write to excel---------------------------#

def task(strange, endrange, output):
    pass
    # for x in range(0, len(columns) - 1):
    #     colhead = my_sheet.cell(row=1, column=x + 1)
    #     colhead.value = str(columns[x])
    # for i in range(strange, endrange):
    #     for col in range(0, len(columns) - 1):
    #         try:
    #             coldata = getdata(i, col)
    #             coldata_cell = my_sheet.cell(row=i + 2, column=col + 1)
    #             if coldata == False:
    #                 coldata_cell.value = coldata
    #             else:
    #                 coldata_cell.value = str(coldata[0])
    #                 if columns[col] in datefields:
    #                     modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(coldata[0], clienttimezone)
    #                     coldata_cell.value = str(modifiedtimestamp)
    #                 if columns[col] in amountfields:
    #                     coldata_cell.value = str(coldata[0] / 100)
    #         except Exception as e:
    #             print("in exception", e)
    # # finally save excel
    # my_wb.save(output)
    # print("Execution completed")
    # end = timer()
    # print('Total time taken: ', end - start, ' seconds')

def task1(strange, endrange ):
    output = "t1.xlsx"
    print(output)
    # task(strange, endrange, output)
    for i in range(0, 100000):
        print("output file execution of {};{};{};{}".format(i,strange,endrange,output))

def task2(strange, endrange ):
    output = "t2.xlsx"
    print(output)
    # task(strange, endrange, output)
    for i in range(0, 100000):
        print("output file execution of {};{};{};{}".format(i,strange,endrange,output))

def task3(strange, endrange ):
    output = "t3.xlsx"
    # task(strange, endrange, output)
    for i in range(0, 100000):
        print("output file execution of {};{};{};{}".format(i,strange,endrange,output))

def task4(strange, endrange ):
    output = "t4.xlsx"
    # task(strange, endrange, output)
    for i in range(0, 100000):
        print("output file execution of {};{};{};{}".format(i,strange,endrange,output))

def task5(strange, endrange ):
    output = "t5.xlsx"
    # task(strange, endrange, output)
    for i in range(0, 100):
        print("output file execution of {};{};{};{}".format(i,strange,endrange,output))

def task6(strange, endrange ):
    output = "t6.xlsx"
    # task(strange, endrange, output)
    for i in range(0, 10000):
        print("output file execution of {};{};{};{}".format(i,strange,endrange,output))

def task7(strange, endrange ):
    output = "t7.xlsx"
    # task(strange, endrange, output)
    for i in range(0, 100):
        print("output file execution of {};{};{};{}".format(i,strange,endrange,output))

if __name__ == "__main__":
    # print ID of current process
    print("ID of process running main program: {}".format(os.getpid()))

    # print name of main thread
    print("Main thread name: {}".format(threading.current_thread().name))

    # creating threads
    t1 = threading.Thread(target=task1(0, 100000), name='t1')
    t2 = threading.Thread(target=task2(100001, 200000), name='t2')
    t3 = threading.Thread(target=task3(200001, 300000), name='t3')
    t4 = threading.Thread(target=task4(300001, 400000), name='t4')
    t5 = threading.Thread(target=task5(400001, 500000), name='t5')
    t6 = threading.Thread(target=task6(500001, 600000), name='t6')
    # t7 = threading.Thread(target=task7(600001, totalRecordCountInResp), name='t7')

    # starting threads
    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    t6.start()
    # t7.start()

    # wait until all threads finish
    t1.join()
    t2.join()
    t3.join()
    t4.join()
    t5.join()
    t6.join()
    # t7.join()
