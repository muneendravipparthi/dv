import json
import os
from timeit import default_timer as timer

import jsonpath
import openpyxl
from jproperties import Properties

from utils.Ephoc2DateTime import Ephoc2DateTime as tzconverter

configs = Properties()
ROOT_DIR1 = os.path.abspath(os.curdir)
ROOT_DIR = ROOT_DIR1.replace('writeExcel', 'configuration.properties')
with open(ROOT_DIR, 'rb') as config_file:
    configs.load(config_file)
Json_DIR = ROOT_DIR1.replace('writeExcel', 'readAPI/jsonfiles/')
start = timer()
input_jsonfile = Json_DIR + configs.get("clientName").data + "_AllInvoices.json"
output_file_name = ROOT_DIR1 + '/outputExcelFiles/' + configs.get("clientName").data + "_Invoice_Actual2.xlsx"
clienttimezone = configs.get("clienttimezone").data

# -------------------Column Nos ----------------------#
currency_code_CNo = 1
customer_id_CNo = 2
date_CNo = 3
due_date_CNo = 4
id_CNo = 5
net_term_days_CNo = 6
price_type_CNo = 8
status_CNo = 9
stripe_id_CNo = 10
subscription_id_CNo = 11
total_CNo = 12
vat_number_CNo = 13
line_items_0_amount_CNo = 14
line_items_0_date_from_CNo = 15
line_items_0_date_to_CNo = 16
line_items_0_description_CNo = 17
line_items_0_entity_id_CNo = 18
line_items_0_entity_type_CNo = 19
line_items_0_id_CNo = 20
line_items_0_quantity_CNo = 21
payments_0_amount_CNo = 22
payments_0_date_CNo = 23
payments_0_payment_method_CNo = 24
discounts_0_amount_CNo = 25
discounts_0_entity_id_CNo = 26
discounts_0_entity_type_CNo = 28
line_items_0_item_level_discount1_entity_id_CNo = 29
line_items_0_item_level_discount1_amount_CNo = 30
line_items_0_unit_amount_CNo =   "NA"
line_items_1_entity_type_CNo =   "NA"
line_items_1_id_CNo =   "NA"
line_items_1_entity_id_CNo =   "NA"
line_items_1_description_CNo =   "NA"
line_items_1_date_from_CNo =   "NA"
line_items_1_date_to_CNo =   "NA"
line_items_1_quantity_CNo =   "NA"
line_items_1_unit_amount_CNo =   "NA"
line_items_1_amount_CNo =   "NA"
line_items_2_entity_type_CNo =   "NA"
line_items_2_entity_id_CNo =   "NA"
line_items_2_description_CNo =   "NA"
line_items_2_date_from_CNo =   "NA"
line_items_2_date_to_CNo =   "NA"
line_items_2_quantity_CNo =   "NA"
line_items_2_unit_amount_CNo =   "NA"
line_items_2_amount_CNo =   "NA"
line_items_3_entity_type_CNo =   "NA"
line_items_3_entity_id_CNo =   "NA"
line_items_3_description_CNo =   "NA"
line_items_3_date_from_CNo =   "NA"
line_items_3_date_to_CNo =   "NA"
line_items_3_quantity_CNo =   "NA"
line_items_3_unit_amount_CNo =   "NA"
line_items_3_amount_CNo =   "NA"
line_items_4_entity_type_CNo =   "NA"
line_items_4_entity_id_CNo =   "NA"
line_items_4_description_CNo =   "NA"
line_items_4_date_from_CNo = "NA"
line_items_4_date_to_CNo =   "NA"
line_items_4_quantity_CNo =   "NA"
line_items_4_unit_amount_CNo =   "NA"
line_items_4_amount_CNo =   "NA"
line_items_5_entity_type_CNo =   "NA"
line_items_5_entity_id_CNo =   "NA"
line_items_5_description_CNo =   "NA"
line_items_5_date_from_CNo =   "NA"
line_items_5_date_to_CNo =   "NA"
line_items_5_quantity_CNo =   "NA"
line_items_5_unit_amount_CNo =   "NA"
line_items_5_amount_CNo =   "NA"
line_items_6_entity_type_CNo =   "NA"
line_items_6_description_CNo =   "NA"
line_items_6_date_from_CNo =   "NA"
line_items_6_date_to_CNo =   "NA"
line_items_6_quantity_CNo =   "NA"
line_items_6_unit_amount_CNo =   "NA"
line_items_6_amount_CNo =   "NA"
line_items_7_entity_type_CNo =   "NA"
line_items_7_description_CNo =   "NA"
line_items_7_date_from_CNo =   "NA"
line_items_7_date_to_CNo =   "NA"
line_items_7_quantity_CNo =   "NA"
line_items_7_unit_amount_CNo =   "NA"
line_items_7_amount_CNo =   "NA"
use_for_proration_CNo =   "NA"
billing_address_first_name_CNo =   "NA"
billing_address_last_name_CNo =   "NA"
billing_address_line1_CNo =   "NA"
billing_address_line2_CNo =   "NA"
billing_address_city_CNo =   "NA"
billing_address_state_CNo =   "NA"
billing_address_zip_CNo =   "NA"
billing_address_country_CNo =   "NA"
billing_address_company_CNo =   "NA"
billing_address_email_CNo =   "NA"
Stripe_customer_id_CNo =   "NA"
shipping_address_city_CNo =   "NA"
shipping_address_country_CNo =   "NA"
shipping_address_first_name_CNo =   "NA"
shipping_address_last_name_CNo =   "NA"
shipping_address_line1_CNo =   "NA"
shipping_address_phone_CNo =   "NA"
shipping_address_zip_CNo =   "NA"
line_items_0_tax1_amount_CNo =   "NA"
line_items_0_tax1_name_CNo =   "NA"
taxes_0_amount_CNo =   "NA"
taxes_0_name_CNo =   "NA"
taxes_0_rate_CNo =   "NA"
line_items_1_tax1_amount_CNo =   "NA"
line_items_1_tax1_name_CNo =   "NA"
line_items_2_id_CNo =   "NA"
line_items_3_id_CNo =   "NA"
line_items_2_tax1_amount_CNo =   "NA"
line_items_3_tax1_amount_CNo =   "NA"
line_items_2_tax1_name_CNo =   "NA"
date_from =   "NA"
date_to =   "NA"
billing_address_line3_CNo =   "NA"
billing_address_phone_CNo =   "NA"
billing_address_state_code_CNo =   "NA"
billing_address_validation_status_CNo =   "NA"
round_off_CNo =   "NA"
shipping_address_company_CNo =   "NA"
shipping_address_email_CNo =   "NA"
shipping_address_line2_CNo =   "NA"
shipping_address_line3_CNo =   "NA"
shipping_address_state_CNo =   "NA"
shipping_address_state_code_CNo =   "NA"
shipping_address_validation_status_CNo =   "NA"
tax_override_reason_CNo =   "NA"
line_items_4_id_CNo =   "NA"
line_items_5_id_CNo =   "NA"
Company_name_CNo =   "NA"
invoice_po_number_CNo =   "NA"
round_off_amount_CNo =   "NA"
line_items_0_item_level_discount2_entity_id_CNo =   "NA"
line_items_0_item_level_discount2_amount_CNo =   "NA"
line_items_0_tax2_name_CNo =   "NA"
line_items_0_tax2_amount_CNo =   "NA"
line_items_0_tax3_name_CNo =   "NA"
line_items_0_tax3_amount_CNo =   "NA"
line_items_0_tax4_name_CNo =   "NA"
line_items_0_tax4_amount_CNo =   "NA"
line_item_0_tiers_line_item_id_CNo =   "NA"
line_item_0_tiers_starting_unit_CNo =   "NA"
line_item_0_tiers_ending_unit_CNo =   "NA"
line_item_0_tiers_quantity_used_CNo =   "NA"
line_item_0_tiers_unit_amount_CNo =   "NA"
discounts_0_description_CNo =   "NA"
taxes_0_description_CNo =   "NA"
taxes_0_juris_type_CNo =   "NA"
taxes_0_juris_name_CNo =   "NA"
taxes_0_juris_code_CNo =   "NA"
payments_0_reference_number_CNo =   "NA"
notes_0_entity_type_CNo =   "NA"
notes_0_entity_id_CNo =   "NA"
notes_0_note_CNo =   "NA"
recurring_CNo =   "NA"
exchange_rate_CNo =   "NA"
amount_paid_CNo =   "NA"
amount_adjusted_CNo =   "NA"
write_off_amount_CNo =   "NA"
credits_applied_CNo =   "NA"
amount_due_CNo =   "NA"
dunning_status_CNo =   "NA"
next_retry_at_CNo =   "NA"
updated_at_CNo =   "NA"
resource_version_CNo =   "NA"
deleted_CNo =   "NA"
object_CNo =   "NA"
first_CNo =   "NA"
amount_to_collect_CNo =   "NA"
has_advance_charges_CNo =   "NA"
base_currency_code_CNo =   "NA"
is_gifted_CNo =   "NA"
term_finalized_CNo =   "NA"
is_digital_CNo =   "NA"
tax_CNo =   "NA"
line_items_CNo =   "NA"
line_items_0__CNo =   "NA"
line_items_0_pricing_model_CNo =   "NA"
line_items_0_is_taxed_CNo =   "NA"
line_items_0_tax_amount_CNo =   "NA"
line_items_0_tax_rate_CNo =   "NA"
line_items_0_object_CNo =   "NA"
line_items_0_subscription_id_CNo =   "NA"
line_items_0_customer_id_CNo =   "NA"
line_items_0_discount_amount_CNo =   "NA"
line_items_0_item_level_discount_amount_CNo =   "NA"
taxes_CNo =   "NA"
taxes_0__CNo =   "NA"
taxes_0_object_CNo =   "NA"
line_item_taxes_CNo =   "NA"
line_item_taxes_0__CNo =   "NA"
line_item_taxes_0_tax_name_CNo =   "NA"
line_item_taxes_0_tax_rate_CNo =   "NA"
line_item_taxes_0_tax_juris_type_CNo =   "NA"
line_item_taxes_0_tax_juris_name_CNo =   "NA"
line_item_taxes_0_tax_juris_code_CNo =   "NA"
line_item_taxes_0_object_CNo =   "NA"
line_item_taxes_0_line_item_id_CNo =   "NA"
line_item_taxes_0_tax_amount_CNo =   "NA"
line_item_taxes_0_is_partial_tax_applied_CNo =   "NA"
line_item_taxes_0_taxable_amount_CNo =   "NA"
line_item_taxes_0_is_non_compliance_tax_CNo =   "NA"
sub_total_CNo =   "NA"
linked_payments_CNo =   "NA"
dunning_attempts_CNo =   "NA"
applied_credits_CNo =   "NA"
adjustment_credit_notes_CNo =   "NA"
issued_credit_notes_CNo =   "NA"
linked_orders_CNo =   "NA"
billing_address_CNo =   "NA"
billing_address_object_CNo =   "NA"
line_items_3_tax1_name_CNo =   "NA"
line_items_4_tax1_name_CNo =   "NA"
line_items_4_tax1_amount_CNo =   "NA"
line_items_5_tax1_name_CNo =   "NA"
line_items_5_tax1_amount_CNo =   "NA"
first_invoice_CNo =   "NA"
line_items_6_entity_id_CNo =   "NA"
line_items_6_tax1_name_CNo =   "NA"
line_items_6_tax1_amount_CNo =   "NA"
line_items_7_entity_id_CNo =   "NA"
line_items_7_tax1_name_CNo =   "NA"
line_items_7_tax1_amount_CNo =   "NA"
line_items_6_id_CNo =   "NA"
line_items_7_id_CNo =   "NA"



jsondata = open(input_jsonfile)
Invoicesdictionary = json.load(jsondata)
print("Final Json File", Invoicesdictionary)
apitojsontime = timer()

jsonpathres = Invoicesdictionary
l = jsonpath.jsonpath(jsonpathres, "list")
totalRecordCountInResp = len(l[0])
# create work book and sheet object
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "Invoices"
# ---------------------------write headers in excel -- first row---------------------------#

if Company_name_CNo != "NA":
    company = my_sheet.cell(row=1, column=Company_name_CNo)
    company.value = str("company")

if id_CNo != "NA":
    id = my_sheet.cell(row=1, column=id_CNo)
    id.value = str("id")

if customer_id_CNo != "NA":
    customer_id = my_sheet.cell(row=1, column=customer_id_CNo)
    customer_id.value = str("customer_id")

if subscription_id_CNo != "NA":
    subscription_id = my_sheet.cell(row=1, column=subscription_id_CNo)
    subscription_id.value = str("subscription_id")

if invoice_po_number_CNo != "NA":
    invoice_po_number = my_sheet.cell(row=1, column=invoice_po_number_CNo)
    invoice_po_number.value = str("invoice_po_number")

if tax_override_reason_CNo != "NA":
    tax_override_reason = my_sheet.cell(row=1, column=tax_override_reason_CNo)
    tax_override_reason.value = str("tax_override_reason")

if recurring_CNo != "NA":
    recurring = my_sheet.cell(row=1, column=recurring_CNo)
    recurring.value = str("recurring")

if status_CNo != "NA":
    status = my_sheet.cell(row=1, column=status_CNo)
    status.value = str("status")

if vat_number_CNo != "NA":
    vat_number = my_sheet.cell(row=1, column=vat_number_CNo)
    vat_number.value = str("vat_number")

if price_type_CNo != "NA":
    price_type = my_sheet.cell(row=1, column=price_type_CNo)
    price_type.value = str("price_type")

if date_CNo != "NA":
    date = my_sheet.cell(row=1, column=date_CNo)
    date.value = str("date")

if due_date_CNo != "NA":
    due_date = my_sheet.cell(row=1, column=due_date_CNo)
    due_date.value = str("due_date")

if net_term_days_CNo != "NA":
    net_term_days = my_sheet.cell(row=1, column=net_term_days_CNo)
    net_term_days.value = str("net_term_days")

if use_for_proration_CNo != "NA":
    use_for_proration = my_sheet.cell(row=1, column=use_for_proration_CNo)
    use_for_proration.value = str("use_for_proration")

if exchange_rate_CNo != "NA":
    exchange_rate = my_sheet.cell(row=1, column=exchange_rate_CNo)
    exchange_rate.value = str("exchange_rate")

if total_CNo != "NA":
    total = my_sheet.cell(row=1, column=total_CNo)
    total.value = str("total")

if amount_paid_CNo != "NA":
    amount_paid = my_sheet.cell(row=1, column=amount_paid_CNo)
    amount_paid.value = str("amount_paid")

if amount_adjusted_CNo != "NA":
    amount_adjusted = my_sheet.cell(row=1, column=amount_adjusted_CNo)
    amount_adjusted.value = str("amount_adjusted")

if write_off_amount_CNo != "NA":
    write_off_amount = my_sheet.cell(row=1, column=write_off_amount_CNo)
    write_off_amount.value = str("write_off_amount")

if credits_applied_CNo != "NA":
    credits_applied = my_sheet.cell(row=1, column=credits_applied_CNo)
    credits_applied.value = str("credits_applied")

if amount_due_CNo != "NA":
    amount_due = my_sheet.cell(row=1, column=amount_due_CNo)
    amount_due.value = str("amount_due")

if dunning_status_CNo != "NA":
    dunning_status = my_sheet.cell(row=1, column=dunning_status_CNo)
    dunning_status.value = str("dunning_status")

if next_retry_at_CNo != "NA":
    next_retry_at = my_sheet.cell(row=1, column=next_retry_at_CNo)
    next_retry_at.value = str("next_retry_at")

if updated_at_CNo != "NA":
    updated_at = my_sheet.cell(row=1, column=updated_at_CNo)
    updated_at.value = str("updated_at")

if resource_version_CNo != "NA":
    resource_version = my_sheet.cell(row=1, column=resource_version_CNo)
    resource_version.value = str("resource_version")

if deleted_CNo != "NA":
    deleted = my_sheet.cell(row=1, column=deleted_CNo)
    deleted.value = str("deleted")

if object_CNo != "NA":
    object = my_sheet.cell(row=1, column=object_CNo)
    object.value = str("object")

if first_invoice_CNo != "NA":
    first_invoice = my_sheet.cell(row=1, column=first_invoice_CNo)
    first_invoice.value = str("first_invoice")

if amount_to_collect_CNo != "NA":
    amount_to_collect = my_sheet.cell(row=1, column=amount_to_collect_CNo)
    amount_to_collect.value = str("amount_to_collect")

if round_off_amount_CNo != "NA":
    round_off_amount = my_sheet.cell(row=1, column=round_off_amount_CNo)
    round_off_amount.value = str("round_off_amount")

if has_advance_charges_CNo != "NA":
    has_advance_charges = my_sheet.cell(row=1, column=has_advance_charges_CNo)
    has_advance_charges.value = str("has_advance_charges")

if currency_code_CNo != "NA":
    currency_code = my_sheet.cell(row=1, column=currency_code_CNo)
    currency_code.value = str("currency_code")

if base_currency_code_CNo != "NA":
    base_currency_code = my_sheet.cell(row=1, column=base_currency_code_CNo)
    base_currency_code.value = str("base_currency_code")

if is_gifted_CNo != "NA":
    is_gifted = my_sheet.cell(row=1, column=is_gifted_CNo)
    is_gifted.value = str("is_gifted")

if term_finalized_CNo != "NA":
    term_finalized = my_sheet.cell(row=1, column=term_finalized_CNo)
    term_finalized.value = str("term_finalized")

if is_digital_CNo != "NA":
    is_digital = my_sheet.cell(row=1, column=is_digital_CNo)
    is_digital.value = str("is_digital")

if tax_CNo != "NA":
    tax = my_sheet.cell(row=1, column=tax_CNo)
    tax.value = str("tax")

if line_items_CNo != "NA":
    line_items = my_sheet.cell(row=1, column=line_items_CNo)
    line_items.value = str("line_items")

if line_items_0__CNo != "NA":
    line_items_0_ = my_sheet.cell(row=1, column=line_items_0__CNo)
    line_items_0_.value = str("line_items_0_")

if line_items_0_id_CNo != "NA":
    line_items_0_id = my_sheet.cell(row=1, column=line_items_0_id_CNo)
    line_items_0_id.value = str("line_items_0_id")

if line_items_0_date_from_CNo != "NA":
    line_items_0_date_from = my_sheet.cell(row=1, column=line_items_0_date_from_CNo)
    line_items_0_date_from.value = str("line_items_0_date_from")

if line_items_0_date_to_CNo != "NA":
    line_items_0_date_to = my_sheet.cell(row=1, column=line_items_0_date_to_CNo)
    line_items_0_date_to.value = str("line_items_0_date_to")

if line_items_0_unit_amount_CNo != "NA":
    line_items_0_unit_amount = my_sheet.cell(row=1, column=line_items_0_unit_amount_CNo)
    line_items_0_unit_amount.value = str("line_items_0_unit_amount")

if line_items_0_quantity_CNo != "NA":
    line_items_0_quantity = my_sheet.cell(row=1, column=line_items_0_quantity_CNo)
    line_items_0_quantity.value = str("line_items_0_quantity")

if line_items_0_amount_CNo != "NA":
    line_items_0_amount = my_sheet.cell(row=1, column=line_items_0_amount_CNo)
    line_items_0_amount.value = str("line_items_0_amount")

if line_items_0_pricing_model_CNo != "NA":
    line_items_0_pricing_model = my_sheet.cell(row=1, column=line_items_0_pricing_model_CNo)
    line_items_0_pricing_model.value = str("line_items_0_pricing_model")

if line_items_0_is_taxed_CNo != "NA":
    line_items_0_is_taxed = my_sheet.cell(row=1, column=line_items_0_is_taxed_CNo)
    line_items_0_is_taxed.value = str("line_items_0_is_taxed")

if line_items_0_tax_amount_CNo != "NA":
    line_items_0_tax_amount = my_sheet.cell(row=1, column=line_items_0_tax_amount_CNo)
    line_items_0_tax_amount.value = str("line_items_0_tax_amount")

if line_items_0_tax_rate_CNo != "NA":
    line_items_0_tax_rate = my_sheet.cell(row=1, column=line_items_0_tax_rate_CNo)
    line_items_0_tax_rate.value = str("line_items_0_tax_rate")

if line_items_0_object_CNo != "NA":
    line_items_0_object = my_sheet.cell(row=1, column=line_items_0_object_CNo)
    line_items_0_object.value = str("line_items_0_object")

if line_items_0_subscription_id_CNo != "NA":
    line_items_0_subscription_id = my_sheet.cell(row=1, column=line_items_0_subscription_id_CNo)
    line_items_0_subscription_id.value = str("line_items_0_subscription_id")

if line_items_0_customer_id_CNo != "NA":
    line_items_0_customer_id = my_sheet.cell(row=1, column=line_items_0_customer_id_CNo)
    line_items_0_customer_id.value = str("line_items_0_customer_id")

if line_items_0_description_CNo != "NA":
    line_items_0_description = my_sheet.cell(row=1, column=line_items_0_description_CNo)
    line_items_0_description.value = str("line_items_0_description")

if line_items_0_entity_type_CNo != "NA":
    line_items_0_entity_type = my_sheet.cell(row=1, column=line_items_0_entity_type_CNo)
    line_items_0_entity_type.value = str("line_items_0_entity_type")

if line_items_0_entity_id_CNo != "NA":
    line_items_0_entity_id = my_sheet.cell(row=1, column=line_items_0_entity_id_CNo)
    line_items_0_entity_id.value = str("line_items_0_entity_id")

if line_items_0_discount_amount_CNo != "NA":
    line_items_0_discount_amount = my_sheet.cell(row=1, column=line_items_0_discount_amount_CNo)
    line_items_0_discount_amount.value = str("line_items_0_discount_amount")

if line_items_0_item_level_discount_amount_CNo != "NA":
    line_items_0_item_level_discount_amount = my_sheet.cell(row=1, column=line_items_0_item_level_discount_amount_CNo)
    line_items_0_item_level_discount_amount.value = str("line_items_0_item_level_discount_amount")

if line_items_0_tax1_name_CNo != "NA":
    line_items_0_tax1_name = my_sheet.cell(row=1, column=line_items_0_tax1_name_CNo)
    line_items_0_tax1_name.value = str("line_items_0_tax1_name")

if line_items_0_tax1_amount_CNo != "NA":
    line_items_0_tax1_amount = my_sheet.cell(row=1, column=line_items_0_tax1_amount_CNo)
    line_items_0_tax1_amount.value = str("line_items_0_tax1_amount")

if taxes_CNo != "NA":
    taxes = my_sheet.cell(row=1, column=taxes_CNo)
    taxes.value = str("taxes")

if taxes_0__CNo != "NA":
    taxes_0_ = my_sheet.cell(row=1, column=taxes_0__CNo)
    taxes_0_.value = str("taxes_0_")

if taxes_0_object_CNo != "NA":
    taxes_0_object = my_sheet.cell(row=1, column=taxes_0_object_CNo)
    taxes_0_object.value = str("taxes_0_object")

if taxes_0_name_CNo != "NA":
    taxes_0_name = my_sheet.cell(row=1, column=taxes_0_name_CNo)
    taxes_0_name.value = str("taxes_0_name")

if taxes_0_rate_CNo != "NA":
    taxes_0_rate = my_sheet.cell(row=1, column=taxes_0_rate_CNo)
    taxes_0_rate.value = str("taxes_0_rate")

if taxes_0_description_CNo != "NA":
    taxes_0_description = my_sheet.cell(row=1, column=taxes_0_description_CNo)
    taxes_0_description.value = str("taxes_0_description")

if taxes_0_amount_CNo != "NA":
    taxes_0_amount = my_sheet.cell(row=1, column=taxes_0_amount_CNo)
    taxes_0_amount.value = str("taxes_0_amount")

if discounts_0_entity_type_CNo != "NA":
    discounts_0_entity_type = my_sheet.cell(row=1, column=discounts_0_entity_type_CNo)
    discounts_0_entity_type.value = str("discounts_0_entity_type")

if discounts_0_entity_id_CNo != "NA":
    discounts_0_entity_id = my_sheet.cell(row=1, column=discounts_0_entity_id_CNo)
    discounts_0_entity_id.value = str("discounts_0_entity_id")

if discounts_0_amount_CNo != "NA":
    discounts_0_amount = my_sheet.cell(row=1, column=discounts_0_amount_CNo)
    discounts_0_amount.value = str("discounts_0_amount")

if payments_0_amount_CNo != "NA":
    payments_0_amount = my_sheet.cell(row=1, column=payments_0_amount_CNo)
    payments_0_amount.value = str("payments_0_amount")

if payments_0_payment_method_CNo != "NA":
    payments_0_payment_method = my_sheet.cell(row=1, column=payments_0_payment_method_CNo)
    payments_0_payment_method.value = str("payments_0_payment_method")

if payments_0_date_CNo != "NA":
    payments_0_date = my_sheet.cell(row=1, column=payments_0_date_CNo)
    payments_0_date.value = str("payments_0_date")

if line_item_taxes_CNo != "NA":
    line_item_taxes = my_sheet.cell(row=1, column=line_item_taxes_CNo)
    line_item_taxes.value = str("line_item_taxes")

if line_item_taxes_0__CNo != "NA":
    line_item_taxes_0_ = my_sheet.cell(row=1, column=line_item_taxes_0__CNo)
    line_item_taxes_0_.value = str("line_item_taxes_0_")

if line_item_taxes_0_tax_name_CNo != "NA":
    line_item_taxes_0_tax_name = my_sheet.cell(row=1, column=line_item_taxes_0_tax_name_CNo)
    line_item_taxes_0_tax_name.value = str("line_item_taxes_0_tax_name")

if line_item_taxes_0_tax_rate_CNo != "NA":
    line_item_taxes_0_tax_rate = my_sheet.cell(row=1, column=line_item_taxes_0_tax_rate_CNo)
    line_item_taxes_0_tax_rate.value = str("line_item_taxes_0_tax_rate")

if line_item_taxes_0_tax_juris_type_CNo != "NA":
    line_item_taxes_0_tax_juris_type = my_sheet.cell(row=1, column=line_item_taxes_0_tax_juris_type_CNo)
    line_item_taxes_0_tax_juris_type.value = str("line_item_taxes_0_tax_juris_type")

if line_item_taxes_0_tax_juris_name_CNo != "NA":
    line_item_taxes_0_tax_juris_name = my_sheet.cell(row=1, column=line_item_taxes_0_tax_juris_name_CNo)
    line_item_taxes_0_tax_juris_name.value = str("line_item_taxes_0_tax_juris_name")

if line_item_taxes_0_tax_juris_code_CNo != "NA":
    line_item_taxes_0_tax_juris_code = my_sheet.cell(row=1, column=line_item_taxes_0_tax_juris_code_CNo)
    line_item_taxes_0_tax_juris_code.value = str("line_item_taxes_0_tax_juris_code")

if line_item_taxes_0_object_CNo != "NA":
    line_item_taxes_0_object = my_sheet.cell(row=1, column=line_item_taxes_0_object_CNo)
    line_item_taxes_0_object.value = str("line_item_taxes_0_object")

if line_item_taxes_0_line_item_id_CNo != "NA":
    line_item_taxes_0_line_item_id = my_sheet.cell(row=1, column=line_item_taxes_0_line_item_id_CNo)
    line_item_taxes_0_line_item_id.value = str("line_item_taxes_0_line_item_id")

if line_item_taxes_0_tax_amount_CNo != "NA":
    line_item_taxes_0_tax_amount = my_sheet.cell(row=1, column=line_item_taxes_0_tax_amount_CNo)
    line_item_taxes_0_tax_amount.value = str("line_item_taxes_0_tax_amount")

if line_item_taxes_0_is_partial_tax_applied_CNo != "NA":
    line_item_taxes_0_is_partial_tax_applied = my_sheet.cell(row=1, column=line_item_taxes_0_is_partial_tax_applied_CNo)
    line_item_taxes_0_is_partial_tax_applied.value = str("line_item_taxes_0_is_partial_tax_applied")

if line_item_taxes_0_taxable_amount_CNo != "NA":
    line_item_taxes_0_taxable_amount = my_sheet.cell(row=1, column=line_item_taxes_0_taxable_amount_CNo)
    line_item_taxes_0_taxable_amount.value = str("line_item_taxes_0_taxable_amount")

if line_item_taxes_0_is_non_compliance_tax_CNo != "NA":
    line_item_taxes_0_is_non_compliance_tax = my_sheet.cell(row=1, column=line_item_taxes_0_is_non_compliance_tax_CNo)
    line_item_taxes_0_is_non_compliance_tax.value = str("line_item_taxes_0_is_non_compliance_tax")

if sub_total_CNo != "NA":
    sub_total = my_sheet.cell(row=1, column=sub_total_CNo)
    sub_total.value = str("sub_total")

if linked_payments_CNo != "NA":
    linked_payments = my_sheet.cell(row=1, column=linked_payments_CNo)
    linked_payments.value = str("linked_payments")

if dunning_attempts_CNo != "NA":
    dunning_attempts = my_sheet.cell(row=1, column=dunning_attempts_CNo)
    dunning_attempts.value = str("dunning_attempts")

if applied_credits_CNo != "NA":
    applied_credits = my_sheet.cell(row=1, column=applied_credits_CNo)
    applied_credits.value = str("applied_credits")

if adjustment_credit_notes_CNo != "NA":
    adjustment_credit_notes = my_sheet.cell(row=1, column=adjustment_credit_notes_CNo)
    adjustment_credit_notes.value = str("adjustment_credit_notes")

if issued_credit_notes_CNo != "NA":
    issued_credit_notes = my_sheet.cell(row=1, column=issued_credit_notes_CNo)
    issued_credit_notes.value = str("issued_credit_notes")

if linked_orders_CNo != "NA":
    linked_orders = my_sheet.cell(row=1, column=linked_orders_CNo)
    linked_orders.value = str("linked_orders")

if billing_address_CNo != "NA":
    billing_address = my_sheet.cell(row=1, column=billing_address_CNo)
    billing_address.value = str("billing_address")

if billing_address_first_name_CNo != "NA":
    billing_address_first_name = my_sheet.cell(row=1, column=billing_address_first_name_CNo)
    billing_address_first_name.value = str("billing_address_first_name")

if billing_address_last_name_CNo != "NA":
    billing_address_last_name = my_sheet.cell(row=1, column=billing_address_last_name_CNo)
    billing_address_last_name.value = str("billing_address_last_name")

if billing_address_email_CNo != "NA":
    billing_address_email = my_sheet.cell(row=1, column=billing_address_email_CNo)
    billing_address_email.value = str("billing_address_email")

if billing_address_company_CNo != "NA":
    billing_address_company = my_sheet.cell(row=1, column=billing_address_company_CNo)
    billing_address_company.value = str("billing_address_company")

if billing_address_phone_CNo != "NA":
    billing_address_phone = my_sheet.cell(row=1, column=billing_address_phone_CNo)
    billing_address_phone.value = str("billing_address_phone")

if billing_address_state_CNo != "NA":
    billing_address_state = my_sheet.cell(row=1, column=billing_address_state_CNo)
    billing_address_state.value = str("billing_address_state")

if billing_address_line1_CNo != "NA":
    billing_address_line1 = my_sheet.cell(row=1, column=billing_address_line1_CNo)
    billing_address_line1.value = str("billing_address_line1")

if billing_address_line2_CNo != "NA":
    billing_address_line2 = my_sheet.cell(row=1, column=billing_address_line2_CNo)
    billing_address_line2.value = str("billing_address_line2")

if billing_address_line3_CNo != "NA":
    billing_address_line3 = my_sheet.cell(row=1, column=billing_address_line3_CNo)
    billing_address_line3.value = str("billing_address_line3")

if billing_address_city_CNo != "NA":
    billing_address_city = my_sheet.cell(row=1, column=billing_address_city_CNo)
    billing_address_city.value = str("billing_address_city")

if billing_address_state_code_CNo != "NA":
    billing_address_state_code = my_sheet.cell(row=1, column=billing_address_state_code_CNo)
    billing_address_state_code.value = str("billing_address_state_code")

if billing_address_country_CNo != "NA":
    billing_address_country = my_sheet.cell(row=1, column=billing_address_country_CNo)
    billing_address_country.value = str("billing_address_country")

if billing_address_zip_CNo != "NA":
    billing_address_zip = my_sheet.cell(row=1, column=billing_address_zip_CNo)
    billing_address_zip.value = str("billing_address_zip")

if billing_address_validation_status_CNo != "NA":
    billing_address_validation_status = my_sheet.cell(row=1, column=billing_address_validation_status_CNo)
    billing_address_validation_status.value = str("billing_address_validation_status")

if billing_address_object_CNo != "NA":
    billing_address_object = my_sheet.cell(row=1, column=billing_address_object_CNo)
    billing_address_object.value = str("billing_address_object")

if shipping_address_first_name_CNo != "NA":
    shipping_address_first_name = my_sheet.cell(row=1, column=shipping_address_first_name_CNo)
    shipping_address_first_name.value = str("shipping_address_first_name")
if shipping_address_last_name_CNo != "NA":
    shipping_address_last_name = my_sheet.cell(row=1, column=shipping_address_last_name_CNo)
    shipping_address_last_name.value = str("shipping_address_last_name")
if shipping_address_email_CNo != "NA":
    shipping_address_email = my_sheet.cell(row=1, column=shipping_address_email_CNo)
    shipping_address_email.value = str("shipping_address_email")
if shipping_address_company_CNo != "NA":
    shipping_address_company = my_sheet.cell(row=1, column=shipping_address_company_CNo)
    shipping_address_company.value = str("shipping_address_company")
if shipping_address_phone_CNo != "NA":
    shipping_address_phone = my_sheet.cell(row=1, column=shipping_address_phone_CNo)
    shipping_address_phone.value = str("shipping_address_phone")
if shipping_address_line1_CNo != "NA":
    shipping_address_line1 = my_sheet.cell(row=1, column=shipping_address_line1_CNo)
    shipping_address_line1.value = str("shipping_address_line1")
if shipping_address_line2_CNo != "NA":
    shipping_address_line2 = my_sheet.cell(row=1, column=shipping_address_line2_CNo)
    shipping_address_line2.value = str("shipping_address_line2")
if shipping_address_line3_CNo != "NA":
    shipping_address_line3 = my_sheet.cell(row=1, column=shipping_address_line3_CNo)
    shipping_address_line3.value = str("shipping_address_line3")
if shipping_address_city_CNo != "NA":
    shipping_address_city = my_sheet.cell(row=1, column=shipping_address_city_CNo)
    shipping_address_city.value = str("shipping_address_city")
if shipping_address_state_code_CNo != "NA":
    shipping_address_state_code = my_sheet.cell(row=1, column=shipping_address_state_code_CNo)
    shipping_address_state_code.value = str("shipping_address_state_code")
if shipping_address_state_CNo != "NA":
    shipping_address_state = my_sheet.cell(row=1, column=shipping_address_state_CNo)
    shipping_address_state.value = str("shipping_address_state")
if shipping_address_zip_CNo != "NA":
    shipping_address_zip = my_sheet.cell(row=1, column=shipping_address_zip_CNo)
    shipping_address_zip.value = str("shipping_address_zip")
if shipping_address_country_CNo != "NA":
    shipping_address_country = my_sheet.cell(row=1, column=shipping_address_country_CNo)
    shipping_address_country.value = str("shipping_address_country")
if shipping_address_validation_status_CNo != "NA":
    shipping_address_validation_status = my_sheet.cell(row=1, column=shipping_address_validation_status_CNo)
    shipping_address_validation_status.value = str("shipping_address_validation_status")

# line_items_1
if line_items_1_id_CNo != "NA":
    line_items_1_id = my_sheet.cell(row=1, column=line_items_1_id_CNo)
    line_items_1_id.value = str("line_items_1_id")
if line_items_1_description_CNo != "NA":
    line_items_1_description = my_sheet.cell(row=1, column=line_items_1_description_CNo)
    line_items_1_description.value = str("line_items_1_description")
if line_items_1_date_from_CNo != "NA":
    line_items_1_date_from = my_sheet.cell(row=1, column=line_items_1_date_from_CNo)
    line_items_1_date_from.value = str("line_items_1_date_from")
if line_items_1_date_to_CNo != "NA":
    line_items_1_date_to = my_sheet.cell(row=1, column=line_items_1_date_to_CNo)
    line_items_1_date_to.value = str("line_items_1_date_to")
if line_items_1_unit_amount_CNo != "NA":
    line_items_1_unit_amount = my_sheet.cell(row=1, column=line_items_1_unit_amount_CNo)
    line_items_1_unit_amount.value = str("line_items_1_unit_amount")
if line_items_1_quantity_CNo != "NA":
    line_items_1_quantity = my_sheet.cell(row=1, column=line_items_1_quantity_CNo)
    line_items_1_quantity.value = str("line_items_1_quantity")
if line_items_1_amount_CNo != "NA":
    line_items_1_amount = my_sheet.cell(row=1, column=line_items_1_amount_CNo)
    line_items_1_amount.value = str("line_items_1_amount")
if line_items_1_entity_type_CNo != "NA":
    line_items_1_entity_type = my_sheet.cell(row=1, column=line_items_1_entity_type_CNo)
    line_items_1_entity_type.value = str("line_items_1_entity_type")
if line_items_1_entity_id_CNo != "NA":
    line_items_1_entity_id = my_sheet.cell(row=1, column=line_items_1_entity_id_CNo)
    line_items_1_entity_id.value = str("line_items_1_entity_id")
if line_items_1_tax1_name_CNo != "NA":
    line_items_1_tax1_name = my_sheet.cell(row=1, column=line_items_1_tax1_name_CNo)
    line_items_1_tax1_name.value = str("line_items_1_tax1_name")
if line_items_1_tax1_amount_CNo != "NA":
    line_items_1_tax1_amount = my_sheet.cell(row=1, column=line_items_1_tax1_amount_CNo)
    line_items_1_tax1_amount.value = str("line_items_1_tax1_amount")

# line_items_2
if line_items_2_id_CNo != "NA":
    line_items_2_id = my_sheet.cell(row=1, column=line_items_2_id_CNo)
    line_items_2_id.value = str("line_items_2_id")
if line_items_2_description_CNo != "NA":
    line_items_2_description = my_sheet.cell(row=1, column=line_items_2_description_CNo)
    line_items_2_description.value = str("line_items_2_description")
if line_items_2_date_from_CNo != "NA":
    line_items_2_date_from = my_sheet.cell(row=1, column=line_items_2_date_from_CNo)
    line_items_2_date_from.value = str("line_items_2_date_from")
if line_items_2_date_to_CNo != "NA":
    line_items_2_date_to = my_sheet.cell(row=1, column=line_items_2_date_to_CNo)
    line_items_2_date_to.value = str("line_items_2_date_to")
if line_items_2_unit_amount_CNo != "NA":
    line_items_2_unit_amount = my_sheet.cell(row=1, column=line_items_2_unit_amount_CNo)
    line_items_2_unit_amount.value = str("line_items_2_unit_amount")
if line_items_2_quantity_CNo != "NA":
    line_items_2_quantity = my_sheet.cell(row=1, column=line_items_2_quantity_CNo)
    line_items_2_quantity.value = str("line_items_2_quantity")
if line_items_2_amount_CNo != "NA":
    line_items_2_amount = my_sheet.cell(row=1, column=line_items_2_amount_CNo)
    line_items_2_amount.value = str("line_items_2_amount")
if line_items_2_entity_type_CNo != "NA":
    line_items_2_entity_type = my_sheet.cell(row=1, column=line_items_2_entity_type_CNo)
    line_items_2_entity_type.value = str("line_items_2_entity_type")
if line_items_2_entity_id_CNo != "NA":
    line_items_2_entity_id = my_sheet.cell(row=1, column=line_items_2_entity_id_CNo)
    line_items_2_entity_id.value = str("line_items_2_entity_id")
if line_items_2_tax1_name_CNo != "NA":
    line_items_2_tax1_name = my_sheet.cell(row=1, column=line_items_2_tax1_name_CNo)
    line_items_2_tax1_name.value = str("line_items_2_tax1_name")
if line_items_2_tax1_amount_CNo != "NA":
    line_items_2_tax1_amount = my_sheet.cell(row=1, column=line_items_2_tax1_amount_CNo)
    line_items_2_tax1_amount.value = str("line_items_2_tax1_amount")

# line_items_3
if line_items_3_id_CNo != "NA":
    line_items_3_id = my_sheet.cell(row=1, column=line_items_3_id_CNo)
    line_items_3_id.value = str("line_items_3_id")
if line_items_3_description_CNo != "NA":
    line_items_3_description = my_sheet.cell(row=1, column=line_items_3_description_CNo)
    line_items_3_description.value = str("line_items_3_description")
if line_items_3_date_from_CNo != "NA":
    line_items_3_date_from = my_sheet.cell(row=1, column=line_items_3_date_from_CNo)
    line_items_3_date_from.value = str("line_items_3_date_from")
if line_items_3_date_to_CNo != "NA":
    line_items_3_date_to = my_sheet.cell(row=1, column=line_items_3_date_to_CNo)
    line_items_3_date_to.value = str("line_items_3_date_to")
if line_items_3_unit_amount_CNo != "NA":
    line_items_3_unit_amount = my_sheet.cell(row=1, column=line_items_3_unit_amount_CNo)
    line_items_3_unit_amount.value = str("line_items_3_unit_amount")
if line_items_3_quantity_CNo != "NA":
    line_items_3_quantity = my_sheet.cell(row=1, column=line_items_3_quantity_CNo)
    line_items_3_quantity.value = str("line_items_3_quantity")
if line_items_3_amount_CNo != "NA":
    line_items_3_amount = my_sheet.cell(row=1, column=line_items_3_amount_CNo)
    line_items_3_amount.value = str("line_items_3_amount")
if line_items_3_entity_type_CNo != "NA":
    line_items_3_entity_type = my_sheet.cell(row=1, column=line_items_3_entity_type_CNo)
    line_items_3_entity_type.value = str("line_items_3_entity_type")
if line_items_3_entity_id_CNo != "NA":
    line_items_3_entity_id = my_sheet.cell(row=1, column=line_items_3_entity_id_CNo)
    line_items_3_entity_id.value = str("line_items_3_entity_id")
if line_items_3_tax1_name_CNo != "NA":
    line_items_3_tax1_name = my_sheet.cell(row=1, column=line_items_3_tax1_name_CNo)
    line_items_3_tax1_name.value = str("line_items_3_tax1_name")
if line_items_3_tax1_amount_CNo != "NA":
    line_items_3_tax1_amount = my_sheet.cell(row=1, column=line_items_3_tax1_amount_CNo)
    line_items_3_tax1_amount.value = str("line_items_3_tax1_amount")

# line_items_4
if line_items_4_id_CNo != "NA":
    line_items_4_id = my_sheet.cell(row=1, column=line_items_4_id_CNo)
    line_items_4_id.value = str("line_items_4_id")
if line_items_4_description_CNo != "NA":
    line_items_4_description = my_sheet.cell(row=1, column=line_items_4_description_CNo)
    line_items_4_description.value = str("line_items_4_description")
if line_items_4_date_from_CNo != "NA":
    line_items_4_date_from = my_sheet.cell(row=1, column=line_items_4_date_from_CNo)
    line_items_4_date_from.value = str("line_items_4_date_from")
if line_items_4_date_to_CNo != "NA":
    line_items_4_date_to = my_sheet.cell(row=1, column=line_items_4_date_to_CNo)
    line_items_4_date_to.value = str("line_items_4_date_to")
if line_items_4_unit_amount_CNo != "NA":
    line_items_4_unit_amount = my_sheet.cell(row=1, column=line_items_4_unit_amount_CNo)
    line_items_4_unit_amount.value = str("line_items_4_unit_amount")
if line_items_4_quantity_CNo != "NA":
    line_items_4_quantity = my_sheet.cell(row=1, column=line_items_4_quantity_CNo)
    line_items_4_quantity.value = str("line_items_4_quantity")
if line_items_4_amount_CNo != "NA":
    line_items_4_amount = my_sheet.cell(row=1, column=line_items_4_amount_CNo)
    line_items_4_amount.value = str("line_items_4_amount")
if line_items_4_entity_type_CNo != "NA":
    line_items_4_entity_type = my_sheet.cell(row=1, column=line_items_4_entity_type_CNo)
    line_items_4_entity_type.value = str("line_items_4_entity_type")
if line_items_4_entity_id_CNo != "NA":
    line_items_4_entity_id = my_sheet.cell(row=1, column=line_items_4_entity_id_CNo)
    line_items_4_entity_id.value = str("line_items_4_entity_id")
if line_items_4_tax1_name_CNo != "NA":
    line_items_4_tax1_name = my_sheet.cell(row=1, column=line_items_4_tax1_name_CNo)
    line_items_4_tax1_name.value = str("line_items_4_tax1_name")
if line_items_4_tax1_amount_CNo != "NA":
    line_items_4_tax1_amount = my_sheet.cell(row=1, column=line_items_4_tax1_amount_CNo)
    line_items_4_tax1_amount.value = str("line_items_4_tax1_amount")

# line_items_5
if line_items_5_id_CNo != "NA":
    line_items_5_id = my_sheet.cell(row=1, column=line_items_5_id_CNo)
    line_items_5_id.value = str("line_items_5_id")
if line_items_5_description_CNo != "NA":
    line_items_5_description = my_sheet.cell(row=1, column=line_items_5_description_CNo)
    line_items_5_description.value = str("line_items_5_description")
if line_items_5_date_from_CNo != "NA":
    line_items_5_date_from = my_sheet.cell(row=1, column=line_items_5_date_from_CNo)
    line_items_5_date_from.value = str("line_items_5_date_from")
if line_items_5_date_to_CNo != "NA":
    line_items_5_date_to = my_sheet.cell(row=1, column=line_items_5_date_to_CNo)
    line_items_5_date_to.value = str("line_items_5_date_to")
if line_items_5_unit_amount_CNo != "NA":
    line_items_5_unit_amount = my_sheet.cell(row=1, column=line_items_5_unit_amount_CNo)
    line_items_5_unit_amount.value = str("line_items_5_unit_amount")
if line_items_5_quantity_CNo != "NA":
    line_items_5_quantity = my_sheet.cell(row=1, column=line_items_5_quantity_CNo)
    line_items_5_quantity.value = str("line_items_5_quantity")
if line_items_5_amount_CNo != "NA":
    line_items_5_amount = my_sheet.cell(row=1, column=line_items_5_amount_CNo)
    line_items_5_amount.value = str("line_items_5_amount")
if line_items_5_entity_type_CNo != "NA":
    line_items_5_entity_type = my_sheet.cell(row=1, column=line_items_5_entity_type_CNo)
    line_items_5_entity_type.value = str("line_items_5_entity_type")
if line_items_5_entity_id_CNo != "NA":
    line_items_5_entity_id = my_sheet.cell(row=1, column=line_items_5_entity_id_CNo)
    line_items_5_entity_id.value = str("line_items_5_entity_id")
if line_items_5_tax1_name_CNo != "NA":
    line_items_5_tax1_name = my_sheet.cell(row=1, column=line_items_5_tax1_name_CNo)
    line_items_5_tax1_name.value = str("line_items_5_tax1_name")
if line_items_5_tax1_amount_CNo != "NA":
    line_items_5_tax1_amount = my_sheet.cell(row=1, column=line_items_5_tax1_amount_CNo)
    line_items_5_tax1_amount.value = str("line_items_5_tax1_amount")

# line_items_6
if line_items_6_id_CNo != "NA":
    line_items_6_id = my_sheet.cell(row=1, column=line_items_6_id_CNo)
    line_items_6_id.value = str("line_items_6_id")
if line_items_6_description_CNo != "NA":
    line_items_6_description = my_sheet.cell(row=1, column=line_items_6_description_CNo)
    line_items_6_description.value = str("line_items_6_description")
if line_items_6_date_from_CNo != "NA":
    line_items_6_date_from = my_sheet.cell(row=1, column=line_items_6_date_from_CNo)
    line_items_6_date_from.value = str("line_items_6_date_from")
if line_items_6_date_to_CNo != "NA":
    line_items_6_date_to = my_sheet.cell(row=1, column=line_items_6_date_to_CNo)
    line_items_6_date_to.value = str("line_items_6_date_to")
if line_items_6_unit_amount_CNo != "NA":
    line_items_6_unit_amount = my_sheet.cell(row=1, column=line_items_6_unit_amount_CNo)
    line_items_6_unit_amount.value = str("line_items_6_unit_amount")
if line_items_6_quantity_CNo != "NA":
    line_items_6_quantity = my_sheet.cell(row=1, column=line_items_6_quantity_CNo)
    line_items_6_quantity.value = str("line_items_6_quantity")
if line_items_6_amount_CNo != "NA":
    line_items_6_amount = my_sheet.cell(row=1, column=line_items_6_amount_CNo)
    line_items_6_amount.value = str("line_items_6_amount")
if line_items_6_entity_type_CNo != "NA":
    line_items_6_entity_type = my_sheet.cell(row=1, column=line_items_6_entity_type_CNo)
    line_items_6_entity_type.value = str("line_items_6_entity_type")
if line_items_6_entity_id_CNo != "NA":
    line_items_6_entity_id = my_sheet.cell(row=1, column=line_items_6_entity_id_CNo)
    line_items_6_entity_id.value = str("line_items_6_entity_id")
if line_items_6_tax1_name_CNo != "NA":
    line_items_6_tax1_name = my_sheet.cell(row=1, column=line_items_6_tax1_name_CNo)
    line_items_6_tax1_name.value = str("line_items_6_tax1_name")
if line_items_6_tax1_amount_CNo != "NA":
    line_items_6_tax1_amount = my_sheet.cell(row=1, column=line_items_6_tax1_amount_CNo)
    line_items_6_tax1_amount.value = str("line_items_6_tax1_amount")

# line_items_7
if line_items_7_id_CNo != "NA":
    line_items_7_id = my_sheet.cell(row=1, column=line_items_7_id_CNo)
    line_items_7_id.value = str("line_items_7_id")
if line_items_7_description_CNo != "NA":
    line_items_7_description = my_sheet.cell(row=1, column=line_items_7_description_CNo)
    line_items_7_description.value = str("line_items_7_description")
if line_items_7_date_from_CNo != "NA":
    line_items_7_date_from = my_sheet.cell(row=1, column=line_items_7_date_from_CNo)
    line_items_7_date_from.value = str("line_items_7_date_from")
if line_items_7_date_to_CNo != "NA":
    line_items_7_date_to = my_sheet.cell(row=1, column=line_items_7_date_to_CNo)
    line_items_7_date_to.value = str("line_items_7_date_to")
if line_items_7_unit_amount_CNo != "NA":
    line_items_7_unit_amount = my_sheet.cell(row=1, column=line_items_7_unit_amount_CNo)
    line_items_7_unit_amount.value = str("line_items_7_unit_amount")
if line_items_7_quantity_CNo != "NA":
    line_items_7_quantity = my_sheet.cell(row=1, column=line_items_7_quantity_CNo)
    line_items_7_quantity.value = str("line_items_7_quantity")
if line_items_7_amount_CNo != "NA":
    line_items_7_amount = my_sheet.cell(row=1, column=line_items_7_amount_CNo)
    line_items_7_amount.value = str("line_items_7_amount")
if line_items_7_entity_type_CNo != "NA":
    line_items_7_entity_type = my_sheet.cell(row=1, column=line_items_7_entity_type_CNo)
    line_items_7_entity_type.value = str("line_items_7_entity_type")
if line_items_7_entity_id_CNo != "NA":
    line_items_7_entity_id = my_sheet.cell(row=1, column=line_items_7_entity_id_CNo)
    line_items_7_entity_id.value = str("line_items_7_entity_id")
if line_items_7_tax1_name_CNo != "NA":
    line_items_7_tax1_name = my_sheet.cell(row=1, column=line_items_7_tax1_name_CNo)
    line_items_7_tax1_name.value = str("line_items_7_tax1_name")
if line_items_7_tax1_amount_CNo != "NA":
    line_items_7_tax1_amount = my_sheet.cell(row=1, column=line_items_7_tax1_amount_CNo)
    line_items_7_tax1_amount.value = str("line_items_7_tax1_amount")


# parse each element/field from response and return

def get_id(i):
    try:
        id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.id")
    except:
        print("in status exception")
    return id


def get_customer_id(i):
    try:
        customer_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.customer_id")
    except:
        print("in status exception")
    return customer_id


def get_subscription_id(i):
    try:
        subscription_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.subscription_id")
    except:
        print("in status exception")
    return subscription_id


def get_invoice_po_number(i):
    try:
        invoice_po_number = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.po_number")
    except:
        print("in status exception")
    return invoice_po_number


def get_tax_override_reason(i):
    try:
        tax_override_reason = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.tax_override_reason")
    except:
        print("in status exception")
    return tax_override_reason


def get_recurring(i):
    try:
        recurring = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.recurring")
    except:
        print("in status exception")
    return recurring


def get_status(i):
    try:
        status = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.status")
    except:
        print("in status exception")
    return status


def get_vat_number(i):
    try:
        vat_number = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.vat_number")
    except:
        print("in status exception")
    return vat_number


def get_price_type(i):
    try:
        price_type = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.price_type")
    except:
        print("in status exception")
    return price_type


def get_date(i):
    try:
        date = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.date")
    except:
        print("in status exception")
    return date


def get_due_date(i):
    try:
        due_date = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.due_date")
    except:
        print("in status exception")
    return due_date


def get_net_term_days(i):
    try:
        net_term_days = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.net_term_days")
    except:
        print("in status exception")
    return net_term_days


def get_use_for_proration(i):
    try:
        use_for_proration = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.use_for_proration")
    except:
        print("in status exception")
    return use_for_proration


def get_exchange_rate(i):
    try:
        exchange_rate = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.exchange_rate")
    except:
        print("in status exception")
    return exchange_rate


def get_total(i):
    try:
        total = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.total")
    except:
        print("in status exception")
    return total


def get_amount_paid(i):
    try:
        amount_paid = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.amount_paid")
    except:
        print("in status exception")
    return amount_paid


def get_amount_adjusted(i):
    try:
        amount_adjusted = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.amount_adjusted")
    except:
        print("in status exception")
    return amount_adjusted


def get_write_off_amount(i):
    try:
        write_off_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.write_off_amount")
    except:
        print("in status exception")
    return write_off_amount


def get_credits_applied(i):
    try:
        credits_applied = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.credits_applied")
    except:
        print("in status exception")
    return credits_applied


def get_amount_due(i):
    try:
        amount_due = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.amount_due")
    except:
        print("in status exception")
    return amount_due


def get_dunning_status(i):
    try:
        dunning_status = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.dunning_status")
    except:
        print("in status exception")
    return dunning_status


def get_next_retry_at(i):
    try:
        next_retry_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.next_retry_at")
    except:
        print("in status exception")
    return next_retry_at


def get_updated_at(i):
    try:
        updated_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.updated_at")
    except:
        print("in status exception")
    return updated_at


def get_resource_version(i):
    try:
        resource_version = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.resource_version")
    except:
        print("in status exception")
    return resource_version


def get_deleted(i):
    try:
        deleted = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.deleted")
    except:
        print("in status exception")
    return deleted


def get_object(i):
    try:
        object = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.object")
    except:
        print("in status exception")
    return object


def get_first_invoice(i):
    try:
        first_invoice = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.first_invoice")
    except:
        print("in status exception")
    return first_invoice


def get_amount_to_collect(i):
    try:
        amount_to_collect = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.amount_to_collect")
    except:
        print("in status exception")
    return amount_to_collect


def get_round_off_amount(i):
    try:
        round_off_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.round_off_amount")
    except:
        print("in status exception")
    return round_off_amount


def get_has_advance_charges(i):
    try:
        has_advance_charges = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.has_advance_charges")
    except:
        print("in status exception")
    return has_advance_charges


def get_currency_code(i):
    try:
        currency_code = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.currency_code")
    except:
        print("in status exception")
    return currency_code


def get_base_currency_code(i):
    try:
        base_currency_code = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.base_currency_code")
    except:
        print("in status exception")
    return base_currency_code


def get_is_gifted(i):
    try:
        is_gifted = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.is_gifted")
    except:
        print("in status exception")
    return is_gifted


def get_term_finalized(i):
    try:
        term_finalized = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.term_finalized")
    except:
        print("in status exception")
    return term_finalized


def get_is_digital(i):
    try:
        is_digital = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.is_digital")
    except:
        print("in status exception")
    return is_digital


def get_tax(i):
    try:
        tax = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.tax")
    except:
        print("in status exception")
    return tax


def get_line_items(i):
    try:
        line_items = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items")
    except:
        print("in status exception")
    return line_items


def get_line_items_0_(i):
    try:
        line_items_0_ = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0]")
    except:
        print("in status exception")
    return line_items_0_


def get_line_items_0_id(i):
    try:
        line_items_0_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].id")
    except:
        print("in status exception")
    return line_items_0_id


def get_line_items_0_date_from(i):
    try:
        line_items_0_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].date_from")
    except:
        print("in status exception")
    return line_items_0_date_from


def get_line_items_0_date_to(i):
    try:
        line_items_0_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].date_to")
    except:
        print("in status exception")
    return line_items_0_date_to


def get_line_items_0_unit_amount(i):
    try:
        line_items_0_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[0].unit_amount")
    except:
        print("in status exception")
    return line_items_0_unit_amount


def get_line_items_0_quantity(i):
    try:
        line_items_0_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].quantity")
    except:
        print("in status exception")
    return line_items_0_quantity


def get_line_items_0_amount(i):
    try:
        line_items_0_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].amount")
    except:
        print("in status exception")
    return line_items_0_amount


def get_line_items_0_pricing_model(i):
    try:
        line_items_0_pricing_model = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].invoice.line_items[0].pricing_model")
    except:
        print("in status exception")
    return line_items_0_pricing_model


def get_line_items_0_is_taxed(i):
    try:
        line_items_0_is_taxed = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].is_taxed")
    except:
        print("in status exception")
    return line_items_0_is_taxed


def get_line_items_0_tax_amount(i):
    try:
        line_items_0_tax_amount = jsonpath.jsonpath(jsonpathres,
                                                    "list[" + str(i) + "].invoice.line_items[0].tax_amount")
    except:
        print("in status exception")
    return line_items_0_tax_amount


def get_line_items_0_tax_rate(i):
    try:
        line_items_0_tax_rate = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].tax_rate")
    except:
        print("in status exception")
    return line_items_0_tax_rate


def get_line_items_0_object(i):
    try:
        line_items_0_object = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].object")
    except:
        print("in status exception")
    return line_items_0_object


def get_line_items_0_subscription_id(i):
    try:
        line_items_0_subscription_id = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].invoice.line_items[0].subscription_id")
    except:
        print("in status exception")
    return line_items_0_subscription_id


def get_line_items_0_customer_id(i):
    try:
        line_items_0_customer_id = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[0].customer_id")
    except:
        print("in status exception")
    return line_items_0_customer_id


def get_line_items_0_description(i):
    try:
        line_items_0_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[0].description")
    except:
        print("in status exception")
    return line_items_0_description


def get_line_items_0_entity_type(i):
    try:
        line_items_0_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[0].entity_type")
    except:
        print("in status exception")
    return line_items_0_entity_type


def get_line_items_0_entity_id(i):
    try:
        line_items_0_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[0].entity_id")
    except:
        print("in status exception")
    return line_items_0_entity_id


def get_line_items_0_discount_amount(i):
    try:
        line_items_0_discount_amount = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].invoice.line_items[0].discount_amount")
    except:
        print("in status exception")
    return line_items_0_discount_amount


def get_line_items_0_item_level_discount_amount(i):
    try:
        line_items_0_item_level_discount_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_items[0].item_level_discount_amount")
    except:
        print("in status exception")
    return line_items_0_item_level_discount_amount


def get_line_items_0_tax1_name(i):
    try:
        line_items_0_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[0].tax_name")
    except:
        print("in status exception")
    return line_items_0_tax1_name


def get_line_items_0_tax1_amount(i):
    try:
        line_items_0_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[0].tax_amount")
    except:
        print("in status exception")
    return line_items_0_tax1_amount


def get_taxes(i):
    try:
        taxes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.taxes")
    except:
        print("in status exception")
    return taxes


def get_taxes_0_(i):
    try:
        taxes_0_ = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.taxes[0]")
    except:
        print("in status exception")
    return taxes_0_


def get_taxes_0_object(i):
    try:
        taxes_0_object = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.taxes[0].object")
    except:
        print("in status exception")
    return taxes_0_object


def get_taxes_0_name(i):
    try:
        taxes_0_name = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.taxes[0].name")
    except:
        print("in status exception")
    return taxes_0_name


def get_taxes_0_rate(i):
    try:
        taxes_0_rate = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.taxes[0].rate")
    except:
        print("in status exception")
    return taxes_0_rate


def get_taxes_0_description(i):
    try:
        taxes_0_description = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.taxes[0].description")
    except:
        print("in status exception")
    return taxes_0_description


def get_taxes_0_amount(i):
    try:
        taxes_0_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.taxes[0].amount")
    except:
        print("in status exception")
    return taxes_0_amount


def get_discounts_0_entity_type(i):
    try:
        discounts_0_entity_type = jsonpath.jsonpath(jsonpathres,
                                                    "list[" + str(i) + "].invoice.discounts[0].entity_type")
    except:
        print("in status exception")
    return discounts_0_entity_type


def get_discounts_0_entity_id(i):
    try:
        discounts_0_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.discounts[0].entity_id")
    except:
        print("in status exception")
    return discounts_0_entity_id


def get_discounts_0_amount(i):
    try:
        discounts_0_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.discounts[0].amount")
    except:
        print("in status exception")
    return discounts_0_amount


def get_payments_0_amount(i):
    try:
        payments_0_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.linked_payments[0].txn_amount")
    except:
        print("in status exception")
    return payments_0_amount


def get_payments_0_payment_method(i):
    try:
        payments_0_payment_method = jsonpath.jsonpath(jsonpathres,
                                                      "list[" + str(i) + "].invoice.linked_payments[0].payment_method")
    except:
        print("in status exception - payment method")
    return payments_0_payment_method


def get_payments_0_date(i):
    try:
        payments_0_date = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.linked_payments[0].txn_date")
    except:
        print("in status exception")
    return payments_0_date


def get_line_item_taxes(i):
    try:
        line_item_taxes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_item_taxes")
    except:
        print("in status exception")
    return line_item_taxes


def get_line_item_taxes_0_(i):
    try:
        line_item_taxes_0_ = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_item_taxes[0]")
    except:
        print("in status exception")
    return line_item_taxes_0_


def get_line_item_taxes_0_tax_name(i):
    try:
        line_item_taxes_0_tax_name = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].invoice.line_item_taxes[0].tax_name")
    except:
        print("in status exception")
    return line_item_taxes_0_tax_name


def get_line_item_taxes_0_tax_rate(i):
    try:
        line_item_taxes_0_tax_rate = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].invoice.line_item_taxes[0].tax_rate")
    except:
        print("in status exception")
    return line_item_taxes_0_tax_rate


def get_line_item_taxes_0_tax_juris_type(i):
    try:
        line_item_taxes_0_tax_juris_type = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_item_taxes[0].tax_juris_type")
    except:
        print("in status exception")
    return line_item_taxes_0_tax_juris_type


def get_line_item_taxes_0_tax_juris_name(i):
    try:
        line_item_taxes_0_tax_juris_name = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_item_taxes[0].tax_juris_name")
    except:
        print("in status exception")
    return line_item_taxes_0_tax_juris_name


def get_line_item_taxes_0_tax_juris_code(i):
    try:
        line_item_taxes_0_tax_juris_code = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_item_taxes[0].tax_juris_code")
    except:
        print("in status exception")
    return line_item_taxes_0_tax_juris_code


def get_line_item_taxes_0_object(i):
    try:
        line_item_taxes_0_object = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[0].object")
    except:
        print("in status exception")
    return line_item_taxes_0_object


def get_line_item_taxes_0_line_item_id(i):
    try:
        line_item_taxes_0_line_item_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_item_taxes[0].line_item_id")
    except:
        print("in status exception")
    return line_item_taxes_0_line_item_id


def get_line_item_taxes_0_tax_amount(i):
    try:
        line_item_taxes_0_tax_amount = jsonpath.jsonpath(jsonpathres,
                                                         "list[" + str(i) + "].invoice.line_item_taxes[0].tax_amount")
    except:
        print("in status exception")
    return line_item_taxes_0_tax_amount


def get_line_item_taxes_0_is_partial_tax_applied(i):
    try:
        line_item_taxes_0_is_partial_tax_applied = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_item_taxes[0].is_partial_tax_applied")
    except:
        print("in status exception")
    return line_item_taxes_0_is_partial_tax_applied


def get_line_item_taxes_0_taxable_amount(i):
    try:
        line_item_taxes_0_taxable_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_item_taxes[0].taxable_amount")
    except:
        print("in status exception")
    return line_item_taxes_0_taxable_amount


def get_line_item_taxes_0_is_non_compliance_tax(i):
    try:
        line_item_taxes_0_is_non_compliance_tax = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.line_item_taxes[0].is_non_compliance_tax")
    except:
        print("in status exception")
    return line_item_taxes_0_is_non_compliance_tax


def get_sub_total(i):
    try:
        sub_total = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.sub_total")
    except:
        print("in status exception")
    return sub_total


def get_linked_payments(i):
    try:
        linked_payments = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.linked_payments")
    except:
        print("in status exception")
    return linked_payments


def get_dunning_attempts(i):
    try:
        dunning_attempts = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.dunning_attempts")
    except:
        print("in status exception")
    return dunning_attempts


def get_applied_credits(i):
    try:
        applied_credits = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.applied_credits")
    except:
        print("in status exception")
    return applied_credits


def get_adjustment_credit_notes(i):
    try:
        adjustment_credit_notes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.adjustment_credit_notes")
    except:
        print("in status exception")
    return adjustment_credit_notes


def get_issued_credit_notes(i):
    try:
        issued_credit_notes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.issued_credit_notes")
    except:
        print("in status exception")
    return issued_credit_notes


def get_linked_orders(i):
    try:
        linked_orders = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.linked_orders")
    except:
        print("in status exception")
    return linked_orders


def get_billing_address(i):
    try:
        billing_address = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address")
    except:
        print("in status exception")
    return billing_address


def get_billing_address_first_name(i):
    try:
        billing_address_first_name = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].invoice.billing_address.first_name")
    except:
        print("in status exception")
    return billing_address_first_name


def get_billing_address_last_name(i):
    try:
        billing_address_last_name = jsonpath.jsonpath(jsonpathres,
                                                      "list[" + str(i) + "].invoice.billing_address.last_name")
    except:
        print("in status exception")
    return billing_address_last_name


def get_billing_address_email(i):
    try:
        billing_address_email = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.email")
    except:
        print("in status exception")
    return billing_address_email


def get_billing_address_company(i):
    try:
        billing_address_company = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.company")
    except:
        print("in status exception")
    return billing_address_company


def get_billing_address_phone(i):
    try:
        billing_address_phone = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.phone")
    except:
        print("in status exception")
    return billing_address_phone


def get_billing_address_state(i):
    try:
        billing_address_state = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.state")
    except:
        print("in status exception")
    return billing_address_state


def get_billing_address_line1(i):
    try:
        billing_address_line1 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.line1")
    except:
        print("in status exception")
    return billing_address_line1


def get_billing_address_line2(i):
    try:
        billing_address_line2 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.line2")
    except:
        print("in status exception")
    return billing_address_line2


def get_billing_address_line3(i):
    try:
        billing_address_line3 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.line3")
    except:
        print("in status exception")
    return billing_address_line3


def get_billing_address_city(i):
    try:
        billing_address_city = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.city")
    except:
        print("in status exception")
    return billing_address_city


def get_billing_address_state_code(i):
    try:
        billing_address_state_code = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].invoice.billing_address.state_code")
    except:
        print("in status exception")
    return billing_address_state_code


def get_billing_address_country(i):
    try:
        billing_address_country = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.country")
    except:
        print("in status exception")
    return billing_address_country


def get_billing_address_zip(i):
    try:
        billing_address_zip = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.zip")
    except:
        print("in status exception")
    return billing_address_zip


def get_billing_address_validation_status(i):
    try:
        billing_address_validation_status = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.billing_address.validation_status")
    except:
        print("in status exception")
    return billing_address_validation_status


def get_billing_address_object(i):
    try:
        billing_address_object = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.billing_address.object")
    except:
        print("in status exception")
    return billing_address_object


def get_shipping_address_first_name(i):
    try:
        shipping_address_first_name = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].invoice.shipping_address.first_name")
    except:
        print("in status exception")
    return shipping_address_first_name


def get_shipping_address_last_name(i):
    try:
        shipping_address_last_name = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].invoice.shipping_address.last_name")
    except:
        print("in status exception")
    return shipping_address_last_name


def get_shipping_address_email(i):
    try:
        shipping_address_email = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.email")
    except:
        print("in status exception")
    return shipping_address_email


def get_shipping_address_company(i):
    try:
        shipping_address_company = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.shipping_address.company")
    except:
        print("in status exception")
    return shipping_address_company


def get_shipping_address_phone(i):
    try:
        shipping_address_phone = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.phone")
    except:
        print("in status exception")
    return shipping_address_phone


def get_shipping_address_line1(i):
    try:
        shipping_address_line1 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.line1")
    except:
        print("in status exception")
    return shipping_address_line1


def get_shipping_address_line2(i):
    try:
        shipping_address_line2 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.line2")
    except:
        print("in status exception")
    return shipping_address_line2


def get_shipping_address_line3(i):
    try:
        shipping_address_line3 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.line3")
    except:
        print("in status exception")
    return shipping_address_line3


def get_shipping_address_city(i):
    try:
        shipping_address_city = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.city")
    except:
        print("in status exception")
    return shipping_address_city


def get_shipping_address_state_code(i):
    try:
        shipping_address_state_code = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].invoice.shipping_address.state_code")
    except:
        print("in status exception")
    return shipping_address_state_code


def get_shipping_address_state(i):
    try:
        shipping_address_state = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.state")
    except:
        print("in status exception")
    return shipping_address_state


def get_shipping_address_zip(i):
    try:
        shipping_address_zip = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.shipping_address.zip")
    except:
        print("in status exception")
    return shipping_address_zip


def get_shipping_address_country(i):
    try:
        shipping_address_country = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.shipping_address.country")
    except:
        print("in status exception")
    return shipping_address_country


def get_shipping_address_validation_status(i):
    try:
        shipping_address_validation_status = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].invoice.shipping_address.validation_status")
    except:
        print("in status exception")
    return shipping_address_validation_status


# line_items_1
def get_line_items_1_id(i):
    try:
        line_items_1_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[1].id")
    except:
        print("in status exception")
    return line_items_1_id


def get_line_items_1_description(i):
    try:
        line_items_1_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[1].description")
    except:
        print("in status exception")
    return line_items_1_description


def get_line_items_1_date_from(i):
    try:
        line_items_1_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[1].date_from")
    except:
        print("in status exception")
    return line_items_1_date_from


def get_line_items_1_date_to(i):
    try:
        line_items_1_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[1].date_to")
    except:
        print("in status exception")
    return line_items_1_date_to


def get_line_items_1_unit_amount(i):
    try:
        line_items_1_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[1].unit_amount")
    except:
        print("in status exception")
    return line_items_1_unit_amount


def get_line_items_1_quantity(i):
    try:
        line_items_1_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[1].quantity")
    except:
        print("in status exception")
    return line_items_1_quantity


def get_line_items_1_amount(i):
    try:
        line_items_1_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[1].amount")
    except:
        print("in status exception")
    return line_items_1_amount


def get_line_items_1_entity_type(i):
    try:
        line_items_1_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[1].entity_type")
    except:
        print("in status exception")
    return line_items_1_entity_type


def get_line_items_1_entity_id(i):
    try:
        line_items_1_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[1].entity_id")
    except:
        print("in status exception")
    return line_items_1_entity_id


def get_line_items_1_tax1_name(i):
    try:
        line_items_1_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[1].tax_name")
    except:
        print("in status exception")
    return line_items_1_tax1_name


def get_line_items_1_tax1_amount(i):
    try:
        line_items_1_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[1].tax_amount")
    except:
        print("in status exception")
    return line_items_1_tax1_amount


# line_items_2
def get_line_items_2_id(i):
    try:
        line_items_2_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[2].id")
    except:
        print("in status exception")
    return line_items_2_id


def get_line_items_2_description(i):
    try:
        line_items_2_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[2].description")
    except:
        print("in status exception")
    return line_items_2_description


def get_line_items_2_date_from(i):
    try:
        line_items_2_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[2].date_from")
    except:
        print("in status exception")
    return line_items_2_date_from


def get_line_items_2_date_to(i):
    try:
        line_items_2_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[2].date_to")
    except:
        print("in status exception")
    return line_items_2_date_to


def get_line_items_2_unit_amount(i):
    try:
        line_items_2_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[2].unit_amount")
    except:
        print("in status exception")
    return line_items_2_unit_amount


def get_line_items_2_quantity(i):
    try:
        line_items_2_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[2].quantity")
    except:
        print("in status exception")
    return line_items_2_quantity


def get_line_items_2_amount(i):
    try:
        line_items_2_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[2].amount")
    except:
        print("in status exception")
    return line_items_2_amount


def get_line_items_2_entity_type(i):
    try:
        line_items_2_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[2].entity_type")
    except:
        print("in status exception")
    return line_items_2_entity_type


def get_line_items_2_entity_id(i):
    try:
        line_items_2_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[2].entity_id")
    except:
        print("in status exception")
    return line_items_2_entity_id


def get_line_items_2_tax1_name(i):
    try:
        line_items_2_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[2].tax_name")
    except:
        print("in status exception")
    return line_items_2_tax1_name


def get_line_items_2_tax1_amount(i):
    try:
        line_items_2_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[2].tax_amount")
    except:
        print("in status exception")
    return line_items_2_tax1_amount


# line_items_3
def get_line_items_3_id(i):
    try:
        line_items_3_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[3].id")
    except:
        print("in status exception")
    return line_items_3_id


def get_line_items_3_description(i):
    try:
        line_items_3_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[3].description")
    except:
        print("in status exception")
    return line_items_3_description


def get_line_items_3_date_from(i):
    try:
        line_items_3_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[3].date_from")
    except:
        print("in status exception")
    return line_items_3_date_from


def get_line_items_3_date_to(i):
    try:
        line_items_3_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[3].date_to")
    except:
        print("in status exception")
    return line_items_3_date_to


def get_line_items_3_unit_amount(i):
    try:
        line_items_3_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[3].unit_amount")
    except:
        print("in status exception")
    return line_items_3_unit_amount


def get_line_items_3_quantity(i):
    try:
        line_items_3_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[3].quantity")
    except:
        print("in status exception")
    return line_items_3_quantity


def get_line_items_3_amount(i):
    try:
        line_items_3_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[3].amount")
    except:
        print("in status exception")
    return line_items_3_amount


def get_line_items_3_entity_type(i):
    try:
        line_items_3_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[3].entity_type")
    except:
        print("in status exception")
    return line_items_3_entity_type


def get_line_items_3_entity_id(i):
    try:
        line_items_3_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[3].entity_id")
    except:
        print("in status exception")
    return line_items_3_entity_id


def get_line_items_3_tax1_name(i):
    try:
        line_items_3_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[3].tax_name")
    except:
        print("in status exception")
    return line_items_3_tax1_name


def get_line_items_3_tax1_amount(i):
    try:
        line_items_3_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[3].tax_amount")
    except:
        print("in status exception")
    return line_items_3_tax1_amount


# line_items_4
def get_line_items_4_id(i):
    try:
        line_items_4_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[4].id")
    except:
        print("in status exception")
    return line_items_4_id


def get_line_items_4_description(i):
    try:
        line_items_4_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[4].description")
    except:
        print("in status exception")
    return line_items_4_description


def get_line_items_4_date_from(i):
    try:
        line_items_4_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[4].date_from")
    except:
        print("in status exception")
    return line_items_4_date_from


def get_line_items_4_date_to(i):
    try:
        line_items_4_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[4].date_to")
    except:
        print("in status exception")
    return line_items_4_date_to


def get_line_items_4_unit_amount(i):
    try:
        line_items_4_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[4].unit_amount")
    except:
        print("in status exception")
    return line_items_4_unit_amount


def get_line_items_4_quantity(i):
    try:
        line_items_4_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[4].quantity")
    except:
        print("in status exception")
    return line_items_4_quantity


def get_line_items_4_amount(i):
    try:
        line_items_4_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[4].amount")
    except:
        print("in status exception")
    return line_items_4_amount


def get_line_items_4_entity_type(i):
    try:
        line_items_4_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[4].entity_type")
    except:
        print("in status exception")
    return line_items_4_entity_type


def get_line_items_4_entity_id(i):
    try:
        line_items_4_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[4].entity_id")
    except:
        print("in status exception")
    return line_items_4_entity_id


def get_line_items_4_tax1_name(i):
    try:
        line_items_4_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[4].tax_name")
    except:
        print("in status exception")
    return line_items_4_tax1_name


def get_line_items_4_tax1_amount(i):
    try:
        line_items_4_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[4].tax_amount")
    except:
        print("in status exception")
    return line_items_4_tax1_amount


# line_items_5
def get_line_items_5_id(i):
    try:
        line_items_5_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[5].id")
    except:
        print("in status exception")
    return line_items_5_id


def get_line_items_5_description(i):
    try:
        line_items_5_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[5].description")
    except:
        print("in status exception")
    return line_items_5_description


def get_line_items_5_date_from(i):
    try:
        line_items_5_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[5].date_from")
    except:
        print("in status exception")
    return line_items_5_date_from


def get_line_items_5_date_to(i):
    try:
        line_items_5_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[5].date_to")
    except:
        print("in status exception")
    return line_items_5_date_to


def get_line_items_5_unit_amount(i):
    try:
        line_items_5_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[5].unit_amount")
    except:
        print("in status exception")
    return line_items_5_unit_amount


def get_line_items_5_quantity(i):
    try:
        line_items_5_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[5].quantity")
    except:
        print("in status exception")
    return line_items_5_quantity


def get_line_items_5_amount(i):
    try:
        line_items_5_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[5].amount")
    except:
        print("in status exception")
    return line_items_5_amount


def get_line_items_5_entity_type(i):
    try:
        line_items_5_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[5].entity_type")
    except:
        print("in status exception")
    return line_items_5_entity_type


def get_line_items_5_entity_id(i):
    try:
        line_items_5_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[5].entity_id")
    except:
        print("in status exception")
    return line_items_5_entity_id


def get_line_items_5_tax1_name(i):
    try:
        line_items_5_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[5].tax_name")
    except:
        print("in status exception")
    return line_items_5_tax1_name


def get_line_items_5_tax1_amount(i):
    try:
        line_items_5_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[5].tax_amount")
    except:
        print("in status exception")
    return line_items_5_tax1_amount


# line_items_6
def get_line_items_6_id(i):
    try:
        line_items_6_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[6].id")
    except:
        print("in status exception")
    return line_items_6_id


def get_line_items_6_description(i):
    try:
        line_items_6_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[6].description")
    except:
        print("in status exception")
    return line_items_6_description


def get_line_items_6_date_from(i):
    try:
        line_items_6_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[6].date_from")
    except:
        print("in status exception")
    return line_items_6_date_from


def get_line_items_6_date_to(i):
    try:
        line_items_6_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[6].date_to")
    except:
        print("in status exception")
    return line_items_6_date_to


def get_line_items_6_unit_amount(i):
    try:
        line_items_6_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[6].unit_amount")
    except:
        print("in status exception")
    return line_items_6_unit_amount


def get_line_items_6_quantity(i):
    try:
        line_items_6_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[6].quantity")
    except:
        print("in status exception")
    return line_items_6_quantity


def get_line_items_6_amount(i):
    try:
        line_items_6_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[6].amount")
    except:
        print("in status exception")
    return line_items_6_amount


def get_line_items_6_entity_type(i):
    try:
        line_items_6_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[6].entity_type")
    except:
        print("in status exception")
    return line_items_6_entity_type


def get_line_items_6_entity_id(i):
    try:
        line_items_6_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[6].entity_id")
    except:
        print("in status exception")
    return line_items_6_entity_id


def get_line_items_6_tax1_name(i):
    try:
        line_items_6_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[6].tax_name")
    except:
        print("in status exception")
    return line_items_6_tax1_name


def get_line_items_6_tax1_amount(i):
    try:
        line_items_6_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[6].tax_amount")
    except:
        print("in status exception")
    return line_items_6_tax1_amount


# line_items_7
def get_line_items_7_id(i):
    try:
        line_items_7_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[7].id")
    except:
        print("in status exception")
    return line_items_7_id


def get_line_items_7_description(i):
    try:
        line_items_7_description = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[7].description")
    except:
        print("in status exception")
    return line_items_7_description


def get_line_items_7_date_from(i):
    try:
        line_items_7_date_from = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[7].date_from")
    except:
        print("in status exception")
    return line_items_7_date_from


def get_line_items_7_date_to(i):
    try:
        line_items_7_date_to = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[7].date_to")
    except:
        print("in status exception")
    return line_items_7_date_to


def get_line_items_7_unit_amount(i):
    try:
        line_items_7_unit_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[7].unit_amount")
    except:
        print("in status exception")
    return line_items_7_unit_amount


def get_line_items_7_quantity(i):
    try:
        line_items_7_quantity = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[7].quantity")
    except:
        print("in status exception")
    return line_items_7_quantity


def get_line_items_7_amount(i):
    try:
        line_items_7_amount = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[7].amount")
    except:
        print("in status exception")
    return line_items_7_amount


def get_line_items_7_entity_type(i):
    try:
        line_items_7_entity_type = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_items[7].entity_type")
    except:
        print("in status exception")
    return line_items_7_entity_type


def get_line_items_7_entity_id(i):
    try:
        line_items_7_entity_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].invoice.line_items[7].entity_id")
    except:
        print("in status exception")
    return line_items_7_entity_id


def get_line_items_7_tax1_name(i):
    try:
        line_items_7_tax1_name = jsonpath.jsonpath(jsonpathres,
                                                   "list[" + str(i) + "].invoice.line_item_taxes[7].tax_name")
    except:
        print("in status exception")
    return line_items_7_tax1_name


def get_line_items_7_tax1_amount(i):
    try:
        line_items_7_tax1_amount = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].invoice.line_item_taxes[7].tax_amount")
    except:
        print("in status exception")
    return line_items_7_tax1_amount



# ---------------------------loop through all records  and write to excel---------------------------#
for i in range(0, totalRecordCountInResp):
    try:

        if Company_name_CNo != "NA":
            Company = get_billing_address_company(i)
            Company_cell = my_sheet.cell(row=i + 2, column=Company_name_CNo)
            if Company == False:
                Company_cell.value = Company
            else:
                Company_cell.value = str(Company[0])

        if id_CNo != "NA":
            id = get_id(i)
            id_cell = my_sheet.cell(row=i + 2, column=id_CNo)
            if id == False:
                id_cell.value = id
            else:
                id_cell.value = str(id[0])

        if customer_id_CNo != "NA":
            customer_id = get_customer_id(i)
            customer_id_cell = my_sheet.cell(row=i + 2, column=customer_id_CNo)
            if customer_id == False:
                customer_id_cell.value = customer_id
            else:
                customer_id_cell.value = str(customer_id[0])

        if subscription_id_CNo != "NA":
            subscription_id = get_subscription_id(i)
            subscription_id_cell = my_sheet.cell(row=i + 2, column=subscription_id_CNo)
            if subscription_id == False:
                subscription_id_cell.value = subscription_id
            else:
                subscription_id_cell.value = str(subscription_id[0])

        if invoice_po_number_CNo != "NA":
            invoice_po_number = get_invoice_po_number(i)
            invoice_po_number_cell = my_sheet.cell(row=i + 2, column=invoice_po_number_CNo)
            if invoice_po_number == False:
                invoice_po_number_cell.value = invoice_po_number
            else:
                invoice_po_number_cell.value = str(invoice_po_number[0])

        if tax_override_reason_CNo != "NA":
            tax_override_reason = get_tax_override_reason(i)
            tax_override_reason_cell = my_sheet.cell(row=i + 2, column=tax_override_reason_CNo)
            if tax_override_reason == False:
                tax_override_reason_cell.value = tax_override_reason
            else:
                tax_override_reason_cell.value = str(tax_override_reason[0])

        if recurring_CNo != "NA":
            recurring = get_recurring(i)
            recurring_cell = my_sheet.cell(row=i + 2, column=recurring_CNo)
            if recurring == False:
                recurring_cell.value = recurring
            else:
                recurring_cell.value = str(recurring[0])

        if status_CNo != "NA":
            status = get_status(i)
            status_cell = my_sheet.cell(row=i + 2, column=status_CNo)
            if status == False:
                status_cell.value = status
            else:
                status_cell.value = str(status[0])

        if vat_number_CNo != "NA":
            vat_number = get_vat_number(i)
            vat_number_cell = my_sheet.cell(row=i + 2, column=vat_number_CNo)
            if vat_number == False:
                vat_number_cell.value = vat_number
            else:
                vat_number_cell.value = str(vat_number[0])

        if price_type_CNo != "NA":
            price_type = get_price_type(i)
            price_type_cell = my_sheet.cell(row=i + 2, column=price_type_CNo)
            if price_type == False:
                price_type_cell.value = price_type
            else:
                price_type_cell.value = str(price_type[0])

        if date_CNo != "NA":
            date = get_date(i)
            date_cell = my_sheet.cell(row=i + 2, column=date_CNo)
            if date == False:
                date_cell.value = date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(date[0], clienttimezone)
                date_cell.value = str(modifiedtimestamp)

        if due_date_CNo != "NA":
            due_date = get_due_date(i)
            due_date_cell = my_sheet.cell(row=i + 2, column=due_date_CNo)
            if due_date == False:
                due_date_cell.value = due_date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(due_date[0], clienttimezone)
                due_date_cell.value = str(modifiedtimestamp)

        if net_term_days_CNo != "NA":
            net_term_days = get_net_term_days(i)
            net_term_days_cell = my_sheet.cell(row=i + 2, column=net_term_days_CNo)
            if net_term_days == False:
                net_term_days_cell.value = net_term_days
            else:
                net_term_days_cell.value = str(net_term_days[0])

        if use_for_proration_CNo != "NA":
            use_for_proration = get_use_for_proration(i)
            use_for_proration_cell = my_sheet.cell(row=i + 2, column=use_for_proration_CNo)
            if use_for_proration == False:
                use_for_proration_cell.value = use_for_proration
            else:
                use_for_proration_cell.value = str(use_for_proration[0])

        if exchange_rate_CNo != "NA":
            exchange_rate = get_exchange_rate(i)
            exchange_rate_cell = my_sheet.cell(row=i + 2, column=exchange_rate_CNo)
            if exchange_rate == False:
                exchange_rate_cell.value = exchange_rate
            else:
                exchange_rate_cell.value = str(exchange_rate[0])

        if total_CNo != "NA":
            total = get_total(i)
            total_cell = my_sheet.cell(row=i + 2, column=total_CNo)
            if total == False:
                total_cell.value = total
            else:
                dollarconverted = tzconverter.centToDollar(total[0])
                total_cell.value = str(dollarconverted)
                # total_cell.value = str(total[0])

        if amount_paid_CNo != "NA":
            amount_paid = get_amount_paid(i)
            amount_paid_cell = my_sheet.cell(row=i + 2, column=amount_paid_CNo)
            if amount_paid == False:
                amount_paid_cell.value = amount_paid
            else:
                dollarconverted = tzconverter.centToDollar(total[0])
                amount_paid_cell.value = str(dollarconverted)

        if amount_adjusted_CNo != "NA":
            amount_adjusted = get_amount_adjusted(i)
            amount_adjusted_cell = my_sheet.cell(row=i + 2, column=amount_adjusted_CNo)
            if amount_adjusted == False:
                amount_adjusted_cell.value = amount_adjusted
            else:
                amount_adjusted_cell.value = str(amount_adjusted[0])

        if write_off_amount_CNo != "NA":
            write_off_amount = get_write_off_amount(i)
            write_off_amount_cell = my_sheet.cell(row=i + 2, column=write_off_amount_CNo)
            if write_off_amount == False:
                write_off_amount_cell.value = write_off_amount
            else:
                write_off_amount_cell.value = str(write_off_amount[0])

        if credits_applied_CNo != "NA":
            credits_applied = get_credits_applied(i)
            credits_applied_cell = my_sheet.cell(row=i + 2, column=credits_applied_CNo)
            if credits_applied == False:
                credits_applied_cell.value = credits_applied
            else:
                credits_applied_cell.value = str(credits_applied[0])

        if amount_due_CNo != "NA":
            amount_due = get_amount_due(i)
            amount_due_cell = my_sheet.cell(row=i + 2, column=amount_due_CNo)
            if amount_due == False:
                amount_due_cell.value = amount_due
            else:
                amount_due_cell.value = str(amount_due[0])

        if dunning_status_CNo != "NA":
            dunning_status = get_dunning_status(i)
            dunning_status_cell = my_sheet.cell(row=i + 2, column=dunning_status_CNo)
            if dunning_status == False:
                dunning_status_cell.value = dunning_status
            else:
                dunning_status_cell.value = str(dunning_status[0])

        if next_retry_at_CNo != "NA":
            next_retry_at = get_next_retry_at(i)
            next_retry_at_cell = my_sheet.cell(row=i + 2, column=next_retry_at_CNo)
            if next_retry_at == False:
                next_retry_at_cell.value = next_retry_at
            else:
                next_retry_at_cell.value = str(next_retry_at[0])

        if updated_at_CNo != "NA":
            updated_at = get_updated_at(i)
            updated_at_cell = my_sheet.cell(row=i + 2, column=updated_at_CNo)
            if updated_at == False:
                updated_at_cell.value = updated_at
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(updated_at[0], clienttimezone)
                updated_at_cell.value = str(modifiedtimestamp)

        if resource_version_CNo != "NA":
            resource_version = get_resource_version(i)
            resource_version_cell = my_sheet.cell(row=i + 2, column=resource_version_CNo)
            if resource_version == False:
                resource_version_cell.value = resource_version
            else:
                resource_version_cell.value = str(resource_version[0])

        if deleted_CNo != "NA":
            deleted = get_deleted(i)
            deleted_cell = my_sheet.cell(row=i + 2, column=deleted_CNo)
            if deleted == False:
                deleted_cell.value = deleted
            else:
                deleted_cell.value = str(deleted[0])

        if object_CNo != "NA":
            object = get_object(i)
            object_cell = my_sheet.cell(row=i + 2, column=object_CNo)
            if object == False:
                object_cell.value = object
            else:
                object_cell.value = str(object[0])

        if first_invoice_CNo != "NA":
            first_invoice = get_first_invoice(i)
            first_invoice_cell = my_sheet.cell(row=i + 2, column=first_invoice_CNo)
            if first_invoice == False:
                first_invoice_cell.value = first_invoice
            else:
                first_invoice_cell.value = str(first_invoice[0])

        if amount_to_collect_CNo != "NA":
            amount_to_collect = get_amount_to_collect(i)
            amount_to_collect_cell = my_sheet.cell(row=i + 2, column=amount_to_collect_CNo)
            if amount_to_collect == False:
                amount_to_collect_cell.value = amount_to_collect
            else:
                amount_to_collect_cell.value = str(amount_to_collect[0])

        if round_off_amount_CNo != "NA":
            round_off_amount = get_round_off_amount(i)
            round_off_amount_cell = my_sheet.cell(row=i + 2, column=round_off_amount_CNo)
            if round_off_amount == False:
                round_off_amount_cell.value = round_off_amount
            else:
                round_off_amount_cell.value = str(round_off_amount[0])

        if has_advance_charges_CNo != "NA":
            has_advance_charges = get_has_advance_charges(i)
            has_advance_charges_cell = my_sheet.cell(row=i + 2, column=has_advance_charges_CNo)
            if has_advance_charges == False:
                has_advance_charges_cell.value = has_advance_charges
            else:
                has_advance_charges_cell.value = str(has_advance_charges[0])

        if currency_code_CNo != "NA":
            currency_code = get_currency_code(i)
            currency_code_cell = my_sheet.cell(row=i + 2, column=currency_code_CNo)
            if currency_code == False:
                currency_code_cell.value = currency_code
            else:
                currency_code_cell.value = str(currency_code[0])

        if base_currency_code_CNo != "NA":
            base_currency_code = get_base_currency_code(i)
            base_currency_code_cell = my_sheet.cell(row=i + 2, column=base_currency_code_CNo)
            if base_currency_code == False:
                base_currency_code_cell.value = base_currency_code
            else:
                base_currency_code_cell.value = str(base_currency_code[0])

        if is_gifted_CNo != "NA":
            is_gifted = get_is_gifted(i)
            is_gifted_cell = my_sheet.cell(row=i + 2, column=is_gifted_CNo)
            if is_gifted == False:
                is_gifted_cell.value = is_gifted
            else:
                is_gifted_cell.value = str(is_gifted[0])

        if term_finalized_CNo != "NA":
            term_finalized = get_term_finalized(i)
            term_finalized_cell = my_sheet.cell(row=i + 2, column=term_finalized_CNo)
            if term_finalized == False:
                term_finalized_cell.value = term_finalized
            else:
                term_finalized_cell.value = str(term_finalized[0])

        if is_digital_CNo != "NA":
            is_digital = get_is_digital(i)
            is_digital_cell = my_sheet.cell(row=i + 2, column=is_digital_CNo)
            if is_digital == False:
                is_digital_cell.value = is_digital
            else:
                is_digital_cell.value = str(is_digital[0])

        if tax_CNo != "NA":
            tax = get_tax(i)
            tax_cell = my_sheet.cell(row=i + 2, column=tax_CNo)
            if tax == False:
                tax_cell.value = tax
            else:
                tax_cell.value = str(tax[0])

        if line_items_CNo != "NA":
            line_items = get_line_items(i)
            line_items_cell = my_sheet.cell(row=i + 2, column=line_items_CNo)
            if line_items == False:
                line_items_cell.value = line_items
            else:
                line_items_cell.value = str(line_items[0])

        if line_items_0__CNo != "NA":
            line_items_0_ = get_line_items_0_(i)
            line_items_0__cell = my_sheet.cell(row=i + 2, column=line_items_0__CNo)
            if line_items_0_ == False:
                line_items_0__cell.value = line_items_0_
            else:
                line_items_0__cell.value = str(line_items_0_[0])

        if line_items_0_id_CNo != "NA":
            line_items_0_id = get_line_items_0_id(i)
            line_items_0_id_cell = my_sheet.cell(row=i + 2, column=line_items_0_id_CNo)
            if line_items_0_id == False:
                line_items_0_id_cell.value = line_items_0_id
            else:
                line_items_0_id_cell.value = str(line_items_0_id[0])

        if line_items_0_date_from_CNo != "NA":
            line_items_0_date_from = get_line_items_0_date_from(i)
            line_items_0_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_0_date_from_CNo)
            if line_items_0_date_from == False:
                line_items_0_date_from_cell.value = line_items_0_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_0_date_from[0], clienttimezone)
                line_items_0_date_from_cell.value = str(modifiedtimestamp)

        if line_items_0_date_to_CNo != "NA":
            line_items_0_date_to = get_line_items_0_date_to(i)
            line_items_0_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_0_date_to_CNo)
            if line_items_0_date_to == False:
                line_items_0_date_to_cell.value = line_items_0_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_0_date_to[0], clienttimezone)
                line_items_0_date_to_cell.value = str(modifiedtimestamp)

        if line_items_0_unit_amount_CNo != "NA":
            line_items_0_unit_amount = get_line_items_0_unit_amount(i)
            line_items_0_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_0_unit_amount_CNo)
            if line_items_0_unit_amount == False:
                line_items_0_unit_amount_cell.value = line_items_0_unit_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_0_unit_amount[0])
                # line_items_0_unit_amount_cell.value = str(dollarconverted)
                line_items_0_unit_amount_cell.value = str(line_items_0_unit_amount[0])

        if line_items_0_quantity_CNo != "NA":
            line_items_0_quantity = get_line_items_0_quantity(i)
            line_items_0_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_0_quantity_CNo)
            if line_items_0_quantity == False:
                line_items_0_quantity_cell.value = line_items_0_quantity
            else:
                line_items_0_quantity_cell.value = str(line_items_0_quantity[0])

        if line_items_0_amount_CNo != "NA":
            line_items_0_amount = get_line_items_0_amount(i)
            line_items_0_amount_cell = my_sheet.cell(row=i + 2, column=line_items_0_amount_CNo)
            if line_items_0_amount == False:
                line_items_0_amount_cell.value = line_items_0_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_0_amount[0])
                line_items_0_amount_cell.value = str(dollarconverted)
                # line_items_0_amount_cell.value = str(line_items_0_amount[0])

        if line_items_0_pricing_model_CNo != "NA":
            line_items_0_pricing_model = get_line_items_0_pricing_model(i)
            line_items_0_pricing_model_cell = my_sheet.cell(row=i + 2, column=line_items_0_pricing_model_CNo)
            if line_items_0_pricing_model == False:
                line_items_0_pricing_model_cell.value = line_items_0_pricing_model
            else:
                line_items_0_pricing_model_cell.value = str(line_items_0_pricing_model[0])

        if line_items_0_is_taxed_CNo != "NA":
            line_items_0_is_taxed = get_line_items_0_is_taxed(i)
            line_items_0_is_taxed_cell = my_sheet.cell(row=i + 2, column=line_items_0_is_taxed_CNo)
            if line_items_0_is_taxed == False:
                line_items_0_is_taxed_cell.value = line_items_0_is_taxed
            else:
                line_items_0_is_taxed_cell.value = str(line_items_0_is_taxed[0])

        if line_items_0_tax_amount_CNo != "NA":
            line_items_0_tax_amount = get_line_items_0_tax_amount(i)
            line_items_0_tax_amount_cell = my_sheet.cell(row=i + 2, column=line_items_0_tax_amount_CNo)
            if line_items_0_tax_amount == False:
                line_items_0_tax_amount_cell.value = line_items_0_tax_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_0_tax_amount[0])
                line_items_0_tax_amount_cell.value = str(dollarconverted)
                # line_items_0_tax_amount_cell.value = str(line_items_0_tax_amount[0])

        if line_items_0_tax_rate_CNo != "NA":
            line_items_0_tax_rate = get_line_items_0_tax_rate(i)
            line_items_0_tax_rate_cell = my_sheet.cell(row=i + 2, column=line_items_0_tax_rate_CNo)
            if line_items_0_tax_rate == False:
                line_items_0_tax_rate_cell.value = line_items_0_tax_rate
            else:
                line_items_0_tax_rate_cell.value = str(line_items_0_tax_rate[0])

        if line_items_0_object_CNo != "NA":
            line_items_0_object = get_line_items_0_object(i)
            line_items_0_object_cell = my_sheet.cell(row=i + 2, column=line_items_0_object_CNo)
            if line_items_0_object == False:
                line_items_0_object_cell.value = line_items_0_object
            else:
                line_items_0_object_cell.value = str(line_items_0_object[0])

        if line_items_0_subscription_id_CNo != "NA":
            line_items_0_subscription_id = get_line_items_0_subscription_id(i)
            line_items_0_subscription_id_cell = my_sheet.cell(row=i + 2, column=line_items_0_subscription_id_CNo)
            if line_items_0_subscription_id == False:
                line_items_0_subscription_id_cell.value = line_items_0_subscription_id
            else:
                line_items_0_subscription_id_cell.value = str(line_items_0_subscription_id[0])

        if line_items_0_customer_id_CNo != "NA":
            line_items_0_customer_id = get_line_items_0_customer_id(i)
            line_items_0_customer_id_cell = my_sheet.cell(row=i + 2, column=line_items_0_customer_id_CNo)
            if line_items_0_customer_id == False:
                line_items_0_customer_id_cell.value = line_items_0_customer_id
            else:
                line_items_0_customer_id_cell.value = str(line_items_0_customer_id[0])

        if line_items_0_description_CNo != "NA":
            line_items_0_description = get_line_items_0_description(i)
            line_items_0_description_cell = my_sheet.cell(row=i + 2, column=line_items_0_description_CNo)
            if line_items_0_description == False:
                line_items_0_description_cell.value = line_items_0_description
            else:
                line_items_0_description_cell.value = str(line_items_0_description[0])

        if line_items_0_entity_type_CNo != "NA":
            line_items_0_entity_type = get_line_items_0_entity_type(i)
            line_items_0_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_0_entity_type_CNo)
            if line_items_0_entity_type == False:
                line_items_0_entity_type_cell.value = line_items_0_entity_type
            else:
                line_items_0_entity_type_cell.value = str(line_items_0_entity_type[0])

        if line_items_0_entity_id_CNo != "NA":
            line_items_0_entity_id = get_line_items_0_entity_id(i)
            line_items_0_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_0_entity_id_CNo)
            if line_items_0_entity_id == False:
                line_items_0_entity_id_cell.value = line_items_0_entity_id
            else:
                line_items_0_entity_id_cell.value = str(line_items_0_entity_id[0])

        if line_items_0_discount_amount_CNo != "NA":
            line_items_0_discount_amount = get_line_items_0_discount_amount(i)
            line_items_0_discount_amount_cell = my_sheet.cell(row=i + 2, column=line_items_0_discount_amount_CNo)
            if line_items_0_discount_amount == False:
                line_items_0_discount_amount_cell.value = line_items_0_discount_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_0_discount_amount[0])
                # line_items_0_discount_amount_cell.value = str(dollarconverted)
                line_items_0_discount_amount_cell.value = str(line_items_0_discount_amount[0])

        if line_items_0_item_level_discount_amount_CNo != "NA":
            line_items_0_item_level_discount_amount = get_line_items_0_item_level_discount_amount(i)
            line_items_0_item_level_discount_amount_cell = my_sheet.cell(row=i + 2,
                                                                         column=line_items_0_item_level_discount_amount_CNo)
            if line_items_0_item_level_discount_amount == False:
                line_items_0_item_level_discount_amount_cell.value = line_items_0_item_level_discount_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_0_item_level_discount_amount[0])
                # line_items_0_item_level_discount_amount_cell.value = str(dollarconverted)
                line_items_0_item_level_discount_amount_cell.value = str(line_items_0_item_level_discount_amount[0])

        if line_items_0_tax1_name_CNo != "NA":
            line_items_0_tax1_name = get_line_items_0_tax1_name(i)
            line_items_0_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_0_tax1_name_CNo)
            if line_items_0_tax1_name == False:
                line_items_0_tax1_name_cell.value = line_items_0_tax1_name
            else:
                line_items_0_tax1_name_cell.value = str(line_items_0_tax1_name[0])

        if line_items_0_tax1_amount_CNo != "NA":
            line_items_0_tax1_amount = get_line_items_0_tax1_amount(i)
            line_items_0_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_0_tax1_amount_CNo)
            if line_items_0_tax1_amount == False:
                line_items_0_tax1_amount_cell.value = line_items_0_tax1_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_0_tax1_amount[0])
                # line_items_0_tax1_amount_cell.value = str(dollarconverted)
                line_items_0_tax1_amount_cell.value = str(line_items_0_tax1_amount[0])

        if taxes_CNo != "NA":
            taxes = get_taxes(i)
            taxes_cell = my_sheet.cell(row=i + 2, column=taxes_CNo)
            if taxes == False:
                taxes_cell.value = taxes
            else:
                taxes_cell.value = str(taxes[0])

        if taxes_0__CNo != "NA":
            taxes_0_ = get_taxes_0_(i)
            taxes_0__cell = my_sheet.cell(row=i + 2, column=taxes_0__CNo)
            if taxes_0_ == False:
                taxes_0__cell.value = taxes_0_
            else:
                taxes_0__cell.value = str(taxes_0_[0])

        if taxes_0_object_CNo != "NA":
            taxes_0_object = get_taxes_0_object(i)
            taxes_0_object_cell = my_sheet.cell(row=i + 2, column=taxes_0_object_CNo)
            if taxes_0_object == False:
                taxes_0_object_cell.value = taxes_0_object
            else:
                taxes_0_object_cell.value = str(taxes_0_object[0])

        if taxes_0_name_CNo != "NA":
            taxes_0_name = get_taxes_0_name(i)
            taxes_0_name_cell = my_sheet.cell(row=i + 2, column=taxes_0_name_CNo)
            if taxes_0_name == False:
                taxes_0_name_cell.value = taxes_0_name
            else:
                taxes_0_name_cell.value = str(taxes_0_name[0])

        if taxes_0_rate_CNo != "NA":
            taxes_0_rate = get_taxes_0_rate(i)
            taxes_0_rate_cell = my_sheet.cell(row=i + 2, column=taxes_0_rate_CNo)
            if taxes_0_rate == False:
                taxes_0_rate_cell.value = taxes_0_rate
            else:
                taxes_0_rate_cell.value = str(taxes_0_rate[0])

        if taxes_0_description_CNo != "NA":
            taxes_0_description = get_taxes_0_description(i)
            taxes_0_description_cell = my_sheet.cell(row=i + 2, column=taxes_0_description_CNo)
            if taxes_0_description == False:
                taxes_0_description_cell.value = taxes_0_description
            else:
                taxes_0_description_cell.value = str(taxes_0_description[0])

        if taxes_0_amount_CNo != "NA":
            taxes_0_amount = get_taxes_0_amount(i)
            taxes_0_amount_cell = my_sheet.cell(row=i + 2, column=taxes_0_amount_CNo)
            if taxes_0_amount == False:
                taxes_0_amount_cell.value = taxes_0_amount
            else:
                # dollarconverted = tzconverter.centToDollar(taxes_0_amount[0])
                # taxes_0_amount_cell.value = str(dollarconverted)
                taxes_0_amount_cell.value = str(taxes_0_amount[0])

        if discounts_0_entity_type_CNo != "NA":
            discounts_0_entity_type = get_discounts_0_entity_type(i)
            discounts_0_entity_type_cell = my_sheet.cell(row=i + 2, column=discounts_0_entity_type_CNo)
            if discounts_0_entity_type == False:
                discounts_0_entity_type_cell.value = discounts_0_entity_type
            else:
                discounts_0_entity_type_cell.value = str(discounts_0_entity_type[0])

        if discounts_0_entity_id_CNo != "NA":
            discounts_0_entity_id = get_discounts_0_entity_id(i)
            discounts_0_entity_id_cell = my_sheet.cell(row=i + 2, column=discounts_0_entity_id_CNo)
            if discounts_0_entity_id == False:
                discounts_0_entity_id_cell.value = discounts_0_entity_id
            else:
                discounts_0_entity_id_cell.value = str(discounts_0_entity_id[0])

        if discounts_0_amount_CNo != "NA":
            discounts_0_amount = get_discounts_0_amount(i)
            discounts_0_amount_cell = my_sheet.cell(row=i + 2, column=discounts_0_amount_CNo)
            if discounts_0_amount == False:
                discounts_0_amount_cell.value = discounts_0_amount
            else:
                # dollarconverted = tzconverter.centToDollar(discounts_0_amount[0])
                # discounts_0_amount_cell.value = str(dollarconverted)
                discounts_0_amount_cell.value = str(discounts_0_amount[0])

        if payments_0_amount_CNo != "NA":
            payments_0_amount = get_payments_0_amount(i)
            payments_0_amount_cell = my_sheet.cell(row=i + 2, column=payments_0_amount_CNo)
            if payments_0_amount == False:
                payments_0_amount_cell.value = payments_0_amount
            else:
                # dollarconverted = tzconverter.centToDollar(payments_0_amount[0])
                # payments_0_amount_cell.value = str(dollarconverted)
                payments_0_amount_cell.value = str(payments_0_amount[0])

        if payments_0_payment_method_CNo != "NA":
            payments_0_payment_method = get_payments_0_payment_method(i)
            payments_0_payment_method_cell = my_sheet.cell(row=i + 2, column=payments_0_payment_method_CNo)
            if payments_0_payment_method == False:
                payments_0_payment_method_cell.value = payments_0_payment_method
            else:
                payments_0_payment_method_cell.value = str(payments_0_payment_method[0])

        if payments_0_date_CNo != "NA":
            payments_0_date = get_payments_0_date(i)
            payments_0_date_cell = my_sheet.cell(row=i + 2, column=payments_0_date_CNo)
            if payments_0_date == False:
                payments_0_date_cell.value = payments_0_date
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(payments_0_date[0], clienttimezone)
                payments_0_date_cell.value = str(modifiedtimestamp)

        if line_item_taxes_CNo != "NA":
            line_item_taxes = get_line_item_taxes(i)
            line_item_taxes_cell = my_sheet.cell(row=i + 2, column=line_item_taxes_CNo)
            if line_item_taxes == False:
                line_item_taxes_cell.value = line_item_taxes
            else:
                line_item_taxes_cell.value = str(line_item_taxes[0])

        if line_item_taxes_0__CNo != "NA":
            line_item_taxes_0_ = get_line_item_taxes_0_(i)
            line_item_taxes_0__cell = my_sheet.cell(row=i + 2, column=line_item_taxes_0__CNo)
            if line_item_taxes_0_ == False:
                line_item_taxes_0__cell.value = line_item_taxes_0_
            else:
                line_item_taxes_0__cell.value = str(line_item_taxes_0_[0])

        if line_item_taxes_0_tax_name_CNo != "NA":
            line_item_taxes_0_tax_name = get_line_item_taxes_0_tax_name(i)
            line_item_taxes_0_tax_name_cell = my_sheet.cell(row=i + 2, column=line_item_taxes_0_tax_name_CNo)
            if line_item_taxes_0_tax_name == False:
                line_item_taxes_0_tax_name_cell.value = line_item_taxes_0_tax_name
            else:
                line_item_taxes_0_tax_name_cell.value = str(line_item_taxes_0_tax_name[0])

        if line_item_taxes_0_tax_rate_CNo != "NA":
            line_item_taxes_0_tax_rate = get_line_item_taxes_0_tax_rate(i)
            line_item_taxes_0_tax_rate_cell = my_sheet.cell(row=i + 2, column=line_item_taxes_0_tax_rate_CNo)
            if line_item_taxes_0_tax_rate == False:
                line_item_taxes_0_tax_rate_cell.value = line_item_taxes_0_tax_rate
            else:
                line_item_taxes_0_tax_rate_cell.value = str(line_item_taxes_0_tax_rate[0])

        if line_item_taxes_0_tax_juris_type_CNo != "NA":
            line_item_taxes_0_tax_juris_type = get_line_item_taxes_0_tax_juris_type(i)
            line_item_taxes_0_tax_juris_type_cell = my_sheet.cell(row=i + 2,
                                                                  column=line_item_taxes_0_tax_juris_type_CNo)
            if line_item_taxes_0_tax_juris_type == False:
                line_item_taxes_0_tax_juris_type_cell.value = line_item_taxes_0_tax_juris_type
            else:
                line_item_taxes_0_tax_juris_type_cell.value = str(line_item_taxes_0_tax_juris_type[0])

        if line_item_taxes_0_tax_juris_name_CNo != "NA":
            line_item_taxes_0_tax_juris_name = get_line_item_taxes_0_tax_juris_name(i)
            line_item_taxes_0_tax_juris_name_cell = my_sheet.cell(row=i + 2,
                                                                  column=line_item_taxes_0_tax_juris_name_CNo)
            if line_item_taxes_0_tax_juris_name == False:
                line_item_taxes_0_tax_juris_name_cell.value = line_item_taxes_0_tax_juris_name
            else:
                line_item_taxes_0_tax_juris_name_cell.value = str(line_item_taxes_0_tax_juris_name[0])

        if line_item_taxes_0_tax_juris_code_CNo != "NA":
            line_item_taxes_0_tax_juris_code = get_line_item_taxes_0_tax_juris_code(i)
            line_item_taxes_0_tax_juris_code_cell = my_sheet.cell(row=i + 2,
                                                                  column=line_item_taxes_0_tax_juris_code_CNo)
            if line_item_taxes_0_tax_juris_code == False:
                line_item_taxes_0_tax_juris_code_cell.value = line_item_taxes_0_tax_juris_code
            else:
                line_item_taxes_0_tax_juris_code_cell.value = str(line_item_taxes_0_tax_juris_code[0])

        if line_item_taxes_0_object_CNo != "NA":
            line_item_taxes_0_object = get_line_item_taxes_0_object(i)
            line_item_taxes_0_object_cell = my_sheet.cell(row=i + 2, column=line_item_taxes_0_object_CNo)
            if line_item_taxes_0_object == False:
                line_item_taxes_0_object_cell.value = line_item_taxes_0_object
            else:
                line_item_taxes_0_object_cell.value = str(line_item_taxes_0_object[0])

        if line_item_taxes_0_line_item_id_CNo != "NA":
            line_item_taxes_0_line_item_id = get_line_item_taxes_0_line_item_id(i)
            line_item_taxes_0_line_item_id_cell = my_sheet.cell(row=i + 2, column=line_item_taxes_0_line_item_id_CNo)
            if line_item_taxes_0_line_item_id == False:
                line_item_taxes_0_line_item_id_cell.value = line_item_taxes_0_line_item_id
            else:
                line_item_taxes_0_line_item_id_cell.value = str(line_item_taxes_0_line_item_id[0])

        if line_item_taxes_0_tax_amount_CNo != "NA":
            line_item_taxes_0_tax_amount = get_line_item_taxes_0_tax_amount(i)
            line_item_taxes_0_tax_amount_cell = my_sheet.cell(row=i + 2, column=line_item_taxes_0_tax_amount_CNo)
            if line_item_taxes_0_tax_amount == False:
                line_item_taxes_0_tax_amount_cell.value = line_item_taxes_0_tax_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_item_taxes_0_tax_amount[0])
                # line_item_taxes_0_tax_amount_cell.value = str(dollarconverted)
                line_item_taxes_0_tax_amount_cell.value = str(line_item_taxes_0_tax_amount[0])

        if line_item_taxes_0_is_partial_tax_applied_CNo != "NA":
            line_item_taxes_0_is_partial_tax_applied = get_line_item_taxes_0_is_partial_tax_applied(i)
            line_item_taxes_0_is_partial_tax_applied_cell = my_sheet.cell(row=i + 2,
                                                                          column=line_item_taxes_0_is_partial_tax_applied_CNo)
            if line_item_taxes_0_is_partial_tax_applied == False:
                line_item_taxes_0_is_partial_tax_applied_cell.value = line_item_taxes_0_is_partial_tax_applied
            else:
                line_item_taxes_0_is_partial_tax_applied_cell.value = str(line_item_taxes_0_is_partial_tax_applied[0])

        if line_item_taxes_0_taxable_amount_CNo != "NA":
            line_item_taxes_0_taxable_amount = get_line_item_taxes_0_taxable_amount(i)
            line_item_taxes_0_taxable_amount_cell = my_sheet.cell(row=i + 2,
                                                                  column=line_item_taxes_0_taxable_amount_CNo)
            if line_item_taxes_0_taxable_amount == False:
                line_item_taxes_0_taxable_amount_cell.value = line_item_taxes_0_taxable_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_item_taxes_0_taxable_amount[0])
                # line_item_taxes_0_taxable_amount_cell.value = str(dollarconverted)
                line_item_taxes_0_taxable_amount_cell.value = str(line_item_taxes_0_taxable_amount[0])

        if line_item_taxes_0_is_non_compliance_tax_CNo != "NA":
            line_item_taxes_0_is_non_compliance_tax = get_line_item_taxes_0_is_non_compliance_tax(i)
            line_item_taxes_0_is_non_compliance_tax_cell = my_sheet.cell(row=i + 2,
                                                                         column=line_item_taxes_0_is_non_compliance_tax_CNo)
            if line_item_taxes_0_is_non_compliance_tax == False:
                line_item_taxes_0_is_non_compliance_tax_cell.value = line_item_taxes_0_is_non_compliance_tax
            else:
                line_item_taxes_0_is_non_compliance_tax_cell.value = str(line_item_taxes_0_is_non_compliance_tax[0])

        if sub_total_CNo != "NA":
            sub_total = get_sub_total(i)
            sub_total_cell = my_sheet.cell(row=i + 2, column=sub_total_CNo)
            if sub_total == False:
                sub_total_cell.value = sub_total
            else:
                # dollarconverted = tzconverter.centToDollar(sub_total[0])
                # sub_total_cell.value = str(dollarconverted)
                sub_total_cell.value = str(sub_total[0])

        if linked_payments_CNo != "NA":
            linked_payments = get_linked_payments(i)
            linked_payments_cell = my_sheet.cell(row=i + 2, column=linked_payments_CNo)
            if linked_payments == False:
                linked_payments_cell.value = linked_payments
            else:
                linked_payments_cell.value = str(linked_payments[0])

        if dunning_attempts_CNo != "NA":
            dunning_attempts = get_dunning_attempts(i)
            dunning_attempts_cell = my_sheet.cell(row=i + 2, column=dunning_attempts_CNo)
            if dunning_attempts == False:
                dunning_attempts_cell.value = dunning_attempts
            else:
                dunning_attempts_cell.value = str(dunning_attempts[0])

        if applied_credits_CNo != "NA":
            applied_credits = get_applied_credits(i)
            applied_credits_cell = my_sheet.cell(row=i + 2, column=applied_credits_CNo)
            if applied_credits == False:
                applied_credits_cell.value = applied_credits
            else:
                applied_credits_cell.value = str(applied_credits[0])

        if adjustment_credit_notes_CNo != "NA":
            adjustment_credit_notes = get_adjustment_credit_notes(i)
            adjustment_credit_notes_cell = my_sheet.cell(row=i + 2, column=adjustment_credit_notes_CNo)
            if adjustment_credit_notes == False:
                adjustment_credit_notes_cell.value = adjustment_credit_notes
            else:
                adjustment_credit_notes_cell.value = str(adjustment_credit_notes[0])

        if issued_credit_notes_CNo != "NA":
            issued_credit_notes = get_issued_credit_notes(i)
            issued_credit_notes_cell = my_sheet.cell(row=i + 2, column=issued_credit_notes_CNo)
            if issued_credit_notes == False:
                issued_credit_notes_cell.value = issued_credit_notes
            else:
                issued_credit_notes_cell.value = str(issued_credit_notes[0])

        if linked_orders_CNo != "NA":
            linked_orders = get_linked_orders(i)
            linked_orders_cell = my_sheet.cell(row=i + 2, column=linked_orders_CNo)
            if linked_orders == False:
                linked_orders_cell.value = linked_orders
            else:
                linked_orders_cell.value = str(linked_orders[0])

        if billing_address_CNo != "NA":
            billing_address = get_billing_address(i)
            billing_address_cell = my_sheet.cell(row=i + 2, column=billing_address_CNo)
            if billing_address == False:
                billing_address_cell.value = billing_address
            else:
                billing_address_cell.value = str(billing_address[0])

        if billing_address_first_name_CNo != "NA":
            billing_address_first_name = get_billing_address_first_name(i)
            billing_address_first_name_cell = my_sheet.cell(row=i + 2, column=billing_address_first_name_CNo)
            if billing_address_first_name == False:
                billing_address_first_name_cell.value = billing_address_first_name
            else:
                billing_address_first_name_cell.value = str(billing_address_first_name[0])

        if billing_address_last_name_CNo != "NA":
            billing_address_last_name = get_billing_address_last_name(i)
            billing_address_last_name_cell = my_sheet.cell(row=i + 2, column=billing_address_last_name_CNo)
            if billing_address_last_name == False:
                billing_address_last_name_cell.value = billing_address_last_name
            else:
                billing_address_last_name_cell.value = str(billing_address_last_name[0])

        if billing_address_email_CNo != "NA":
            billing_address_email = get_billing_address_email(i)
            billing_address_email_cell = my_sheet.cell(row=i + 2, column=billing_address_email_CNo)
            if billing_address_email == False:
                billing_address_email_cell.value = billing_address_email
            else:
                emailvalue = billing_address_email[0]
                emailvalue = emailvalue.replace("_AT_", "@")
                emailvalue = emailvalue.replace("@example.com", "")
                billing_address_email_cell.value = str(emailvalue)

        if billing_address_company_CNo != "NA":
            billing_address_company = get_billing_address_company(i)
            billing_address_company_cell = my_sheet.cell(row=i + 2, column=billing_address_company_CNo)
            if billing_address_company == False:
                billing_address_company_cell.value = billing_address_company
            else:
                billing_address_company_cell.value = str(billing_address_company[0])

        if billing_address_phone_CNo != "NA":
            billing_address_phone = get_billing_address_phone(i)
            billing_address_phone_cell = my_sheet.cell(row=i + 2, column=billing_address_phone_CNo)
            if billing_address_phone == False:
                billing_address_phone_cell.value = billing_address_phone
            else:
                billing_address_phone_cell.value = str(billing_address_phone[0])

        if billing_address_state_CNo != "NA":
            billing_address_state = get_billing_address_state(i)
            billing_address_state_cell = my_sheet.cell(row=i + 2, column=billing_address_state_CNo)
            if billing_address_state == False:
                billing_address_state_cell.value = billing_address_state
            else:
                billing_address_state_cell.value = str(billing_address_state[0])

        if billing_address_line1_CNo != "NA":
            billing_address_line1 = get_billing_address_line1(i)
            billing_address_line1_cell = my_sheet.cell(row=i + 2, column=billing_address_line1_CNo)
            if billing_address_line1 == False:
                billing_address_line1_cell.value = billing_address_line1
            else:
                billing_address_line1_cell.value = str(billing_address_line1[0])

        if billing_address_line2_CNo != "NA":
            billing_address_line2 = get_billing_address_line2(i)
            billing_address_line2_cell = my_sheet.cell(row=i + 2, column=billing_address_line2_CNo)
            if billing_address_line2 == False:
                billing_address_line2_cell.value = billing_address_line2
            else:
                billing_address_line2_cell.value = str(billing_address_line2[0])

        if billing_address_line3_CNo != "NA":
            billing_address_line3 = get_billing_address_line3(i)
            billing_address_line3_cell = my_sheet.cell(row=i + 2, column=billing_address_line3_CNo)
            if billing_address_line3 == False:
                billing_address_line3_cell.value = billing_address_line3
            else:
                billing_address_line3_cell.value = str(billing_address_line3[0])

        if billing_address_city_CNo != "NA":
            billing_address_city = get_billing_address_city(i)
            billing_address_city_cell = my_sheet.cell(row=i + 2, column=billing_address_city_CNo)
            if billing_address_city == False:
                billing_address_city_cell.value = billing_address_city
            else:
                billing_address_city_cell.value = str(billing_address_city[0])

        if billing_address_state_code_CNo != "NA":
            billing_address_state_code = get_billing_address_state_code(i)
            billing_address_state_code_cell = my_sheet.cell(row=i + 2, column=billing_address_state_code_CNo)
            if billing_address_state_code == False:
                billing_address_state_code_cell.value = billing_address_state_code
            else:
                billing_address_state_code_cell.value = str(billing_address_state_code[0])

        if billing_address_country_CNo != "NA":
            billing_address_country = get_billing_address_country(i)
            billing_address_country_cell = my_sheet.cell(row=i + 2, column=billing_address_country_CNo)
            if billing_address_country == False:
                billing_address_country_cell.value = billing_address_country
            else:
                billing_address_country_cell.value = str(billing_address_country[0])

        if billing_address_zip_CNo != "NA":
            billing_address_zip = get_billing_address_zip(i)
            billing_address_zip_cell = my_sheet.cell(row=i + 2, column=billing_address_zip_CNo)
            if billing_address_zip == False:
                billing_address_zip_cell.value = billing_address_zip
            else:
                billing_address_zip_cell.value = str(billing_address_zip[0])

        if billing_address_validation_status_CNo != "NA":
            billing_address_validation_status = get_billing_address_validation_status(i)
            billing_address_validation_status_cell = my_sheet.cell(row=i + 2,
                                                                   column=billing_address_validation_status_CNo)
            if billing_address_validation_status == False:
                billing_address_validation_status_cell.value = billing_address_validation_status
            else:
                billing_address_validation_status_cell.value = str(billing_address_validation_status[0])

        if billing_address_object_CNo != "NA":
            billing_address_object = get_billing_address_object(i)
            billing_address_object_cell = my_sheet.cell(row=i + 2, column=billing_address_object_CNo)
            if billing_address_object == False:
                billing_address_object_cell.value = billing_address_object
            else:
                billing_address_object_cell.value = str(billing_address_object[0])

        if shipping_address_first_name_CNo != "NA":
            shipping_address_first_name = get_shipping_address_first_name(i)
            shipping_address_first_name_cell = my_sheet.cell(row=i + 2, column=shipping_address_first_name_CNo)
            if shipping_address_first_name == False:
                shipping_address_first_name_cell.value = shipping_address_first_name
            else:
                shipping_address_first_name_cell.value = str(shipping_address_first_name[0])

        if shipping_address_last_name_CNo != "NA":
            shipping_address_last_name = get_shipping_address_last_name(i)
            shipping_address_last_name_cell = my_sheet.cell(row=i + 2, column=shipping_address_last_name_CNo)
            if shipping_address_last_name == False:
                shipping_address_last_name_cell.value = shipping_address_last_name
            else:
                shipping_address_last_name_cell.value = str(shipping_address_last_name[0])

        if shipping_address_email_CNo != "NA":
            shipping_address_email = get_shipping_address_email(i)
            shipping_address_email_cell = my_sheet.cell(row=i + 2, column=shipping_address_email_CNo)
            if shipping_address_email == False:
                shipping_address_email_cell.value = shipping_address_email
            else:
                emailvalue = shipping_address_email[0]
                emailvalue = emailvalue.replace("_AT_", "@")
                emailvalue = emailvalue.replace("@example.com", "")
                shipping_address_email_cell.value = str(emailvalue)

        if shipping_address_company_CNo != "NA":
            shipping_address_company = get_shipping_address_company(i)
            shipping_address_company_cell = my_sheet.cell(row=i + 2, column=shipping_address_company_CNo)
            if shipping_address_company == False:
                shipping_address_company_cell.value = shipping_address_company
            else:
                shipping_address_company_cell.value = str(shipping_address_company[0])

        if shipping_address_phone_CNo != "NA":
            shipping_address_phone = get_shipping_address_phone(i)
            shipping_address_phone_cell = my_sheet.cell(row=i + 2, column=shipping_address_phone_CNo)
            if shipping_address_phone == False:
                shipping_address_phone_cell.value = shipping_address_phone
            else:
                shipping_address_phone_cell.value = str(shipping_address_phone[0])

        if shipping_address_line1_CNo != "NA":
            shipping_address_line1 = get_shipping_address_line1(i)
            shipping_address_line1_cell = my_sheet.cell(row=i + 2, column=shipping_address_line1_CNo)
            if shipping_address_line1 == False:
                shipping_address_line1_cell.value = shipping_address_line1
            else:
                shipping_address_line1_cell.value = str(shipping_address_line1[0])

        if shipping_address_line2_CNo != "NA":
            shipping_address_line2 = get_shipping_address_line2(i)
            shipping_address_line2_cell = my_sheet.cell(row=i + 2, column=shipping_address_line2_CNo)
            if shipping_address_line2 == False:
                shipping_address_line2_cell.value = shipping_address_line2
            else:
                shipping_address_line2_cell.value = str(shipping_address_line2[0])

        if shipping_address_line3_CNo != "NA":
            shipping_address_line3 = get_shipping_address_line3(i)
            shipping_address_line3_cell = my_sheet.cell(row=i + 2, column=shipping_address_line3_CNo)
            if shipping_address_line3 == False:
                shipping_address_line3_cell.value = shipping_address_line3
            else:
                shipping_address_line3_cell.value = str(shipping_address_line3[0])

        if shipping_address_city_CNo != "NA":
            shipping_address_city = get_shipping_address_city(i)
            shipping_address_city_cell = my_sheet.cell(row=i + 2, column=shipping_address_city_CNo)
            if shipping_address_city == False:
                shipping_address_city_cell.value = shipping_address_city
            else:
                shipping_address_city_cell.value = str(shipping_address_city[0])

        if shipping_address_state_code_CNo != "NA":
            shipping_address_state_code = get_shipping_address_state_code(i)
            shipping_address_state_code_cell = my_sheet.cell(row=i + 2, column=shipping_address_state_code_CNo)
            if shipping_address_state_code == False:
                shipping_address_state_code_cell.value = shipping_address_state_code
            else:
                shipping_address_state_code_cell.value = str(shipping_address_state_code[0])

        if shipping_address_state_CNo != "NA":
            shipping_address_state = get_shipping_address_state(i)
            shipping_address_state_cell = my_sheet.cell(row=i + 2, column=shipping_address_state_CNo)
            if shipping_address_state == False:
                shipping_address_state_cell.value = shipping_address_state
            else:
                shipping_address_state_cell.value = str(shipping_address_state[0])

        if shipping_address_zip_CNo != "NA":
            shipping_address_zip = get_shipping_address_zip(i)
            shipping_address_zip_cell = my_sheet.cell(row=i + 2, column=shipping_address_zip_CNo)
            if shipping_address_zip == False:
                shipping_address_zip_cell.value = shipping_address_zip
            else:
                shipping_address_zip_cell.value = str(shipping_address_zip[0])

        if shipping_address_country_CNo != "NA":
            shipping_address_country = get_shipping_address_country(i)
            shipping_address_country_cell = my_sheet.cell(row=i + 2, column=shipping_address_country_CNo)
            if shipping_address_country == False:
                shipping_address_country_cell.value = shipping_address_country
            else:
                shipping_address_country_cell.value = str(shipping_address_country[0])

        if shipping_address_validation_status_CNo != "NA":
            shipping_address_validation_status = get_shipping_address_validation_status(i)
            shipping_address_validation_status_cell = my_sheet.cell(row=i + 2,
                                                                    column=shipping_address_validation_status_CNo)
            if shipping_address_validation_status == False:
                shipping_address_validation_status_cell.value = shipping_address_validation_status
            else:
                shipping_address_validation_status_cell.value = str(shipping_address_validation_status[0])

        # line_items_1
        if line_items_1_id_CNo != "NA":
            line_items_1_id = get_line_items_1_id(i)
            line_items_1_id_cell = my_sheet.cell(row=i + 2, column=line_items_1_id_CNo)
            if line_items_1_id == False:
                line_items_1_id_cell.value = line_items_1_id
            else:
                line_items_1_id_cell.value = str(line_items_1_id[0])
        if line_items_1_description_CNo != "NA":
            line_items_1_description = get_line_items_1_description(i)
            line_items_1_description_cell = my_sheet.cell(row=i + 2, column=line_items_1_description_CNo)
            if line_items_1_description == False:
                line_items_1_description_cell.value = line_items_1_description
            else:
                line_items_1_description_cell.value = str(line_items_1_description[0])
        if line_items_1_date_from_CNo != "NA":
            line_items_1_date_from = get_line_items_1_date_from(i)
            line_items_1_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_1_date_from_CNo)
            if line_items_1_date_from == False:
                line_items_1_date_from_cell.value = line_items_1_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_1_date_from[0], clienttimezone)
                line_items_1_date_from_cell.value = str(modifiedtimestamp)
        if line_items_1_date_to_CNo != "NA":
            line_items_1_date_to = get_line_items_1_date_to(i)
            line_items_1_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_1_date_to_CNo)
            if line_items_1_date_to == False:
                line_items_1_date_to_cell.value = line_items_1_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_1_date_to[0], clienttimezone)
                line_items_1_date_to_cell.value = str(modifiedtimestamp)
        if line_items_1_unit_amount_CNo != "NA":
            line_items_1_unit_amount = get_line_items_1_unit_amount(i)
            line_items_1_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_1_unit_amount_CNo)
            if line_items_1_unit_amount == False:
                line_items_1_unit_amount_cell.value = line_items_1_unit_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_1_unit_amount[0])
                # line_items_1_unit_amount_cell.value = str(dollarconverted)
                line_items_1_unit_amount_cell.value = str(line_items_1_unit_amount[0])

        if line_items_1_quantity_CNo != "NA":
            line_items_1_quantity = get_line_items_1_quantity(i)
            line_items_1_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_1_quantity_CNo)
            if line_items_1_quantity == False:
                line_items_1_quantity_cell.value = line_items_1_quantity
            else:
                line_items_1_quantity_cell.value = str(line_items_1_quantity[0])
        if line_items_1_amount_CNo != "NA":
            line_items_1_amount = get_line_items_1_amount(i)
            line_items_1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_1_amount_CNo)
            if line_items_1_amount == False:
                line_items_1_amount_cell.value = line_items_1_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_1_amount[0])
                line_items_1_amount_cell.value = str(dollarconverted)
                # line_items_1_amount_cell.value = str(line_items_1_amount[0])

        if line_items_1_entity_type_CNo != "NA":
            line_items_1_entity_type = get_line_items_1_entity_type(i)
            line_items_1_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_1_entity_type_CNo)
            if line_items_1_entity_type == False:
                line_items_1_entity_type_cell.value = line_items_1_entity_type
            else:
                line_items_1_entity_type_cell.value = str(line_items_1_entity_type[0])
        if line_items_1_entity_id_CNo != "NA":
            line_items_1_entity_id = get_line_items_1_entity_id(i)
            line_items_1_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_1_entity_id_CNo)
            if line_items_1_entity_id == False:
                line_items_1_entity_id_cell.value = line_items_1_entity_id
            else:
                line_items_1_entity_id_cell.value = str(line_items_1_entity_id[0])
        if line_items_1_tax1_name_CNo != "NA":
            line_items_1_tax1_name = get_line_items_1_tax1_name(i)
            line_items_1_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_1_tax1_name_CNo)
            if line_items_1_tax1_name == False:
                line_items_1_tax1_name_cell.value = line_items_1_tax1_name
            else:
                line_items_1_tax1_name_cell.value = str(line_items_1_tax1_name[0])
        if line_items_1_tax1_amount_CNo != "NA":
            line_items_1_tax1_amount = get_line_items_1_tax1_amount(i)
            line_items_1_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_1_tax1_amount_CNo)
            if line_items_1_tax1_amount == False:
                line_items_1_tax1_amount_cell.value = line_items_1_tax1_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_1_tax1_amount[0])
                # line_items_1_tax1_amount_cell.value = str(dollarconverted)
                line_items_1_tax1_amount_cell.value = str(line_items_1_tax1_amount[0])

        # line_items_2
        if line_items_2_id_CNo != "NA":
            line_items_2_id = get_line_items_2_id(i)
            line_items_2_id_cell = my_sheet.cell(row=i + 2, column=line_items_2_id_CNo)
            if line_items_2_id == False:
                line_items_2_id_cell.value = line_items_2_id
            else:
                line_items_2_id_cell.value = str(line_items_2_id[0])
        if line_items_2_description_CNo != "NA":
            line_items_2_description = get_line_items_2_description(i)
            line_items_2_description_cell = my_sheet.cell(row=i + 2, column=line_items_2_description_CNo)
            if line_items_2_description == False:
                line_items_2_description_cell.value = line_items_2_description
            else:
                line_items_2_description_cell.value = str(line_items_2_description[0])
        if line_items_2_date_from_CNo != "NA":
            line_items_2_date_from = get_line_items_2_date_from(i)
            line_items_2_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_2_date_from_CNo)
            if line_items_2_date_from == False:
                line_items_2_date_from_cell.value = line_items_2_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_2_date_from[0], clienttimezone)
                line_items_2_date_from_cell.value = str(modifiedtimestamp)
        if line_items_2_date_to_CNo != "NA":
            line_items_2_date_to = get_line_items_2_date_to(i)
            line_items_2_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_2_date_to_CNo)
            if line_items_2_date_to == False:
                line_items_2_date_to_cell.value = line_items_2_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_2_date_to[0], clienttimezone)
                line_items_2_date_to_cell.value = str(modifiedtimestamp)
        if line_items_2_unit_amount_CNo != "NA":
            line_items_2_unit_amount = get_line_items_2_unit_amount(i)
            line_items_2_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_2_unit_amount_CNo)
            if line_items_2_unit_amount == False:
                line_items_2_unit_amount_cell.value = line_items_2_unit_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_2_unit_amount[0])
                # line_items_2_unit_amount_cell.value = str(dollarconverted)
                line_items_2_unit_amount_cell.value = str(line_items_2_unit_amount[0])

        if line_items_2_quantity_CNo != "NA":
            line_items_2_quantity = get_line_items_2_quantity(i)
            line_items_2_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_2_quantity_CNo)
            if line_items_2_quantity == False:
                line_items_2_quantity_cell.value = line_items_2_quantity
            else:
                line_items_2_quantity_cell.value = str(line_items_2_quantity[0])
        if line_items_2_amount_CNo != "NA":
            line_items_2_amount = get_line_items_2_amount(i)
            line_items_2_amount_cell = my_sheet.cell(row=i + 2, column=line_items_2_amount_CNo)
            if line_items_2_amount == False:
                line_items_2_amount_cell.value = line_items_2_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_2_amount[0])
                line_items_2_amount_cell.value = str(dollarconverted)
                # line_items_2_amount_cell.value = str(line_items_2_amount[0])

        if line_items_2_entity_type_CNo != "NA":
            line_items_2_entity_type = get_line_items_2_entity_type(i)
            line_items_2_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_2_entity_type_CNo)
            if line_items_2_entity_type == False:
                line_items_2_entity_type_cell.value = line_items_2_entity_type
            else:
                line_items_2_entity_type_cell.value = str(line_items_2_entity_type[0])
        if line_items_2_entity_id_CNo != "NA":
            line_items_2_entity_id = get_line_items_2_entity_id(i)
            line_items_2_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_2_entity_id_CNo)
            if line_items_2_entity_id == False:
                line_items_2_entity_id_cell.value = line_items_2_entity_id
            else:
                line_items_2_entity_id_cell.value = str(line_items_2_entity_id[0])
        if line_items_2_tax1_name_CNo != "NA":
            line_items_2_tax1_name = get_line_items_2_tax1_name(i)
            line_items_2_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_2_tax1_name_CNo)
            if line_items_2_tax1_name == False:
                line_items_2_tax1_name_cell.value = line_items_2_tax1_name
            else:
                line_items_2_tax1_name_cell.value = str(line_items_2_tax1_name[0])
        if line_items_2_tax1_amount_CNo != "NA":
            line_items_2_tax1_amount = get_line_items_2_tax1_amount(i)
            line_items_2_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_2_tax1_amount_CNo)
            if line_items_2_tax1_amount == False:
                line_items_2_tax1_amount_cell.value = line_items_2_tax1_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_2_tax1_amount[0])
                # line_items_2_tax1_amount_cell.value = str(dollarconverted)
                line_items_2_tax1_amount_cell.value = str(line_items_2_tax1_amount[0])

        # line_items_3
        if line_items_3_id_CNo != "NA":
            line_items_3_id = get_line_items_3_id(i)
            line_items_3_id_cell = my_sheet.cell(row=i + 2, column=line_items_3_id_CNo)
            if line_items_3_id == False:
                line_items_3_id_cell.value = line_items_3_id
            else:
                line_items_3_id_cell.value = str(line_items_3_id[0])
        if line_items_3_description_CNo != "NA":
            line_items_3_description = get_line_items_3_description(i)
            line_items_3_description_cell = my_sheet.cell(row=i + 2, column=line_items_3_description_CNo)
            if line_items_3_description == False:
                line_items_3_description_cell.value = line_items_3_description
            else:
                line_items_3_description_cell.value = str(line_items_3_description[0])
        if line_items_3_date_from_CNo != "NA":
            line_items_3_date_from = get_line_items_3_date_from(i)
            line_items_3_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_3_date_from_CNo)
            if line_items_3_date_from == False:
                line_items_3_date_from_cell.value = line_items_3_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_3_date_from[0], clienttimezone)
                line_items_3_date_from_cell.value = str(modifiedtimestamp)
        if line_items_3_date_to_CNo != "NA":
            line_items_3_date_to = get_line_items_3_date_to(i)
            line_items_3_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_3_date_to_CNo)
            if line_items_3_date_to == False:
                line_items_3_date_to_cell.value = line_items_3_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_3_date_to[0], clienttimezone)
                line_items_3_date_to_cell.value = str(modifiedtimestamp)
        if line_items_3_unit_amount_CNo != "NA":
            line_items_3_unit_amount = get_line_items_3_unit_amount(i)
            line_items_3_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_3_unit_amount_CNo)
            if line_items_3_unit_amount == False:
                line_items_3_unit_amount_cell.value = line_items_3_unit_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_3_unit_amount[0])
                # line_items_3_unit_amount_cell.value = str(dollarconverted)
                line_items_3_unit_amount_cell.value = str(line_items_3_unit_amount[0])

        if line_items_3_quantity_CNo != "NA":
            line_items_3_quantity = get_line_items_3_quantity(i)
            line_items_3_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_3_quantity_CNo)
            if line_items_3_quantity == False:
                line_items_3_quantity_cell.value = line_items_3_quantity
            else:
                line_items_3_quantity_cell.value = str(line_items_3_quantity[0])
        if line_items_3_amount_CNo != "NA":
            line_items_3_amount = get_line_items_3_amount(i)
            line_items_3_amount_cell = my_sheet.cell(row=i + 2, column=line_items_3_amount_CNo)
            if line_items_3_amount == False:
                line_items_3_amount_cell.value = line_items_3_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_3_amount[0])
                line_items_3_amount_cell.value = str(dollarconverted)
                # line_items_3_amount_cell.value = str(line_items_3_amount[0])

        if line_items_3_entity_type_CNo != "NA":
            line_items_3_entity_type = get_line_items_3_entity_type(i)
            line_items_3_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_3_entity_type_CNo)
            if line_items_3_entity_type == False:
                line_items_3_entity_type_cell.value = line_items_3_entity_type
            else:
                line_items_3_entity_type_cell.value = str(line_items_3_entity_type[0])
        if line_items_3_entity_id_CNo != "NA":
            line_items_3_entity_id = get_line_items_3_entity_id(i)
            line_items_3_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_3_entity_id_CNo)
            if line_items_3_entity_id == False:
                line_items_3_entity_id_cell.value = line_items_3_entity_id
            else:
                line_items_3_entity_id_cell.value = str(line_items_3_entity_id[0])
        if line_items_3_tax1_name_CNo != "NA":
            line_items_3_tax1_name = get_line_items_3_tax1_name(i)
            line_items_3_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_3_tax1_name_CNo)
            if line_items_3_tax1_name == False:
                line_items_3_tax1_name_cell.value = line_items_3_tax1_name
            else:
                line_items_3_tax1_name_cell.value = str(line_items_3_tax1_name[0])
        if line_items_3_tax1_amount_CNo != "NA":
            line_items_3_tax1_amount = get_line_items_3_tax1_amount(i)
            line_items_3_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_3_tax1_amount_CNo)
            if line_items_3_tax1_amount == False:
                line_items_3_tax1_amount_cell.value = line_items_3_tax1_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_3_tax1_amount[0])
                # line_items_3_tax1_amount_cell.value = str(dollarconverted)
                line_items_3_tax1_amount_cell.value = str(line_items_3_tax1_amount[0])

        # line_items_4
        if line_items_4_id_CNo != "NA":
            line_items_4_id = get_line_items_4_id(i)
            line_items_4_id_cell = my_sheet.cell(row=i + 2, column=line_items_4_id_CNo)
            if line_items_4_id == False:
                line_items_4_id_cell.value = line_items_4_id
            else:
                line_items_4_id_cell.value = str(line_items_4_id[0])
        if line_items_4_description_CNo != "NA":
            line_items_4_description = get_line_items_4_description(i)
            line_items_4_description_cell = my_sheet.cell(row=i + 2, column=line_items_4_description_CNo)
            if line_items_4_description == False:
                line_items_4_description_cell.value = line_items_4_description
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_4_date_from[0], clienttimezone)
                line_items_4_date_from_cell.value = str(modifiedtimestamp)
        if line_items_4_date_from_CNo != "NA":
            line_items_4_date_from = get_line_items_4_date_from(i)
            line_items_4_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_4_date_from_CNo)
            if line_items_4_date_from == False:
                line_items_4_date_from_cell.value = line_items_4_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_4_date_to[0], clienttimezone)
                line_items_4_date_to_cell.value = str(modifiedtimestamp)
        if line_items_4_date_to_CNo != "NA":
            line_items_4_date_to = get_line_items_4_date_to(i)
            line_items_4_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_4_date_to_CNo)
            if line_items_4_date_to == False:
                line_items_4_date_to_cell.value = line_items_4_date_to
            else:
                line_items_4_date_to_cell.value = str(line_items_4_date_to[0])
        if line_items_4_unit_amount_CNo != "NA":
            line_items_4_unit_amount = get_line_items_4_unit_amount(i)
            line_items_4_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_4_unit_amount_CNo)
            if line_items_4_unit_amount == False:
                line_items_4_unit_amount_cell.value = line_items_4_unit_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_4_unit_amount[0])
                # line_items_4_unit_amount_cell.value = str(dollarconverted)
                line_items_4_unit_amount_cell.value = str(line_items_4_unit_amount[0])

        if line_items_4_quantity_CNo != "NA":
            line_items_4_quantity = get_line_items_4_quantity(i)
            line_items_4_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_4_quantity_CNo)
            if line_items_4_quantity == False:
                line_items_4_quantity_cell.value = line_items_4_quantity
            else:
                line_items_4_quantity_cell.value = str(line_items_4_quantity[0])
        if line_items_4_amount_CNo != "NA":
            line_items_4_amount = get_line_items_4_amount(i)
            line_items_4_amount_cell = my_sheet.cell(row=i + 2, column=line_items_4_amount_CNo)
            if line_items_4_amount == False:
                line_items_4_amount_cell.value = line_items_4_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_4_amount[0])
                line_items_4_amount_cell.value = str(dollarconverted)
                # line_items_4_amount_cell.value = str(line_items_4_amount[0])

        if line_items_4_entity_type_CNo != "NA":
            line_items_4_entity_type = get_line_items_4_entity_type(i)
            line_items_4_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_4_entity_type_CNo)
            if line_items_4_entity_type == False:
                line_items_4_entity_type_cell.value = line_items_4_entity_type
            else:
                line_items_4_entity_type_cell.value = str(line_items_4_entity_type[0])
        if line_items_4_entity_id_CNo != "NA":
            line_items_4_entity_id = get_line_items_4_entity_id(i)
            line_items_4_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_4_entity_id_CNo)
            if line_items_4_entity_id == False:
                line_items_4_entity_id_cell.value = line_items_4_entity_id
            else:
                line_items_4_entity_id_cell.value = str(line_items_4_entity_id[0])
        if line_items_4_tax1_name_CNo != "NA":
            line_items_4_tax1_name = get_line_items_4_tax1_name(i)
            line_items_4_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_4_tax1_name_CNo)
            if line_items_4_tax1_name == False:
                line_items_4_tax1_name_cell.value = line_items_4_tax1_name
            else:
                line_items_4_tax1_name_cell.value = str(line_items_4_tax1_name[0])
        if line_items_4_tax1_amount_CNo != "NA":
            line_items_4_tax1_amount = get_line_items_4_tax1_amount(i)
            line_items_4_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_4_tax1_amount_CNo)
            if line_items_4_tax1_amount == False:
                line_items_4_tax1_amount_cell.value = line_items_4_tax1_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_4_tax1_amount[0])
                # line_items_4_tax1_amount_cell.value = str(dollarconverted)
                line_items_4_tax1_amount_cell.value = str(line_items_4_tax1_amount[0])

        # line_items_5
        if line_items_5_id_CNo != "NA":
            line_items_5_id = get_line_items_5_id(i)
            line_items_5_id_cell = my_sheet.cell(row=i + 2, column=line_items_5_id_CNo)
            if line_items_5_id == False:
                line_items_5_id_cell.value = line_items_5_id
            else:
                line_items_5_id_cell.value = str(line_items_5_id[0])
        if line_items_5_description_CNo != "NA":
            line_items_5_description = get_line_items_5_description(i)
            line_items_5_description_cell = my_sheet.cell(row=i + 2, column=line_items_5_description_CNo)
            if line_items_5_description == False:
                line_items_5_description_cell.value = line_items_5_description
            else:
                line_items_5_description_cell.value = str(line_items_5_description[0])
        if line_items_5_date_from_CNo != "NA":
            line_items_5_date_from = get_line_items_5_date_from(i)
            line_items_5_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_5_date_from_CNo)
            if line_items_5_date_from == False:
                line_items_5_date_from_cell.value = line_items_5_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_5_date_from[0], clienttimezone)
                line_items_5_date_from_cell.value = str(modifiedtimestamp)
        if line_items_5_date_to_CNo != "NA":
            line_items_5_date_to = get_line_items_5_date_to(i)
            line_items_5_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_5_date_to_CNo)
            if line_items_5_date_to == False:
                line_items_5_date_to_cell.value = line_items_5_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_5_date_to[0], clienttimezone)
                line_items_5_date_to_cell.value = str(modifiedtimestamp)
        if line_items_5_unit_amount_CNo != "NA":
            line_items_5_unit_amount = get_line_items_5_unit_amount(i)
            line_items_5_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_5_unit_amount_CNo)
            if line_items_5_unit_amount == False:
                line_items_5_unit_amount_cell.value = line_items_5_unit_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_5_unit_amount[0])
                # line_items_5_unit_amount_cell.value = str(dollarconverted)
                line_items_5_unit_amount_cell.value = str(line_items_5_unit_amount[0])

        if line_items_5_quantity_CNo != "NA":
            line_items_5_quantity = get_line_items_5_quantity(i)
            line_items_5_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_5_quantity_CNo)
            if line_items_5_quantity == False:
                line_items_5_quantity_cell.value = line_items_5_quantity
            else:
                line_items_5_quantity_cell.value = str(line_items_5_quantity[0])
        if line_items_5_amount_CNo != "NA":
            line_items_5_amount = get_line_items_5_amount(i)
            line_items_5_amount_cell = my_sheet.cell(row=i + 2, column=line_items_5_amount_CNo)
            if line_items_5_amount == False:
                line_items_5_amount_cell.value = line_items_5_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_5_amount[0])
                line_items_5_amount_cell.value = str(dollarconverted)
                # line_items_5_amount_cell.value = str(line_items_5_amount[0])

        if line_items_5_entity_type_CNo != "NA":
            line_items_5_entity_type = get_line_items_5_entity_type(i)
            line_items_5_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_5_entity_type_CNo)
            if line_items_5_entity_type == False:
                line_items_5_entity_type_cell.value = line_items_5_entity_type
            else:
                line_items_5_entity_type_cell.value = str(line_items_5_entity_type[0])
        if line_items_5_entity_id_CNo != "NA":
            line_items_5_entity_id = get_line_items_5_entity_id(i)
            line_items_5_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_5_entity_id_CNo)
            if line_items_5_entity_id == False:
                line_items_5_entity_id_cell.value = line_items_5_entity_id
            else:
                line_items_5_entity_id_cell.value = str(line_items_5_entity_id[0])
        if line_items_5_tax1_name_CNo != "NA":
            line_items_5_tax1_name = get_line_items_5_tax1_name(i)
            line_items_5_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_5_tax1_name_CNo)
            if line_items_5_tax1_name == False:
                line_items_5_tax1_name_cell.value = line_items_5_tax1_name
            else:
                line_items_5_tax1_name_cell.value = str(line_items_5_tax1_name[0])
        if line_items_5_tax1_amount_CNo != "NA":
            line_items_5_tax1_amount = get_line_items_5_tax1_amount(i)
            line_items_5_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_5_tax1_amount_CNo)
            if line_items_5_tax1_amount == False:
                line_items_5_tax1_amount_cell.value = line_items_5_tax1_amount
            else:
                # dollarconverted = tzconverter.centToDollar(line_items_5_tax1_amount[0])
                # line_items_5_tax1_amount_cell.value = str(dollarconverted)
                line_items_5_tax1_amount_cell.value = str(line_items_5_tax1_amount[0])

        # line_items_6
        if line_items_6_id_CNo != "NA":
            line_items_6_id = get_line_items_6_id(i)
            line_items_6_id_cell = my_sheet.cell(row=i + 2, column=line_items_6_id_CNo)
            if line_items_6_id == False:
                line_items_6_id_cell.value = line_items_6_id
            else:
                line_items_6_id_cell.value = str(line_items_6_id[0])
        if line_items_6_description_CNo != "NA":
            line_items_6_description = get_line_items_6_description(i)
            line_items_6_description_cell = my_sheet.cell(row=i + 2, column=line_items_6_description_CNo)
            if line_items_6_description == False:
                line_items_6_description_cell.value = line_items_6_description
            else:
                line_items_6_description_cell.value = str(line_items_6_description[0])
        if line_items_6_date_from_CNo != "NA":
            line_items_6_date_from = get_line_items_6_date_from(i)
            line_items_6_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_6_date_from_CNo)
            if line_items_6_date_from == False:
                line_items_6_date_from_cell.value = line_items_6_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_6_date_from[0], clienttimezone)
                line_items_6_date_from_cell.value = str(modifiedtimestamp)
        if line_items_6_date_to_CNo != "NA":
            line_items_6_date_to = get_line_items_6_date_to(i)
            line_items_6_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_6_date_to_CNo)
            if line_items_6_date_to == False:
                line_items_6_date_to_cell.value = line_items_6_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_6_date_to[0], clienttimezone)
                line_items_6_date_to_cell.value = str(modifiedtimestamp)
        if line_items_6_unit_amount_CNo != "NA":
            line_items_6_unit_amount = get_line_items_6_unit_amount(i)
            line_items_6_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_6_unit_amount_CNo)
            if line_items_6_unit_amount == False:
                line_items_6_unit_amount_cell.value = line_items_6_unit_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_6_unit_amount[0])
                line_items_6_unit_amount_cell.value = str(dollarconverted)

        if line_items_6_quantity_CNo != "NA":
            line_items_6_quantity = get_line_items_6_quantity(i)
            line_items_6_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_6_quantity_CNo)
            if line_items_6_quantity == False:
                line_items_6_quantity_cell.value = line_items_6_quantity
            else:
                line_items_6_quantity_cell.value = str(line_items_6_quantity[0])
        if line_items_6_amount_CNo != "NA":
            line_items_6_amount = get_line_items_6_amount(i)
            line_items_6_amount_cell = my_sheet.cell(row=i + 2, column=line_items_6_amount_CNo)
            if line_items_6_amount == False:
                line_items_6_amount_cell.value = line_items_6_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_6_amount[0])
                line_items_6_amount_cell.value = str(dollarconverted)
        if line_items_6_entity_type_CNo != "NA":
            line_items_6_entity_type = get_line_items_6_entity_type(i)
            line_items_6_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_6_entity_type_CNo)
            if line_items_6_entity_type == False:
                line_items_6_entity_type_cell.value = line_items_6_entity_type
            else:
                line_items_6_entity_type_cell.value = str(line_items_6_entity_type[0])
        if line_items_6_entity_id_CNo != "NA":
            line_items_6_entity_id = get_line_items_6_entity_id(i)
            line_items_6_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_6_entity_id_CNo)
            if line_items_6_entity_id == False:
                line_items_6_entity_id_cell.value = line_items_6_entity_id
            else:
                line_items_6_entity_id_cell.value = str(line_items_6_entity_id[0])
        if line_items_6_tax1_name_CNo != "NA":
            line_items_6_tax1_name = get_line_items_6_tax1_name(i)
            line_items_6_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_6_tax1_name_CNo)
            if line_items_6_tax1_name == False:
                line_items_6_tax1_name_cell.value = line_items_6_tax1_name
            else:
                line_items_6_tax1_name_cell.value = str(line_items_6_tax1_name[0])
        if line_items_6_tax1_amount_CNo != "NA":
            line_items_6_tax1_amount = get_line_items_6_tax1_amount(i)
            line_items_6_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_6_tax1_amount_CNo)
            if line_items_6_tax1_amount == False:
                line_items_6_tax1_amount_cell.value = line_items_6_tax1_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_6_tax1_amount[0])
                line_items_6_tax1_amount_cell.value = str(dollarconverted)

        # line_items_7
        if line_items_7_id_CNo != "NA":
            line_items_7_id = get_line_items_7_id(i)
            line_items_7_id_cell = my_sheet.cell(row=i + 2, column=line_items_7_id_CNo)
            if line_items_7_id == False:
                line_items_7_id_cell.value = line_items_7_id
            else:
                line_items_7_id_cell.value = str(line_items_7_id[0])
        if line_items_7_description_CNo != "NA":
            line_items_7_description = get_line_items_7_description(i)
            line_items_7_description_cell = my_sheet.cell(row=i + 2, column=line_items_7_description_CNo)
            if line_items_7_description == False:
                line_items_7_description_cell.value = line_items_7_description
            else:
                line_items_7_description_cell.value = str(line_items_7_description[0])
        if line_items_7_date_from_CNo != "NA":
            line_items_7_date_from = get_line_items_7_date_from(i)
            line_items_7_date_from_cell = my_sheet.cell(row=i + 2, column=line_items_7_date_from_CNo)
            if line_items_7_date_from == False:
                line_items_7_date_from_cell.value = line_items_7_date_from
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_7_date_from[0], clienttimezone)
                line_items_7_date_from_cell.value = str(modifiedtimestamp)
        if line_items_7_date_to_CNo != "NA":
            line_items_7_date_to = get_line_items_7_date_to(i)
            line_items_7_date_to_cell = my_sheet.cell(row=i + 2, column=line_items_7_date_to_CNo)
            if line_items_7_date_to == False:
                line_items_7_date_to_cell.value = line_items_7_date_to
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(line_items_7_date_to[0], clienttimezone)
                line_items_7_date_to_cell.value = str(modifiedtimestamp)
        if line_items_7_unit_amount_CNo != "NA":
            line_items_7_unit_amount = get_line_items_7_unit_amount(i)
            line_items_7_unit_amount_cell = my_sheet.cell(row=i + 2, column=line_items_7_unit_amount_CNo)
            if line_items_7_unit_amount == False:
                line_items_7_unit_amount_cell.value = line_items_7_unit_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_7_unit_amount[0])
                line_items_7_unit_amount_cell.value = str(dollarconverted)
        if line_items_7_quantity_CNo != "NA":
            line_items_7_quantity = get_line_items_7_quantity(i)
            line_items_7_quantity_cell = my_sheet.cell(row=i + 2, column=line_items_7_quantity_CNo)
            if line_items_7_quantity == False:
                line_items_7_quantity_cell.value = line_items_7_quantity
            else:
                line_items_7_quantity_cell.value = str(line_items_7_quantity[0])
        if line_items_7_amount_CNo != "NA":
            line_items_7_amount = get_line_items_7_amount(i)
            line_items_7_amount_cell = my_sheet.cell(row=i + 2, column=line_items_7_amount_CNo)
            if line_items_7_amount == False:
                line_items_7_amount_cell.value = line_items_7_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_7_amount[0])
                line_items_7_amount_cell.value = str(dollarconverted)
        if line_items_7_entity_type_CNo != "NA":
            line_items_7_entity_type = get_line_items_7_entity_type(i)
            line_items_7_entity_type_cell = my_sheet.cell(row=i + 2, column=line_items_7_entity_type_CNo)
            if line_items_7_entity_type == False:
                line_items_7_entity_type_cell.value = line_items_7_entity_type
            else:
                line_items_7_entity_type_cell.value = str(line_items_7_entity_type[0])
        if line_items_7_entity_id_CNo != "NA":
            line_items_7_entity_id = get_line_items_7_entity_id(i)
            line_items_7_entity_id_cell = my_sheet.cell(row=i + 2, column=line_items_7_entity_id_CNo)
            if line_items_7_entity_id == False:
                line_items_7_entity_id_cell.value = line_items_7_entity_id
            else:
                line_items_7_entity_id_cell.value = str(line_items_7_entity_id[0])
        if line_items_7_tax1_name_CNo != "NA":
            line_items_7_tax1_name = get_line_items_7_tax1_name(i)
            line_items_7_tax1_name_cell = my_sheet.cell(row=i + 2, column=line_items_7_tax1_name_CNo)
            if line_items_7_tax1_name == False:
                line_items_7_tax1_name_cell.value = line_items_7_tax1_name
            else:
                line_items_7_tax1_name_cell.value = str(line_items_7_tax1_name[0])
        if line_items_7_tax1_amount_CNo != "NA":
            line_items_7_tax1_amount = get_line_items_7_tax1_amount(i)
            line_items_7_tax1_amount_cell = my_sheet.cell(row=i + 2, column=line_items_7_tax1_amount_CNo)
            if line_items_7_tax1_amount == False:
                line_items_7_tax1_amount_cell.value = line_items_7_tax1_amount
            else:
                dollarconverted = tzconverter.centToDollar(line_items_7_tax1_amount[0])
                line_items_7_tax1_amount_cell.value = str(dollarconverted)

    except Exception as e:
        print("in exception", e)

# finally save excel
my_wb.save(output_file_name)
print("Execution completed")
end = timer()
print('Total time taken: ', end - start, ' seconds')
