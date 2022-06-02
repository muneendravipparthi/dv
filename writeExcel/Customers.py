import json
import os
from timeit import default_timer as timer

import jsonpath
import openpyxl
from jproperties import Properties

from utils.Ephoc2DateTime import Ephoc2DateTime as tzconverter

configs = Properties()
ROOT_DIR1 = os.path.abspath(os.curdir)
ROOT_DIR = ROOT_DIR1.replace('writeExcel', 'configuration.properties')
with open(ROOT_DIR, 'rb') as config_file:
    configs.load(config_file)
Json_DIR = ROOT_DIR1.replace('writeExcel', 'readAPI/jsonfiles/')
start = timer()
input_jsonfile = Json_DIR + configs.get("clientName").data + "_AllCustomers.json"
output_file_name = ROOT_DIR1 + '/outputExcelFiles/' + configs.get("clientName").data + "_Customer_Actual.xlsx"
clienttimezone = configs.get("clienttimezone").data

# ------------------Coulmn_nos----------------------#
id_CNo = 1
created_at_CNo = 2
first_name_CNo = 3
last_name_CNo = 4
email_CNo = 5
phone_CNo = 6
meta_data_CNo = 7
billing_address_first_name_CNo = 8
billing_address_last_name_CNo = 9
billing_address_email_CNo = 10
billing_address_line1_CNo = 11
billing_address_line2_CNo = 12
billing_address_city_CNo = 13
billing_address_state_CNo = 14
billing_address_zip_CNo = 15
billing_address_country_CNo = 16

preferred_currency_code_CNo = 'NA'
locale_CNo = 'NA'
company_CNo = 'NA'
payment_method_type_CNo = 'NA'
payment_method_gateway_account_id_CNo = 'NA'
payment_method_reference_id_CNo = 'NA'
allow_direct_debit_CNo = 'NA'
auto_collection_CNo = 'NA'
taxability_CNo = 'NA'
vat_number_CNo = 'NA'
net_term_days_CNo = 'NA'
consolidated_invoicing_CNo = 'NA'
invoice_notes_CNo = 'NA'
billing_address_company_CNo = 'NA'
billing_address_phone_CNo = 'NA'
billing_address_line3_CNo = 'NA'
billing_address_state_code_CNo = 'NA'
registered_for_gst_CNo = 'NA'
entity_code_CNo = 'NA'
exempt_number_CNo = 'NA'
card_last4_CNo = 'NA'
card_brand_CNo = 'NA'
card_funding_CNo = 'NA'
card_exp_month_CNo = 'NA'
card_exp_year_CNo = 'NA'
card_name_CNo = 'NA'
payment_method_tmp_token_CNo = 'NA'
billing_address_validation_status_CNo = 'NA'
cf_account_number_CNo = 'NA'
cf_xero_id_CNo = 'NA'
cf_xero_name_CNo = 'NA'

jsondata = open(input_jsonfile)
Customerdictionary = json.load(jsondata)
print("Json File", Customerdictionary)

apitojsontime = timer()

jsonpathres = Customerdictionary
l = jsonpath.jsonpath(jsonpathres, "list")
totalRecordCountInResp = len(l[0])
# create work book and sheet object
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "Customers"

# ---------------------------write headers in excel -- first row---------------------------#

if id_CNo != "NA":
    id = my_sheet.cell(row=1, column=id_CNo)
    id.value = str("id")

if first_name_CNo != "NA":
    first_name = my_sheet.cell(row=1, column=first_name_CNo)
    first_name.value = str("first_name")

if last_name_CNo != "NA":
    last_name = my_sheet.cell(row=1, column=last_name_CNo)
    last_name.value = str("last_name")

if email_CNo != "NA":
    email = my_sheet.cell(row=1, column=email_CNo)
    email.value = str("email")

if preferred_currency_code_CNo != "NA":
    preferred_currency_code = my_sheet.cell(row=1, column=preferred_currency_code_CNo)
    preferred_currency_code.value = str("preferred_currency_code")

if phone_CNo != "NA":
    phone = my_sheet.cell(row=1, column=phone_CNo)
    phone.value = str("phone")

if company_CNo != "NA":
    company = my_sheet.cell(row=1, column=company_CNo)
    company.value = str("company")

if auto_collection_CNo != "NA":
    auto_collection = my_sheet.cell(row=1, column=auto_collection_CNo)
    auto_collection.value = str("auto_collection")

if net_term_days_CNo != "NA":
    net_term_days = my_sheet.cell(row=1, column=net_term_days_CNo)
    net_term_days.value = str("net_term_days")

if allow_direct_debit_CNo != "NA":
    allow_direct_debit = my_sheet.cell(row=1, column=allow_direct_debit_CNo)
    allow_direct_debit.value = str("allow_direct_debit")

if vat_number_CNo != "NA":
    vat_number = my_sheet.cell(row=1, column=vat_number_CNo)
    vat_number.value = str("vat_number")

if registered_for_gst_CNo != "NA":
    registered_for_gst = my_sheet.cell(row=1, column=registered_for_gst_CNo)
    registered_for_gst.value = str("registered_for_gst")

if taxability_CNo != "NA":
    taxability = my_sheet.cell(row=1, column=taxability_CNo)
    taxability.value = str("taxability")

if locale_CNo != "NA":
    locale = my_sheet.cell(row=1, column=locale_CNo)
    locale.value = str("locale")

if entity_code_CNo != "NA":
    entity_code = my_sheet.cell(row=1, column=entity_code_CNo)
    entity_code.value = str("entity_code")

if exempt_number_CNo != "NA":
    exempt_number = my_sheet.cell(row=1, column=exempt_number_CNo)
    exempt_number.value = str("exempt_number")

if meta_data_CNo != "NA":
    meta_data = my_sheet.cell(row=1, column=meta_data_CNo)
    meta_data.value = str("meta_data")

if consolidated_invoicing_CNo != "NA":
    consolidated_invoicing = my_sheet.cell(row=1, column=consolidated_invoicing_CNo)
    consolidated_invoicing.value = str("consolidated_invoicing")

if invoice_notes_CNo != "NA":
    invoice_notes = my_sheet.cell(row=1, column=invoice_notes_CNo)
    invoice_notes.value = str("invoice_notes")

if payment_method_type_CNo != "NA":
    payment_method_type = my_sheet.cell(row=1, column=payment_method_type_CNo)
    payment_method_type.value = str("payment_method_type")

if card_last4_CNo != "NA":
    card_last4 = my_sheet.cell(row=1, column=card_last4_CNo)
    card_last4.value = str("card_last4")

if card_brand_CNo != "NA":
    card_brand = my_sheet.cell(row=1, column=card_brand_CNo)
    card_brand.value = str("card_brand")

if card_funding_CNo != "NA":
    card_funding = my_sheet.cell(row=1, column=card_funding_CNo)
    card_funding.value = str("card_funding")

if card_exp_month_CNo != "NA":
    card_exp_month = my_sheet.cell(row=1, column=card_exp_month_CNo)
    card_exp_month.value = str("card_exp_month")

if card_exp_year_CNo != "NA":
    card_exp_year = my_sheet.cell(row=1, column=card_exp_year_CNo)
    card_exp_year.value = str("card_exp_year")

if card_name_CNo != "NA":
    card_name = my_sheet.cell(row=1, column=card_name_CNo)
    card_name.value = str("card_name")

if payment_method_gateway_account_id_CNo != "NA":
    payment_method_gateway_account_id = my_sheet.cell(row=1, column=payment_method_gateway_account_id_CNo)
    payment_method_gateway_account_id.value = str("payment_method_gateway_account_id")

if payment_method_reference_id_CNo != "NA":
    payment_method_reference_id = my_sheet.cell(row=1, column=payment_method_reference_id_CNo)
    payment_method_reference_id.value = str("payment_method_reference_id")

if payment_method_tmp_token_CNo != "NA":
    payment_method_tmp_token = my_sheet.cell(row=1, column=payment_method_tmp_token_CNo)
    payment_method_tmp_token.value = str("payment_method_tmp_token")

if billing_address_first_name_CNo != "NA":
    billing_address_first_name = my_sheet.cell(row=1, column=billing_address_first_name_CNo)
    billing_address_first_name.value = str("billing_address_first_name")

if billing_address_last_name_CNo != "NA":
    billing_address_last_name = my_sheet.cell(row=1, column=billing_address_last_name_CNo)
    billing_address_last_name.value = str("billing_address_last_name")

if billing_address_email_CNo != "NA":
    billing_address_email = my_sheet.cell(row=1, column=billing_address_email_CNo)
    billing_address_email.value = str("billing_address_email")

if billing_address_company_CNo != "NA":
    billing_address_company = my_sheet.cell(row=1, column=billing_address_company_CNo)
    billing_address_company.value = str("billing_address_company")

if billing_address_phone_CNo != "NA":
    billing_address_phone = my_sheet.cell(row=1, column=billing_address_phone_CNo)
    billing_address_phone.value = str("billing_address_phone")

if billing_address_line1_CNo != "NA":
    billing_address_line1 = my_sheet.cell(row=1, column=billing_address_line1_CNo)
    billing_address_line1.value = str("billing_address_line1")

if billing_address_line2_CNo != "NA":
    billing_address_line2 = my_sheet.cell(row=1, column=billing_address_line2_CNo)
    billing_address_line2.value = str("billing_address_line2")

if billing_address_line3_CNo != "NA":
    billing_address_line3 = my_sheet.cell(row=1, column=billing_address_line3_CNo)
    billing_address_line3.value = str("billing_address_line3")

if billing_address_city_CNo != "NA":
    billing_address_city = my_sheet.cell(row=1, column=billing_address_city_CNo)
    billing_address_city.value = str("billing_address_city")

if billing_address_state_code_CNo != "NA":
    billing_address_state_code = my_sheet.cell(row=1, column=billing_address_state_code_CNo)
    billing_address_state_code.value = str("billing_address_state_code")

if billing_address_state_CNo != "NA":
    billing_address_state = my_sheet.cell(row=1, column=billing_address_state_CNo)
    billing_address_state.value = str("billing_address_state")

if billing_address_zip_CNo != "NA":
    billing_address_zip = my_sheet.cell(row=1, column=billing_address_zip_CNo)
    billing_address_zip.value = str("billing_address_zip")

if billing_address_country_CNo != "NA":
    billing_address_country = my_sheet.cell(row=1, column=billing_address_country_CNo)
    billing_address_country.value = str("billing_address_country")

if billing_address_validation_status_CNo != "NA":
    billing_address_validation_status = my_sheet.cell(row=1, column=billing_address_validation_status_CNo)
    billing_address_validation_status.value = str("billing_address_validation_status")

if cf_account_number_CNo != "NA":
    cf_account_number = my_sheet.cell(row=1, column=cf_account_number_CNo)
    cf_account_number.value = str("cf_account_number")

if cf_xero_id_CNo != "NA":
    cf_xero_id = my_sheet.cell(row=1, column=cf_xero_id_CNo)
    cf_xero_id.value = str("cf_xero_id")

if cf_xero_name_CNo != "NA":
    cf_xero_name = my_sheet.cell(row=1, column=cf_xero_name_CNo)
    cf_xero_name.value = str("cf_xero_name")

if created_at_CNo != "NA":
    created_at = my_sheet.cell(row=1, column=created_at_CNo)
    created_at.value = str("created_at")

# parse each element/field from response and return

def get_id(i):
    try:
        id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.id")
    except:
        print("in status exception")
    return id


def get_first_name(i):
    try:
        first_name = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.first_name")
    except:
        print("in status exception")
    return first_name


def get_last_name(i):
    try:
        last_name = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.last_name")
    except:
        print("in status exception")
    return last_name


def get_email(i):
    try:
        email = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.email")
    except:
        print("in status exception")
    return email


def get_phone(i):
    try:
        phone = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.phone")
    except:
        print("in status exception")
    return phone


def get_company(i):
    try:
        company = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.company")
    except:
        print("in status exception")
    return company


def get_vat_number(i):
    try:
        vat_number = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.vat_number")
    except:
        print("in status exception")
    return vat_number


def get_auto_collection(i):
    try:
        auto_collection = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.auto_collection")
    except:
        print("in status exception")
    return auto_collection


def get_net_term_days(i):
    try:
        net_term_days = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.net_term_days")
    except:
        print("in status exception")
    return net_term_days


def get_vat_number_status(i):
    try:
        vat_number_status = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.vat_number_status")
    except:
        print("in status exception")
    return vat_number_status


def get_allow_direct_debit(i):
    try:
        allow_direct_debit = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.allow_direct_debit")
    except:
        print("in status exception")
    return allow_direct_debit


def get_created_at(i):
    try:
        created_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.created_at")
    except:
        print("in status exception")
    return created_at


def get_created_from_ip(i):
    try:
        created_from_ip = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.created_from_ip")
    except:
        print("in status exception")
    return created_from_ip


def get_taxability(i):
    try:
        taxability = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.taxability")
    except:
        print("in status exception")
    return taxability


def get_updated_at(i):
    try:
        updated_at = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.updated_at")
    except:
        print("in status exception")
    return updated_at


def get_locale(i):
    try:
        locale = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.locale")
    except:
        print("in status exception")
    return locale


def get_pii_cleared(i):
    try:
        pii_cleared = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.pii_cleared")
    except:
        print("in status exception")
    return pii_cleared


def get_resource_version(i):
    try:
        resource_version = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.resource_version")
    except:
        print("in status exception")
    return resource_version


def get_deleted(i):
    try:
        deleted = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.deleted")
    except:
        print("in status exception")
    return deleted


def get_object(i):
    try:
        object = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.object")
    except:
        print("in status exception")
    return object


def get_billing_address_first_name(i):
    try:
        billing_address_first_name = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].customer.billing_address.first_name")
    except:
        print("in status exception")
    return billing_address_first_name


def get_billing_address_last_name(i):
    try:
        billing_address_last_name = jsonpath.jsonpath(jsonpathres,
                                                      "list[" + str(i) + "].customer.billing_address.last_name")
    except:
        print("in status exception")
    return billing_address_last_name


def get_billing_address_email(i):
    try:
        billing_address_email = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.email")
    except:
        print("in status exception")
    return billing_address_email


def get_billing_address_company(i):
    try:
        billing_address_company = jsonpath.jsonpath(jsonpathres,
                                                    "list[" + str(i) + "].customer.billing_address.company")
    except:
        print("in status exception")
    return billing_address_company


def get_billing_address_phone(i):
    try:
        billing_address_phone = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.phone")
    except:
        print("in status exception")
    return billing_address_phone


def get_billing_address_line1(i):
    try:
        billing_address_line1 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.line1")
    except:
        print("in status exception")
    return billing_address_line1


def get_billing_address_line2(i):
    try:
        billing_address_line2 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.line2")
    except:
        print("in status exception")
    return billing_address_line2


def get_billing_address_line3(i):
    try:
        billing_address_line3 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.line3")
    except:
        print("in status exception")
    return billing_address_line3


def get_billing_address_city(i):
    try:
        billing_address_city = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.city")
    except:
        print("in status exception")
    return billing_address_city


def get_billing_address_country(i):
    try:
        billing_address_country = jsonpath.jsonpath(jsonpathres,
                                                    "list[" + str(i) + "].customer.billing_address.country")
    except:
        print("in status exception")
    return billing_address_country


def get_billing_address_zip(i):
    try:
        billing_address_zip = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.zip")
    except:
        print("in status exception")
    return billing_address_zip


def get_billing_address_validation_status(i):
    try:
        billing_address_validation_status = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].customer.billing_address.validation_status")
    except:
        print("in status exception")
    return billing_address_validation_status


def get_billing_address_object(i):
    try:
        billing_address_object = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.billing_address.object")
    except:
        print("in status exception")
    return billing_address_object


def get_card_status(i):
    try:
        card_status = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.card_status")
    except:
        print("in status exception")
    return card_status


def get_promotional_credits(i):
    try:
        promotional_credits = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.promotional_credits")
    except:
        print("in status exception")
    return promotional_credits


def get_refundable_credits(i):
    try:
        refundable_credits = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.refundable_credits")
    except:
        print("in status exception")
    return refundable_credits


def get_excess_payments(i):
    try:
        excess_payments = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.excess_payments")
    except:
        print("in status exception")
    return excess_payments


def get_unbilled_charges(i):
    try:
        unbilled_charges = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.unbilled_charges")
    except:
        print("in status exception")
    return unbilled_charges


def get_preferred_currency_code(i):
    try:
        preferred_currency_code = jsonpath.jsonpath(jsonpathres,
                                                    "list[" + str(i) + "].customer.preferred_currency_code")
    except:
        print("in status exception")
    return preferred_currency_code


def get_business_customer_without_vat_number(i):
    try:
        business_customer_without_vat_number = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].customer.business_customer_without_vat_number")
    except:
        print("in status exception")
    return business_customer_without_vat_number


def get_vat_number_validated_time(i):
    try:
        vat_number_validated_time = jsonpath.jsonpath(jsonpathres,
                                                      "list[" + str(i) + "].customer.vat_number_validated_time")
    except:
        print("in status exception")
    return vat_number_validated_time


def get_cf_clientid(i):
    try:
        cf_clientid = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.cf_clientid")
    except:
        print("in status exception")
    return cf_clientid


def get_cf_account_number(i):
    try:
        cf_account_number = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.cf_account_number")
    except:
        print("in status exception")
    return cf_account_number


def get_cf_xero_id(i):
    try:
        cf_xero_id = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.cf_xero_id")
    except:
        print("in status exception")
    return cf_xero_id


def get_cf_xero_name(i):
    try:
        cf_xero_name = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.cf_xero_name")
    except:
        print("in status exception")
    return cf_xero_name


def get_registered_for_gst(i):
    try:
        registered_for_gst = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.registered_for_gst")
    except:
        print("in status exception")
    return registered_for_gst


def get_cf_clientid(i):
    try:
        cf_clientid = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.cf_clientid")
    except:
        print("in status exception")
    return cf_clientid


def get_entity_code(i):
    try:
        entity_code = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.entity_code")
    except:
        print("in status exception")
    return entity_code


def get_exempt_number(i):
    try:
        exempt_number = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.exempt_number")
    except:
        print("in status exception")
    return exempt_number


def get_meta_data(i):
    try:
        meta_data = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.meta_data")
    except:
        print("in status exception")
    return meta_data


def get_consolidated_invoicing(i):
    try:
        consolidated_invoicing = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.consolidated_invoicing")
    except:
        print("in status exception")
    return consolidated_invoicing


def get_invoice_notes(i):
    try:
        invoice_notes = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.invoice_notes")
    except:
        print("in status exception")
    return invoice_notes


def get_payment_method_type(i):
    try:
        payment_method_type = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].customer.payment_method.type")
    except:
        print("in status exception")
    return payment_method_type


def get_card_last4(i):
    try:
        card_last4 = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].card.last4")
    except:
        print("in status exception")
    return card_last4


def get_card_brand(i):
    try:
        card_brand = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].card.card_type")
    except:
        print("in status exception")
    return card_brand


def get_card_funding(i):
    try:
        card_funding = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].card.funding_type")
    except:
        print("in status exception")
    return card_funding


def get_card_exp_month(i):
    try:
        card_exp_month = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].card.expiry_month")
    except:
        print("in status exception")
    return card_exp_month


def get_card_exp_year(i):
    try:
        card_exp_year = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].card.expiry_year")
    except:
        print("in status exception")
    return card_exp_year


def get_card_name(i):
    final_card_name = "BLANK"
    try:
        card_first_name = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].card.first_name")
        card_last_name = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "].card.last_name")
        if card_first_name[0] != '' and card_last_name[0] != '':
            final_card_name = card_first_name[0] + " " + card_last_name[0]
    except:
        print("in status exception card name")
    return final_card_name


def get_payment_method_gateway_account_id(i):
    try:
        payment_method_gateway_account_id = jsonpath.jsonpath(jsonpathres, "list[" + str(
            i) + "].customer.payment_method_gateway_account_id")
    except:
        print("in status exception")
    return payment_method_gateway_account_id


def get_payment_method_reference_id(i):
    try:
        payment_method_reference_id = jsonpath.jsonpath(jsonpathres,
                                                        "list[" + str(i) + "].customer.payment_method_reference_id")
    except:
        print("in status exception")
    return payment_method_reference_id


def get_payment_method_tmp_token(i):
    try:
        payment_method_tmp_token = jsonpath.jsonpath(jsonpathres,
                                                     "list[" + str(i) + "].customer.payment_method_tmp_token")
    except:
        print("in status exception")
    return payment_method_tmp_token


def get_billing_address_state_code(i):
    try:
        billing_address_state_code = jsonpath.jsonpath(jsonpathres,
                                                       "list[" + str(i) + "].customer.billing_address.state_code")
    except:
        print("in status exception")
    return billing_address_state_code


def get_billing_address_state(i):
    try:
        billing_address_state = jsonpath.jsonpath(jsonpathres,
                                                  "list[" + str(i) + "].customer.billing_address.state")
    except:
        print("in status exception")
    return billing_address_state


def get_created_at(i):
    try:
        created_at = jsonpath.jsonpath(Customerdictionary, "list[" + str(i) + "].customer.created_at")
    except:
        print("in status exception")
    return created_at



# ---------------------------loop through all records  and write to excel---------------------------#
for i in range(0, totalRecordCountInResp):
    try:

        if id_CNo != "NA":
            id = get_id(i)
            id_cell = my_sheet.cell(row=i + 2, column=id_CNo)
            if id == False:
                id_cell.value = id
            else:
                id_cell.value = str(id[0])

        if first_name_CNo != "NA":
            first_name = get_first_name(i)
            first_name_cell = my_sheet.cell(row=i + 2, column=first_name_CNo)
            if first_name == False:
                first_name_cell.value = first_name
            else:
                first_name_cell.value = str(first_name[0])

        if last_name_CNo != "NA":
            last_name = get_last_name(i)
            last_name_cell = my_sheet.cell(row=i + 2, column=last_name_CNo)
            if last_name == False:
                last_name_cell.value = last_name
            else:
                last_name_cell.value = str(last_name[0])

        if email_CNo != "NA":
            email = get_email(i)
            email_cell = my_sheet.cell(row=i + 2, column=email_CNo)
            if email == False:
                email_cell.value = email
            else:
                emailvalue = email[0]
                emailvalue = emailvalue.replace("_AT_", "@")
                emailvalue = emailvalue.replace("@example.com", "")
                email_cell.value = str(emailvalue)

        if preferred_currency_code_CNo != "NA":
            preferred_currency_code = get_preferred_currency_code(i)
            preferred_currency_code_cell = my_sheet.cell(row=i + 2, column=preferred_currency_code_CNo)
            if preferred_currency_code == False:
                preferred_currency_code_cell.value = preferred_currency_code
            else:
                preferred_currency_code_cell.value = str(preferred_currency_code[0])

        if phone_CNo != "NA":
            phone = get_phone(i)
            phone_cell = my_sheet.cell(row=i + 2, column=phone_CNo)
            if phone == False:
                phone_cell.value = phone
            else:
                # phone_cell.value = str(phone[0])
                temp = str(phone[0])
                temp = temp.replace("(", "")
                temp = temp.replace("-", "")
                temp = temp.replace(")", "")
                temp = temp.replace(" ", "")
                phone_cell.value = temp
                # temp = temp.replace("(","")
        #                  for ch in ["(",")","-"]:
        #                      if ch in temp:
        #                          temp = temp.replace(ch,"")
        #                          phone_cell.value = temp
        #                      else:
        #                          phone_cell.value = str(phone[0])

        if company_CNo != "NA":
            company = get_company(i)
            company_cell = my_sheet.cell(row=i + 2, column=company_CNo)
            if company == False:
                company_cell.value = company
            else:
                company_cell.value = str(company[0])

        if auto_collection_CNo != "NA":
            auto_collection = get_auto_collection(i)
            auto_collection_cell = my_sheet.cell(row=i + 2, column=auto_collection_CNo)
            if auto_collection == False:
                auto_collection_cell.value = auto_collection
            else:
                auto_collection_cell.value = str(auto_collection[0])

        if net_term_days_CNo != "NA":
            net_term_days = get_net_term_days(i)
            net_term_days_cell = my_sheet.cell(row=i + 2, column=net_term_days_CNo)
            if net_term_days == False:
                net_term_days_cell.value = net_term_days
            else:
                net_term_days_cell.value = str(net_term_days[0])

        if allow_direct_debit_CNo != "NA":
            allow_direct_debit = get_allow_direct_debit(i)
            allow_direct_debit_cell = my_sheet.cell(row=i + 2, column=allow_direct_debit_CNo)
            if allow_direct_debit == False:
                allow_direct_debit_cell.value = allow_direct_debit
            else:
                allow_direct_debit_cell.value = str(allow_direct_debit[0])

        if vat_number_CNo != "NA":
            vat_number = get_vat_number(i)
            vat_number_cell = my_sheet.cell(row=i + 2, column=vat_number_CNo)
            if vat_number == False:
                vat_number_cell.value = vat_number
            else:
                vat_number_cell.value = str(vat_number[0])

        if registered_for_gst_CNo != "NA":
            registered_for_gst = get_registered_for_gst(i)
            registered_for_gst_cell = my_sheet.cell(row=i + 2, column=registered_for_gst_CNo)
            if registered_for_gst == False:
                registered_for_gst_cell.value = registered_for_gst
            else:
                registered_for_gst_cell.value = str(registered_for_gst[0])

        if taxability_CNo != "NA":
            taxability = get_taxability(i)
            taxability_cell = my_sheet.cell(row=i + 2, column=taxability_CNo)
            if taxability == False:
                taxability_cell.value = taxability
            else:
                taxability_cell.value = str(taxability[0])

        if locale_CNo != "NA":
            locale = get_locale(i)
            locale_cell = my_sheet.cell(row=i + 2, column=locale_CNo)
            if locale == False:
                locale_cell.value = locale
            else:
                locale_cell.value = str(locale[0])

        if entity_code_CNo != "NA":
            entity_code = get_entity_code(i)
            entity_code_cell = my_sheet.cell(row=i + 2, column=entity_code_CNo)
            if entity_code == False:
                entity_code_cell.value = entity_code
            else:
                entity_code_cell.value = str(entity_code[0])

        if exempt_number_CNo != "NA":
            exempt_number = get_exempt_number(i)
            exempt_number_cell = my_sheet.cell(row=i + 2, column=exempt_number_CNo)
            if exempt_number == False:
                exempt_number_cell.value = exempt_number
            else:
                exempt_number_cell.value = str(exempt_number[0])

        if meta_data_CNo != "NA":
            meta_data = get_meta_data(i)
            meta_data_cell = my_sheet.cell(row=i + 2, column=meta_data_CNo)
            if meta_data == False:
                meta_data_cell.value = meta_data
            else:
                meta_data_cell.value = str(meta_data[0])

        if consolidated_invoicing_CNo != "NA":
            consolidated_invoicing = get_consolidated_invoicing(i)
            consolidated_invoicing_cell = my_sheet.cell(row=i + 2, column=consolidated_invoicing_CNo)
            if consolidated_invoicing == False:
                consolidated_invoicing_cell.value = consolidated_invoicing
            else:
                consolidated_invoicing_cell.value = str(consolidated_invoicing[0])

        if invoice_notes_CNo != "NA":
            invoice_notes = get_invoice_notes(i)
            invoice_notes_cell = my_sheet.cell(row=i + 2, column=invoice_notes_CNo)
            if invoice_notes == False:
                invoice_notes_cell.value = invoice_notes
            else:
                invoice_notes_cell.value = str(invoice_notes[0])

        if payment_method_type_CNo != "NA":
            payment_method_type = get_payment_method_type(i)
            payment_method_type_cell = my_sheet.cell(row=i + 2, column=payment_method_type_CNo)
            if payment_method_type == False:
                payment_method_type_cell.value = "BLANK"
            else:
                payment_method_type_cell.value = str(payment_method_type[0])

        if card_last4_CNo != "NA":
            card_last4 = get_card_last4(i)
            card_last4_cell = my_sheet.cell(row=i + 2, column=card_last4_CNo)
            if card_last4 == False:
                card_last4_cell.value = card_last4
            else:
                card_last4_cell.value = str(card_last4[0])

        if card_brand_CNo != "NA":
            card_brand = get_card_brand(i)
            card_brand_cell = my_sheet.cell(row=i + 2, column=card_brand_CNo)
            if card_brand == False:
                card_brand_cell.value = card_brand
            else:
                temp = str(card_brand[0])
                card_brand_cell.value = temp.replace("_", " ")

        if card_funding_CNo != "NA":
            card_funding = get_card_funding(i)
            card_funding_cell = my_sheet.cell(row=i + 2, column=card_funding_CNo)
            if card_funding == False:
                card_funding_cell.value = card_funding
            else:
                card_funding_cell.value = str(card_funding[0])

        if card_exp_month_CNo != "NA":
            card_exp_month = get_card_exp_month(i)
            card_exp_month_cell = my_sheet.cell(row=i + 2, column=card_exp_month_CNo)
            if card_exp_month == False:
                card_exp_month_cell.value = card_exp_month
            else:
                card_exp_month_cell.value = str(card_exp_month[0])

        if card_exp_year_CNo != "NA":
            card_exp_year = get_card_exp_year(i)
            card_exp_year_cell = my_sheet.cell(row=i + 2, column=card_exp_year_CNo)
            if card_exp_year == False:
                card_exp_year_cell.value = card_exp_year
            else:
                card_exp_year_cell.value = str(card_exp_year[0])

        if card_name_CNo != "NA":
            card_name = get_card_name(i)
            card_name_cell = my_sheet.cell(row=i + 2, column=card_name_CNo)
            if card_name == False:
                card_name_cell.value = card_name
            else:
                card_name_cell.value = str(card_name)

        if payment_method_gateway_account_id_CNo != "NA":
            payment_method_gateway_account_id = get_payment_method_gateway_account_id(i)
            payment_method_gateway_account_id_cell = my_sheet.cell(row=i + 2,
                                                                   column=payment_method_gateway_account_id_CNo)
            if payment_method_gateway_account_id == False:
                payment_method_gateway_account_id_cell.value = payment_method_gateway_account_id
            else:
                payment_method_gateway_account_id_cell.value = str(payment_method_gateway_account_id[0])

        if payment_method_reference_id_CNo != "NA":
            payment_method_reference_id = get_payment_method_reference_id(i)
            payment_method_reference_id_cell = my_sheet.cell(row=i + 2, column=payment_method_reference_id_CNo)
            if payment_method_reference_id == False:
                payment_method_reference_id_cell.value = payment_method_reference_id
            else:
                payment_method_reference_id_cell.value = str(payment_method_reference_id[0])

        if payment_method_tmp_token_CNo != "NA":
            payment_method_tmp_token = get_payment_method_tmp_token(i)
            payment_method_tmp_token_cell = my_sheet.cell(row=i + 2, column=payment_method_tmp_token_CNo)
            if payment_method_tmp_token == False:
                payment_method_tmp_token_cell.value = payment_method_tmp_token
            else:
                payment_method_tmp_token_cell.value = str(payment_method_tmp_token[0])

        if billing_address_first_name_CNo != "NA":
            billing_address_first_name = get_billing_address_first_name(i)
            billing_address_first_name_cell = my_sheet.cell(row=i + 2, column=billing_address_first_name_CNo)
            if billing_address_first_name == False:
                billing_address_first_name_cell.value = billing_address_first_name
            else:
                billing_address_first_name_cell.value = str(billing_address_first_name[0])

        if billing_address_last_name_CNo != "NA":
            billing_address_last_name = get_billing_address_last_name(i)
            billing_address_last_name_cell = my_sheet.cell(row=i + 2, column=billing_address_last_name_CNo)
            if billing_address_last_name == False:
                billing_address_last_name_cell.value = billing_address_last_name
            else:
                billing_address_last_name_cell.value = str(billing_address_last_name[0])

        if billing_address_email_CNo != "NA":
            billing_address_email = get_billing_address_email(i)
            billing_address_email_cell = my_sheet.cell(row=i + 2, column=billing_address_email_CNo)
            if billing_address_email == False:
                billing_address_email_cell.value = billing_address_email
            else:
                billing_address_emailvalue = billing_address_email[0]
                billing_address_emailvalue = billing_address_emailvalue.replace("_AT_", "@")
                billing_address_emailvalue = billing_address_emailvalue.replace("@example.com", "")
                billing_address_email_cell.value = str(billing_address_emailvalue)

        if billing_address_company_CNo != "NA":
            billing_address_company = get_billing_address_company(i)
            billing_address_company_cell = my_sheet.cell(row=i + 2, column=billing_address_company_CNo)
            if billing_address_company == False:
                billing_address_company_cell.value = billing_address_company
            else:
                billing_address_company_cell.value = str(billing_address_company[0])

        if billing_address_phone_CNo != "NA":
            billing_address_phone = get_billing_address_phone(i)
            billing_address_phone_cell = my_sheet.cell(row=i + 2, column=billing_address_phone_CNo)
            if billing_address_phone == False:
                billing_address_phone_cell.value = billing_address_phone
            else:
                billing_address_phone_cell.value = str(billing_address_phone[0])

        if billing_address_line1_CNo != "NA":
            billing_address_line1 = get_billing_address_line1(i)
            billing_address_line1_cell = my_sheet.cell(row=i + 2, column=billing_address_line1_CNo)
            if billing_address_line1 == False:
                billing_address_line1_cell.value = billing_address_line1
            else:
                billing_address_line1_cell.value = str(billing_address_line1[0])

        if billing_address_line2_CNo != "NA":
            billing_address_line2 = get_billing_address_line2(i)
            billing_address_line2_cell = my_sheet.cell(row=i + 2, column=billing_address_line2_CNo)
            if billing_address_line2 == False:
                billing_address_line2_cell.value = billing_address_line2
            else:
                billing_address_line2_cell.value = str(billing_address_line2[0])

        if billing_address_line3_CNo != "NA":
            billing_address_line3 = get_billing_address_line3(i)
            billing_address_line3_cell = my_sheet.cell(row=i + 2, column=billing_address_line3_CNo)
            if billing_address_line3 == False:
                billing_address_line3_cell.value = billing_address_line3
            else:
                billing_address_line3_cell.value = str(billing_address_line3[0])

        if billing_address_city_CNo != "NA":
            billing_address_city = get_billing_address_city(i)
            billing_address_city_cell = my_sheet.cell(row=i + 2, column=billing_address_city_CNo)
            if billing_address_city == False:
                billing_address_city_cell.value = billing_address_city
            else:
                billing_address_city_cell.value = str(billing_address_city[0])

        if billing_address_state_code_CNo != "NA":
            billing_address_state_code = get_billing_address_state_code(i)
            billing_address_state_code_cell = my_sheet.cell(row=i + 2, column=billing_address_state_code_CNo)
            if billing_address_state_code == False:
                billing_address_state_code_cell.value = billing_address_state_code
            else:
                billing_address_state_code_cell.value = str(billing_address_state_code[0])

        if billing_address_state_CNo != "NA":
            billing_address_state = get_billing_address_state(i)
            billing_address_state_cell = my_sheet.cell(row=i + 2, column=billing_address_state_CNo)
            if billing_address_state == False:
                billing_address_state_cell.value = billing_address_state
            else:
                billing_address_state_cell.value = str(billing_address_state[0])

        if billing_address_zip_CNo != "NA":
            billing_address_zip = get_billing_address_zip(i)
            billing_address_zip_cell = my_sheet.cell(row=i + 2, column=billing_address_zip_CNo)
            if billing_address_zip == False:
                billing_address_zip_cell.value = billing_address_zip
            else:
                billing_address_zip_cell.value = str(billing_address_zip[0])

        if billing_address_country_CNo != "NA":
            billing_address_country = get_billing_address_country(i)
            billing_address_country_cell = my_sheet.cell(row=i + 2, column=billing_address_country_CNo)
            if billing_address_country == False:
                billing_address_country_cell.value = billing_address_country
            else:
                billing_address_country_cell.value = str(billing_address_country[0])

        if billing_address_validation_status_CNo != "NA":
            billing_address_validation_status = get_billing_address_validation_status(i)
            billing_address_validation_status_cell = my_sheet.cell(row=i + 2,
                                                                   column=billing_address_validation_status_CNo)
            if billing_address_validation_status == False:
                billing_address_validation_status_cell.value = billing_address_validation_status
            else:
                billing_address_validation_status_cell.value = str(billing_address_validation_status[0])

        if cf_account_number_CNo != "NA":
            cf_account_number = get_cf_account_number(i)
            cf_account_number_cell = my_sheet.cell(row=i + 2, column=cf_account_number_CNo)
            if cf_account_number == False:
                cf_account_number_cell.value = cf_account_number
            else:
                cf_account_number_cell.value = str(cf_account_number[0])

        if cf_xero_id_CNo != "NA":
            cf_xero_id = get_cf_xero_id(i)
            cf_xero_id_cell = my_sheet.cell(row=i + 2, column=cf_xero_id_CNo)
            if cf_xero_id == False:
                cf_xero_id_cell.value = cf_xero_id
            else:
                cf_xero_id_cell.value = str(cf_xero_id[0])

        if cf_xero_name_CNo != "NA":
            cf_xero_name = get_cf_xero_name(i)
            cf_xero_name_cell = my_sheet.cell(row=i + 2, column=cf_xero_name_CNo)
            if cf_xero_name == False:
                cf_xero_name_cell.value = cf_xero_name
            else:
                cf_xero_name_cell.value = str(cf_xero_name[0])

        if created_at_CNo != "NA":
            created_at = get_created_at(i)
            created_at_cell = my_sheet.cell(row=i + 2, column=created_at_CNo)
            if created_at == False:
                created_at_cell.value = created_at
            else:
                modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(created_at[0], clienttimezone)
                created_at_cell.value = str(modifiedtimestamp)

    except Exception as e:
        print("in exception", e)

# finally save excel
my_wb.save(output_file_name)
print("Execution completed")
end = timer()
print('Total time taken: ', end - start, ' seconds')
