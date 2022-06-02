import json
import os
from timeit import default_timer as timer

import jsonpath
import openpyxl
from jproperties import Properties

from utils.Ephoc2DateTime import Ephoc2DateTime as tzconverter

configs = Properties()
ROOT_DIR1 = os.path.abspath(os.curdir)
ROOT_DIR = ROOT_DIR1.replace('writeExcel', 'configuration.properties')
with open(ROOT_DIR, 'rb') as config_file:
    configs.load(config_file)
Json_DIR = ROOT_DIR1.replace('writeExcel', 'readAPI/jsonfiles/')
start = timer()
# input_jsonfile = "/Users/cb-muneendra/cb_data_validation/utils/AAAABBBPixellu_AllInvoices.json"
input_jsonfile = "/Users/cb-muneendra/cb_data_validation/readAPI/jsonfiles/pixellu_live_AllInvoices.json"
# input_jsonfile = Json_DIR + configs.get("clientName").data + "_AllInvoices.json"
output_file_name = ROOT_DIR1 + '/outputExcelFiles/' + configs.get("clientName").data + "_Invoice_Actual3.xlsx"
clienttimezone = configs.get("clienttimezone").data

jsondata = open(input_jsonfile)
Invoicesdictionary = json.load(jsondata)
print("Final Json File", Invoicesdictionary)
apitojsontime = timer()

jsonpathres = Invoicesdictionary
l = jsonpath.jsonpath(jsonpathres, "list")
totalRecordCountInResp = len(l[0])
# create work book and sheet object
my_wb = openpyxl.Workbook()
my_sheet = my_wb.active
my_sheet.title = "Invoices"

datefields = ["invoice[date]", "invoice[due_date]", "line_items[date_from][0]", "line_items[date_to][0]",
              "payments[date][0]", "line_items[date_from][1]", "line_items[date_to][1]", "line_items[date_from][2]",
              "line_items[date_to][2]", "line_items[date_from][3]", "line_items[date_to][3]",  "line_items[date_from][4]", "line_items[date_to][4]",  "line_items[date_from][5]", "line_items[date_to][5]"]
amountfields = ["invoice[total]", "line_items[unit_amount][0]", "line_items[amount][0]",
                "line_items[item_level_discount1_amount][0]", "line_items[tax1_amount][0]", "discounts[amount][0]",
                "payments[amount][0]", "line_items[unit_amount][1]", "line_items[amount][1]",
                "line_items[unit_amount][2]", "line_items[amount][2]", "line_items[unit_amount][3]", "line_items[amount][3]", "line_items[unit_amount][4]", "line_items[amount][4]", "line_items[unit_amount][5]", "line_items[amount][5]"]
columns = ["invoice[id]", "invoice[currency_code]", "invoice[customer_id]", "invoice[subscription_id]",
           "invoice[status]", "invoice[date]", "invoice[po_number]", "invoice[price_type]", "tax_override_reason",
           "invoice[vat_number]", "round_off", "invoice[total]", "invoice[due_date]", "invoice[net_term_days]",
           "use_for_proration", "line_items[id][0]", "line_items[entity_type][0]", "line_items[entity_id][0]",
           "line_items[description][0]", "line_items[date_from][0]", "line_items[date_to][0]",
           "line_items[quantity][0]", "line_items[unit_amount][0]", "line_items[amount][0]",
           "line_items[item_level_discount1_entity_id][0]", "line_items[item_level_discount1_amount][0]",
           "line_items[tax1_name][0]", "line_items[tax1_amount][0]", "discounts[entity_type][0]",
           "discounts[entity_id][0]", "discounts[description][0]", "discounts[amount][0]", "payments[amount][0]",
           "payments[date][0]", "line_items[entity_type][1]", "line_items[entity_id][1]", "line_items[description][1]",
           "line_items[date_from][1]", "line_items[date_to][1]", "line_items[quantity][1]",
           "line_items[unit_amount][1]", "line_items[amount][1]", "line_items[entity_type][2]",
           "line_items[entity_id][2]", "line_items[description][2]", "line_items[date_from][2]",
           "line_items[date_to][2]", "line_items[quantity][2]", "line_items[unit_amount][2]", "line_items[amount][2]","line_items[entity_type][3]",
           "line_items[entity_id][3]", "line_items[description][3]", "line_items[date_from][3]",
           "line_items[date_to][3]", "line_items[quantity][3]", "line_items[unit_amount][3]", "line_items[amount][3]","line_items[entity_type][4]",
           "line_items[entity_id][4]", "line_items[description][4]", "line_items[date_from][4]",
           "line_items[date_to][4]", "line_items[quantity][4]", "line_items[unit_amount][4]", "line_items[amount][4]","line_items[entity_type][5]",
           "line_items[entity_id][5]", "line_items[description][5]", "line_items[date_from][5]",
           "line_items[date_to][5]", "line_items[quantity][5]", "line_items[unit_amount][5]", "line_items[amount][5]",
           "billing_address[first_name]", "billing_address[last_name]", "billing_address[email]",
           "billing_address[company]", "billing_address[phone]", "billing_address[line1]", "billing_address[line2]",
           "billing_address[line3]", "billing_address[state_code]", "billing_address[city]", "billing_address[state]",
           "billing_address[zip]", "billing_address[country]", "shipping_address[first_name]",
           "shipping_address[last_name]", "shipping_address[email]", "shipping_address[company]",
           "shipping_address[phone]", "shipping_address[line1]", "shipping_address[line2]", "shipping_address[line3]",
           "shipping_address[city]", "shipping_address[state_code]", "shipping_address[state]", "shipping_address[zip]",
           "shipping_address[country]"]

for x in range(0, len(columns) - 1):
    colhead = my_sheet.cell(row=1, column=x + 1)
    colhead.value = str(columns[x])

jsoncall = [".invoice.id", ".invoice.currency_code", ".invoice.customer_id", ".invoice.subscription_id",
            ".invoice.status", ".invoice.date", ".invoice.po_number", ".invoice.price_type",
            ".invoice.tax_override_reason", ".invoice.vat_number", ".invoice.round_off_amount", ".invoice.total",
            ".invoice.due_date", ".invoice.net_term_days", ".invoice.use_for_proration", ".invoice.line_items[0].id",
            ".invoice.line_items[0].entity_type", ".invoice.line_items[0].entity_id",
            ".invoice.line_items[0].description", ".invoice.line_items[0].date_from", ".invoice.line_items[0].date_to",
            ".invoice.line_items[0].quantity", ".invoice.line_items[0].unit_amount", ".invoice.line_items[0].amount",
            ".invoice.line_items[0].item_level_discount_entity_id", ".invoice.line_items[0].item_level_discount_amount",
            ".invoice.line_items[0].tax_name", ".invoice.line_items[0].tax_amount", ".invoice.discounts[0].entity_type",
            ".invoice.discounts[0].entity_id", ".invoice.discounts[0].description", ".invoice.discounts[0].amount",
            ".invoice.linked_payments[0].txn_amount", ".invoice.linked_payments[0].txn_date",
            ".invoice.line_items[1].entity_type", ".invoice.line_items[1].entity_id",
            ".invoice.line_items[1].description", ".invoice.line_items[1].date_from", ".invoice.line_items[1].date_to",
            ".invoice.line_items[1].quantity", ".invoice.line_items[1].unit_amount", ".invoice.line_items[1].amount",
            ".invoice.line_items[2].entity_type", ".invoice.line_items[2].entity_id",
            ".invoice.line_items[2].description", ".invoice.line_items[2].date_from", ".invoice.line_items[2].date_to",
            ".invoice.line_items[2].quantity", ".invoice.line_items[2].unit_amount", ".invoice.line_items[2].amount",".invoice.line_items[3].entity_type", ".invoice.line_items[3].entity_id",
            ".invoice.line_items[3].description", ".invoice.line_items[3].date_from", ".invoice.line_items[3].date_to",
            ".invoice.line_items[3].quantity", ".invoice.line_items[3].unit_amount", ".invoice.line_items[3].amount",".invoice.line_items[4].entity_type", ".invoice.line_items[4].entity_id",
            ".invoice.line_items[4].description", ".invoice.line_items[4].date_from", ".invoice.line_items[4].date_to",
            ".invoice.line_items[4].quantity", ".invoice.line_items[4].unit_amount", ".invoice.line_items[4].amount",".invoice.line_items[5].entity_type", ".invoice.line_items[5].entity_id",
            ".invoice.line_items[5].description", ".invoice.line_items[5].date_from", ".invoice.line_items[5].date_to",
            ".invoice.line_items[5].quantity", ".invoice.line_items[5].unit_amount", ".invoice.line_items[5].amount",
            ".invoice.billing_address.first_name", ".invoice.billing_address.last_name",
            ".invoice.billing_address.email", ".invoice.billing_address.company", ".invoice.billing_address.phone",
            ".invoice.billing_address.line1", ".invoice.billing_address.line2", ".invoice.billing_address.line3",
            ".invoice.billing_address.state_code", ".invoice.billing_address.city", ".invoice.billing_address.state",
            ".invoice.billing_address.zip", ".invoice.billing_address.country", ".invoice.shipping_address.first_name",
            ".invoice.shipping_address.last_name", ".invoice.shipping_address.email",
            ".invoice.shipping_address.company", ".invoice.shipping_address.phone", ".invoice.shipping_address.line1",
            ".invoice.shipping_address.line2", ".invoice.shipping_address.line3",
            ".invoice.shipping_address.state_code", ".invoice.shipping_address.city", ".invoice.shipping_address.state",
            ".invoice.shipping_address.zip", ".invoice.shipping_address.country"]


def getdata(i, col):
    try:
        data = jsonpath.jsonpath(jsonpathres, "list[" + str(i) + "]" + jsoncall[col])
    except:
        print("in status exception : ", jsoncall[col])
    return data


# ---------------------------loop through all records  and write to excel---------------------------#
for i in range(0, totalRecordCountInResp):
    try:
        for col in range(0, len(columns) - 1):
            coldata = getdata(i, col)
            coldata_cell = my_sheet.cell(row=i + 2, column=col + 1)
            if coldata == False:
                coldata_cell.value = coldata
            else:
                coldata_cell.value = str(coldata[0])
                if columns[col] in datefields:
                    modifiedtimestamp = tzconverter.epoch_To_Datetime_Convert(coldata[0], clienttimezone)
                    coldata_cell.value = str(modifiedtimestamp)
                if columns[col] in amountfields:
                    coldata_cell.value = str(coldata[0] / 100)
    except Exception as e:
        print("in exception", e)

# finally save excel
my_wb.save(output_file_name)
print("Execution completed")
end = timer()
print('Total time taken: ', end - start, ' seconds')
