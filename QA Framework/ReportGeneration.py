import datetime
import pandas as pd
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import load_workbook
from openpyxl.chart import (
    PieChart,
    Reference
)

from openpyxl.chart.series import DataPoint
import numpy as np


def report_generation(type_in, df1, site_name, client_name,module,sheet_name):
    # sheet_name = 'DS2VsDS3Subscription_diff'
    validation_type = (str(type_in)).split("_")
    now = datetime.datetime.now()
    file_name = "./TestResults/" + client_name + "_" + site_name + "_" + validation_type[0] + "_" + now.strftime("%Y%m%d%H%M%S") + "_Diff.xlsx"

    df1 = df1.replace(to_replace='nan --> False', value='')
    df1 = df1.replace(to_replace='nan --> 0', value='')
    df1 = df1.replace(to_replace='0 --> False', value='')
    df1 = df1.replace(to_replace='nan [-->]', value='BLANK -', regex=True)
    df1 = df1.replace(to_replace='False [-->]', value='BLANK -', regex=True)
    df1 = df1.replace(to_replace='false [-->]', value='BLANK -', regex=True)
    df1 = df1.replace(to_replace='[-->] nan', value='> BLANK', regex=True)
    df1 = df1.replace(to_replace='[-->] False', value='> BLANK', regex=True)
    df1 = df1.replace(to_replace='[-->] false', value='> BLANK', regex=True)
    df1 = df1.replace(to_replace='False', value='')
    df1 = df1.replace(to_replace='false', value='')
    df1 = df1.replace(to_replace='BLANK --> nan', value='')
    df1 = df1.replace(to_replace='BLANK --> BLANK', value='')
    df1 = df1.replace(to_replace='on --> off', value='on <ignore> off')
    df1 = df1.replace(to_replace=' --> False', value=' --> BLANK')
    df1 = df1.replace(to_replace='BLANK --> None', value='')
    df1 = df1.replace(to_replace='--> None', value='')
    df1 = df1.replace(to_replace='Paid --> PAID', value='PAID')
    df1 = df1.replace(to_replace='Payment Due --> PAYMENT_DUE', value='PAYMENT_DUE')
    df1 = df1.replace(to_replace='Active --> active', value='Active')
    df1 = df1.replace(to_replace='Cancelled --> cancelled', value='Cancelled')
    df1 = df1.replace(to_replace='0 --> BLANK', value='')
    df1 = df1.replace(to_replace=' --> BLANK', value='')
    df1 = df1.replace(to_replace=' --> false', value='')
    df1 = df1.replace(to_replace='false --> ', value='Blank --> ')
    df1 = df1.replace(to_replace='PAYMENT_DUE --> payment_due', value='PAYMENT_DUE')
    df1 = df1.replace(to_replace='NaT --> False', value='')
    df1 = df1.replace(to_replace='eur --> EUR', value='EUR')
    df1 = df1.replace(to_replace='usd --> USD', value='USD')
    df1 = df1.replace(to_replace=' --> IT', value='USD')
    df1 = df1.replace(to_replace='charge --> charge_item_price', value='charge_item_price')
    df1 = df1.replace(to_replace='plan --> plan_item_price', value='plan_item_price')
    df1 = df1.replace(to_replace='directdebit --> direct_debit', value='direct_debit')
    df1 = df1.replace(to_replace='creditcard --> card', value='creditcard <ignore> card')
    df1 = df1.replace(to_replace='addon --> addon_item_price', value='addon_item_price')
    df1 = df1.replace(to_replace='None --> ', value='')
    df1 = df1.replace(to_replace='1.0 --> 1', value='1')
    df1 = df1.replace(to_replace='2.0 --> 2', value='2')
    df1 = df1.replace(to_replace='3.0 --> 3', value='3')
    df1 = df1.replace(to_replace='4.0 --> 4', value='4')
    # df1 = df1.replace(to_replace='None', value='')
    df1 = df1.replace(to_replace='None --> BLANK', value='')
    # df1['invoice[stripe_id]'] = df1['invoice[stripe_id]'].replace(to_replace=' --> ', value='')


    # Insert an empty column to write the formulas
    df1.insert(len(df1.columns), 'Execution_Status', np.nan)

    writer = pd.ExcelWriter(file_name, engine='xlsxwriter')

    # Convert the dataframe to an XlsxWriter Excel object.
    df1.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

    # Get the xlsxwriter workbook and worksheet objects.
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes(1, 1)

    # Create a for loop to start writing the formulas to each row
    for row in range(2, df1.shape[0]+2):
        formula = f'=IF(COUNTIF(A{row}: {get_column_letter(len(df1.columns)-1)}{row}, "*-->*"), "FAIL", "PASS")'

        worksheet.write_formula(f"{get_column_letter(len(df1.columns))}{row}", formula)

    status_column_position = get_column_letter(len(df1.columns))

    # Add a format. Light red fill with dark red text.
    format1 = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
    format5 = workbook.add_format({'bg_color': '#FFDEAD', 'font_color': '#000000'})

    # Apply a conditional format to the cell range.
    worksheet.conditional_format(1, 0, len(df1), len(df1.columns),
                                 {'type':     'text',
                                  'criteria': 'containing',
                                  'value':    '-->',
                                  'format':   format1})

    worksheet.conditional_format(1, 0, len(df1), len(df1.columns),
                                 {'type': 'text',
                                  'criteria': 'containing',
                                  'value': '<ignore>',
                                  'format': format5})

    # Add a format. Blue fill with White text.
    format2 = workbook.add_format({'bg_color':   '#1589FF', 'font_color': '#FFFFFF'})

    worksheet.conditional_format(0, 0, 0, len(df1.columns)-1,
                                 {'type':     'text', 'criteria': 'not containing', 'value': '-->', 'format':   format2})

    # Add a format. Green fill with White text.
    format3 = workbook.add_format({'bg_color':   '#41A317', 'font_color': '#FFFFFF'})

    # Add a format. Red fill with White text.
    format4 = workbook.add_format({'bg_color':   '#E41B17', 'font_color': '#FFFFFF'})

    worksheet.conditional_format(1, len(df1.columns)-1, len(df1), len(df1.columns)-1,
                                 {'type':     'text',
                                  'criteria': 'containing',
                                  'value':    'PASS',
                                  'format':   format3})

    worksheet.conditional_format(1, len(df1.columns)-1, len(df1), len(df1.columns)-1,
                                 {'type':     'text',
                                  'criteria': 'containing',
                                  'value':    'FAIL',
                                  'format':   format4})

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()

    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    number_columns = sheet.max_column

    # set the font style to bold and column and row width
    for j in range(number_columns):
        sheet.cell(row=1, column=j+1).font = Font(size=14, bold=True)
        sheet.column_dimensions[get_column_letter(j+1)].width = 30

    sheet.row_dimensions[1].height = 20

    wb.save(file_name)
    wb.close()

    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book

    no_of_executions_formula = '=ROWS(' + sheet_name + '!A:A)-COUNTBLANK(' + sheet_name + '!A:A)-1'
    pass_formula = '=COUNTIF(' + sheet_name + '!' + status_column_position + ':' + status_column_position + ',"PASS")'
    fail_formula = '=COUNTIF(' + sheet_name + '!' + status_column_position + ':' + status_column_position + ',"FAIL")'
    other_formula = '=B4-(C4+D4)'
    total_deviations = '="Total no of mismatches  : "&COUNTIF(' + sheet_name + '!A:' + status_column_position + ',"* --> *")'

    farm_1 = {'Total No Of Records Validated': no_of_executions_formula, 'No Of Records Pass': pass_formula, 'No Of Records Fail': fail_formula,
              'Other': other_formula}

    df3 = pd.DataFrame([farm_1], index=['Farm 1'])

    df3.to_excel(writer, index=False, header=True, sheet_name='Execution_Report')

    writer.save()

    columns_data = list(df1.columns.values)
    del columns_data[-1]
    length_of_coulmns = len(columns_data)
    list_str_sno = list(range(1, length_of_coulmns+1))

    lst = [None] * length_of_coulmns

    no_of_mismatch_data = []
    qa_updates = []
    colwithf = []

    for col in range(len(columns_data)):
        c_letter = get_column_letter(col+1)
        link = '#'+sheet_name+'!' + c_letter + '1:' + c_letter+str(len(df1)+1)
        colwithf.append('=HYPERLINK("{}", "{}")'.format(link, columns_data[col]))

    for col in list_str_sno:
        c_letter = get_column_letter(col)
        no_of_mismatch_formula = '=COUNTIF('+sheet_name+'!' + c_letter + ':' + c_letter + ',"* --> *")'
        # NoOfMismathesformula = '=COUNTIF(DS2vsDS3_Subscription_diff!'+cLetter+':'+cLetter+',"* --> *")'
        no_of_mismatch_data.append(no_of_mismatch_formula)
        qaupdateformula = '=IF(D'+str(col + 1)+'>0,"Fail","Pass")'
        qa_updates.append(qaupdateformula)

    data = {'SNo': list_str_sno,
            # 'Data Field': columnsData,
            'Data Field': colwithf,
            'Issue Description': lst,
            '#Mismatches': no_of_mismatch_data,
            'ME Review Comments': lst,
            'QA Updates': qa_updates
            }
    df4 = pd.DataFrame(data)
    df4.to_excel(writer, index=False, header=True, sheet_name='Review')
    writer.save()

    wb1 = openpyxl.load_workbook(file_name)
    sheet1 = wb1['Execution_Report']
    sheet2 = wb1['Review']
    wb1.move_sheet('Review', 3)
    sheet1.insert_rows(idx=1, amount=2)
    sheet1.insert_cols(idx=1, amount=1)
    sheet1.merge_cells(range_string='Execution_Report!B1:E1')
    sheet1.merge_cells(range_string='Execution_Report!A1:A50')
    sheet1.merge_cells(range_string='Execution_Report!F1:K50')
    sheet1.merge_cells(range_string='Execution_Report!A22:K50')
    sheet1.merge_cells(range_string='Execution_Report!B6:E6')
    sheet1.merge_cells(range_string='Execution_Report!B7:E21')
    sheet1.cell(row=1, column=2).value = str(client_name).upper() + ' CHARGEBEE REPORTS'
    sheet1.cell(row=1, column=2).alignment = Alignment(horizontal='center')
    sheet1.cell(row=7, column=2).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=1, column=2).font = Font(color="FFFFFF", bold='True')
    sheet1.cell(row=1, column=2).fill = PatternFill(fgColor='FFA500', fill_type='solid')
    for c in range(2, 6):
        sheet1.cell(row=4, column=c).alignment = Alignment(horizontal='center')
        sheet1.cell(row=3, column=c).fill = PatternFill(fgColor='1589FF', fill_type='solid')
        sheet1.cell(row=3, column=c).font = Font(color="FFFFFF", bold='True')
    sheet1.cell(row=4, column=4).font = Font(color="E41B17")
    sheet1.cell(row=6, column=2).value = total_deviations
    sheet1.cell(row=6, column=2).alignment = Alignment(horizontal='center')
    sheet1.cell(row=6, column=2).font = Font(color="17202A", bold='True')
    sheet1.cell(row=6, column=2).fill = PatternFill(fgColor='CACFD2', fill_type='solid')

    sheet1.column_dimensions['B'].width = 25
    sheet1.column_dimensions['C'].width = 25
    sheet1.column_dimensions['D'].width = 25
    sheet1.column_dimensions['E'].width = 25
    sheet2.column_dimensions['B'].width = 25
    sheet2.column_dimensions['C'].width = 25
    sheet2.column_dimensions['D'].width = 25
    sheet2.column_dimensions['E'].width = 25
    sheet2.column_dimensions['F'].width = 25
    pie = PieChart()
    labels = Reference(range_string='Execution_Report!C3:E3')
    data = Reference(range_string='Execution_Report!C4:E4')
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    # pie.title = "Subscriptions Execution Report"
    pie.title = module

    # Cut the first slice out of the pie
    slice_data = DataPoint(idx=0, explosion=10)
    pie.series[0].data_points = [slice_data]
    pie.anchor = 'B7'
    pie.height = 6.5  # default is 7.5
    pie.width = 12  # default is 15

    sheet1.add_chart(pie, "C8")

    sheet1.cell(1, 2).border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
                                      right=Side(border_style=BORDER_THIN, color='00000000'),
                                      top=Side(border_style=BORDER_THIN, color='00000000'),
                                      bottom=Side(border_style=BORDER_THIN, color='00000000'))

    sheet1.cell(1, 2).border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
                                      right=Side(border_style=BORDER_THIN, color='00000000'),
                                      top=Side(border_style=BORDER_THIN, color='00000000'),
                                      bottom=Side(border_style=BORDER_THIN, color='00000000'))

    sheet1.cell(4, 2).border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
                                      right=Side(border_style=BORDER_THIN, color='00000000'),
                                      top=Side(border_style=BORDER_THIN, color='00000000'),
                                      bottom=Side(border_style=BORDER_THIN, color='00000000'))

    sheet1.cell(4, 3).border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
                                      right=Side(border_style=BORDER_THIN, color='00000000'),
                                      top=Side(border_style=BORDER_THIN, color='00000000'),
                                      bottom=Side(border_style=BORDER_THIN, color='00000000'))

    sheet1.cell(4, 4).border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
                                      right=Side(border_style=BORDER_THIN, color='00000000'),
                                      top=Side(border_style=BORDER_THIN, color='00000000'),
                                      bottom=Side(border_style=BORDER_THIN, color='00000000'))

    sheet1.cell(4, 5).border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
                                      right=Side(border_style=BORDER_THIN, color='00000000'),
                                      top=Side(border_style=BORDER_THIN, color='00000000'),
                                      bottom=Side(border_style=BORDER_THIN, color='00000000'))
    for column in range(2, 6):
        sheet1.cell(6, column).border = Border(bottom=Side(border_style=BORDER_THIN, color='00000000'))
        sheet1.cell(21, column).border = Border(bottom=Side(border_style=BORDER_THIN, color='00000000'))
    for row in range(1, 22):
        sheet1.cell(row, 1).border = Border(right=Side(border_style=BORDER_THIN, color='00000000'))
        sheet1.cell(row, 6).border = Border(left=Side(border_style=BORDER_THIN, color='00000000'))

    wb1.save(file_name)

    print('!!! Execution Completed !!!')
    return True
